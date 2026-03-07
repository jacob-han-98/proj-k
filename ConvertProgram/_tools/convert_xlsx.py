"""
XLSX → AI-Readable Markdown 변환 파이프라인
===========================================
Tier 1: 시트 셀 데이터 (openpyxl) → 구조화 Markdown
Tier 1.5: 도형/플로우차트 (OOXML direct parsing) → 도형 Markdown

사용법:
  python convert_xlsx.py "경로/파일.xlsx"                     # 단일 파일
  python convert_xlsx.py "경로/파일.xlsx" --out "출력/폴더"    # 출력 폴더 지정
  python convert_xlsx.py --batch "폴더경로"                   # 폴더 내 전체 xlsx 일괄

의존성: openpyxl (pip install openpyxl)
"""

import openpyxl
import zipfile
import os
import sys
import json
import re
import argparse
import colorsys
from xml.etree import ElementTree as ET
from collections import defaultdict
from pathlib import Path


# ============================================================
# Tier 1: 시트 셀 데이터 → Markdown
# ============================================================

def build_merge_info(ws):
    """병합 셀 정보 구축: origin vs continuation 구분"""
    merge_origins = {}
    merge_skips = {}
    merge_spans = {}

    for mg in ws.merged_cells.ranges:
        merge_origins[(mg.min_row, mg.min_col)] = True
        merge_spans[(mg.min_row, mg.min_col)] = (mg.max_row, mg.max_col)
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    merge_skips[(r, c)] = (mg.min_row, mg.min_col)
    return merge_origins, merge_skips, merge_spans


def get_row_data(ws, r, merge_skips):
    """행에서 병합 continuation이 아닌 실제 데이터만 추출"""
    cells = []
    for c in range(1, ws.max_column + 1):
        if (r, c) in merge_skips:
            continue
        v = ws.cell(r, c).value
        if v is not None:
            cells.append((c, v))
    return cells


def clean(v):
    if v is None:
        return ''
    return str(v).strip()


# 강조로 간주할 글자색 (기본 검정/흰색이 아닌 눈에 띄는 색)
_EMPHASIS_COLORS = {
    'FF0000', 'CC0000', 'C00000', 'FF3300',  # 빨강 계열
    '0000FF', '0070C0',                        # 파랑 계열
    '00B050', '008000',                        # 초록 계열
    'FF00FF', '800080', '7030A0',              # 보라 계열
}


def get_font_color(cell, theme_colors=None):
    """셀의 글자색이 강조색이면 해당 색 이름을 반환. 아니면 None."""
    try:
        font = cell.font
        if not font or not font.color:
            return None

        color = font.color
        rgb = None

        # 1) 직접 RGB
        try:
            rgb_val = color.rgb
            if isinstance(rgb_val, str) and rgb_val != '00000000':
                rgb = rgb_val
                if len(rgb) == 8:
                    rgb = rgb[2:]
        except:
            pass

        # 2) 테마 색상
        if (not rgb or rgb == '000000') and theme_colors:
            try:
                theme_idx = color.theme
                if isinstance(theme_idx, int) and 0 <= theme_idx < len(theme_colors):
                    rgb = theme_colors[theme_idx]
                    tint = color.tint
                    if tint and tint != 0:
                        rgb = _apply_tint(rgb, tint)
            except:
                pass

        if not rgb:
            return None

        rgb = rgb.upper()
        if rgb in _EMPHASIS_COLORS:
            return rgb
        return None
    except:
        return None


# 잘 알려진 색상 이름 매핑 (RGB hex 대문자 → 한글명)
_COLOR_NAMES = {
    'FF0000': '빨강', 'FF3300': '빨강', 'CC0000': '진빨강', 'C00000': '진빨강',
    '00FF00': '초록', '008000': '초록', '92D050': '연두',
    '0000FF': '파랑', '0070C0': '파랑', '00B0F0': '하늘',
    'FFFF00': '노랑', 'FFC000': '주황',
    'FF00FF': '분홍', '800080': '보라', '7030A0': '보라',
    '00FFFF': '하늘', 'FFA500': '주황', 'FF8C00': '주황',
    'C0C0C0': '은색', '808080': '회색', 'A6A6A6': '회색',
    'BFBFBF': '밝은회색', 'D9D9D9': '밝은회색',
}


def _extract_theme_colors(wb):
    """워크북의 theme XML에서 테마 색상 팔레트를 추출"""
    try:
        # openpyxl 내부의 theme XML 접근
        theme_xml = wb.loaded_theme
        if not theme_xml:
            return []
        root = ET.fromstring(theme_xml)
        ns_a = '{http://schemas.openxmlformats.org/drawingml/2006/main}'

        # themeElements/clrScheme 에서 색상 추출
        colors = []
        clr_scheme = root.find(f'.//{ns_a}clrScheme')
        if clr_scheme is None:
            return []

        # 테마 색상 순서: dk1, lt1, dk2, lt2, accent1~6, hlink, folHlink
        color_tags = ['dk1', 'lt1', 'dk2', 'lt2',
                      'accent1', 'accent2', 'accent3', 'accent4',
                      'accent5', 'accent6', 'hlink', 'folHlink']
        for tag in color_tags:
            elem = clr_scheme.find(f'{ns_a}{tag}')
            if elem is not None:
                # srgbClr 또는 sysClr
                srgb = elem.find(f'{ns_a}srgbClr')
                if srgb is not None:
                    colors.append(srgb.attrib.get('val', '000000'))
                else:
                    sys_clr = elem.find(f'{ns_a}sysClr')
                    if sys_clr is not None:
                        colors.append(sys_clr.attrib.get('lastClr', '000000'))
                    else:
                        colors.append('000000')
            else:
                colors.append('000000')
        return colors
    except:
        return []


def _apply_tint(rgb_hex, tint):
    """테마 색상에 tint 적용 (밝기 조절)"""
    try:
        r = int(rgb_hex[0:2], 16) / 255.0
        g = int(rgb_hex[2:4], 16) / 255.0
        b = int(rgb_hex[4:6], 16) / 255.0
        if tint < 0:
            # 어둡게
            r = r * (1.0 + tint)
            g = g * (1.0 + tint)
            b = b * (1.0 + tint)
        else:
            # 밝게
            r = r * (1.0 - tint) + tint
            g = g * (1.0 - tint) + tint
            b = b * (1.0 - tint) + tint
        r = max(0, min(255, int(r * 255)))
        g = max(0, min(255, int(g * 255)))
        b = max(0, min(255, int(b * 255)))
        return f"{r:02X}{g:02X}{b:02X}"
    except:
        return rgb_hex


def get_cell_color(cell, theme_colors=None):
    """셀의 배경색(fill color)을 HEX 문자열로 반환. 없으면 None."""
    try:
        fill = cell.fill
        if not fill or not fill.fgColor:
            return None

        color = fill.fgColor
        rgb = None

        # 1) 직접 RGB(ARGB) 먼저 시도
        try:
            rgb_val = color.rgb
            if isinstance(rgb_val, str) and rgb_val != '00000000':
                rgb = rgb_val
                if len(rgb) == 8:
                    rgb = rgb[2:]
        except:
            pass

        # 2) 직접 RGB가 없으면 테마 색상 해석
        if (not rgb or rgb == '000000') and theme_colors:
            try:
                theme_idx = color.theme
                if isinstance(theme_idx, int) and 0 <= theme_idx < len(theme_colors):
                    rgb = theme_colors[theme_idx]
                    tint = color.tint
                    if tint and tint != 0:
                        rgb = _apply_tint(rgb, tint)
            except:
                pass

        if not rgb or rgb == '000000' or rgb.upper() == 'FFFFFF':
            return None

        rgb = rgb.upper()
        name = _COLOR_NAMES.get(rgb, '')
        if name:
            return f"[{name} #{rgb}]"
        return f"[#{rgb}]"
    except:
        return None


def _build_image_map(ws):
    """시트에서 이미지가 위치한 행 범위를 매핑

    Returns:
        list of (start_row, end_row, start_col, end_col) — 1-based rows/cols
    """
    image_regions = []
    for img in ws._images:
        anchor = img.anchor
        try:
            if hasattr(anchor, '_from') and anchor._from is not None:
                fr = anchor._from
                start_row = fr.row + 1  # 0-based → 1-based
                start_col = fr.col + 1
                if hasattr(anchor, 'to') and anchor.to is not None:
                    end_row = anchor.to.row + 1
                    end_col = anchor.to.col + 1
                else:
                    end_row = start_row + 5
                    end_col = start_col + 5
                image_regions.append((start_row, end_row, start_col, end_col))
        except Exception:
            pass
    return image_regions


def _find_image_at_row(image_regions, row, emitted_markers):
    """주어진 행에 걸치는 이미지 영역이 있으면 반환 (이미 마커를 삽입한 영역은 제외)"""
    for region in image_regions:
        sr, er, sc, ec = region
        if sr <= row <= er and region not in emitted_markers:
            return region
    return None


def _find_column_clusters(header_cells, merge_spans=None, header_row=0, max_gap=1):
    """헤더 셀을 근접 클러스터로 분리. 병합 스팬을 고려하여 실효 열 끝 위치 사용.
    max_gap=1 → 실효 열 끝과 다음 열 시작의 차이가 1 초과이면 별도 클러스터.
    (예: col 14→16은 차이 2 > 1 → 분리, 단 col 6[span→7]→8은 차이 1 → 동일)"""
    if not header_cells:
        return []
    sorted_cells = sorted(header_cells, key=lambda x: x[0])

    # 각 셀의 실효 끝 열 계산 (merge span 포함)
    effective = []
    for c, text in sorted_cells:
        end_c = c
        if merge_spans and (header_row, c) in merge_spans:
            end_c = merge_spans[(header_row, c)][1]
        effective.append((c, end_c))

    clusters = [[effective[0][0]]]
    prev_end = effective[0][1]
    for start_c, end_c in effective[1:]:
        if start_c - prev_end > max_gap:
            clusters.append([start_c])
        else:
            clusters[-1].append(start_c)
        prev_end = max(prev_end, end_c)
    return clusters


def _match_headers_to_clusters(group_cells, clusters):
    """그룹 헤더 셀(col, text)을 컬럼 클러스터에 매칭 → {cluster_idx: label}"""
    matched = {}
    for gc, gv in group_cells:
        best_ci = -1
        best_dist = float('inf')
        for ci, cluster in enumerate(clusters):
            min_c, max_c = min(cluster), max(cluster)
            if min_c - 2 <= gc <= max_c:
                dist = abs(gc - min_c)
                if dist < best_dist:
                    best_dist = dist
                    best_ci = ci
        if best_ci >= 0:
            matched[best_ci] = clean(gv)
    return matched


def _extract_single_table_data(ws, start_row, header_cols, merge_skips, theme_colors):
    """단일 테이블의 데이터 행들을 추출 (data_rows, end_row 반환)"""
    data_rows = []
    dr = start_row + 1
    min_header_col = min(header_cols)
    while dr <= ws.max_row:
        row_cells = get_row_data(ws, dr, merge_skips)
        if row_cells:
            first_col, first_val = row_cells[0]
            first_str = clean(first_val)
            if first_col < min_header_col and (
                first_str.startswith('▶') or
                re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩]', first_str) or
                re.match(r'^\(\d+\)', first_str)):
                break

        vals = []
        any_val = False
        has_own_data = False
        for hc in header_cols:
            if (dr, hc) in merge_skips:
                origin_row, origin_col = merge_skips[(dr, hc)]
                origin_val = ws.cell(origin_row, origin_col).value
                if origin_val is not None:
                    vals.append(clean(origin_val))
                    any_val = True
                else:
                    vals.append('')
            else:
                cell = ws.cell(dr, hc)
                v = cell.value
                if v is not None:
                    text = clean(v)
                    if text and get_font_color(cell, theme_colors):
                        text = f"**{text}**"
                    vals.append(text)
                    any_val = True
                    has_own_data = True
                else:
                    color = get_cell_color(cell, theme_colors)
                    if color:
                        vals.append(color)
                        any_val = True
                        has_own_data = True
                    else:
                        vals.append('')

        if not any_val:
            all_skips = all((dr, hc) in merge_skips for hc in header_cols)
            if all_skips:
                dr += 1
                continue
            break
        if not has_own_data:
            dr += 1
            continue
        data_rows.append(vals)
        dr += 1
    return data_rows, dr


def _extract_parallel_tables(ws, start_row, headers, clusters, group_cells,
                              merge_skips, theme_colors, original_start):
    """병렬 테이블 추출 — 가로 배치된 별도 테이블들을 분리 출력"""
    group_map = _match_headers_to_clusters(group_cells, clusters)

    # 각 클러스터의 서브헤더
    cluster_headers = []
    for cluster in clusters:
        sub = [(c, n) for c, n in headers if c in cluster]
        cluster_headers.append(sub)

    # 데이터 행 읽기 (모든 클러스터에서 동시에)
    all_header_cols = [c for c, _ in headers]
    data_rows_raw, end_row = _extract_single_table_data(
        ws, start_row, all_header_cols, merge_skips, theme_colors)

    # 각 클러스터별 데이터 분리
    col_to_cluster = {}
    for ci, cluster in enumerate(clusters):
        for c in cluster:
            col_to_cluster[c] = ci

    cluster_data = [[] for _ in clusters]
    for row_vals in data_rows_raw:
        per_cluster = [[] for _ in clusters]
        for i, hc in enumerate(all_header_cols):
            ci = col_to_cluster.get(hc, 0)
            per_cluster[ci].append(row_vals[i])
        for ci in range(len(clusters)):
            if any(v for v in per_cluster[ci]):
                cluster_data[ci].append(per_cluster[ci])

    # Markdown 출력
    md = ""

    # 병렬 관계 표시 (그룹 헤더에서 * 마커 제거)
    labels = []
    for ci in range(len(clusters)):
        raw = group_map.get(ci, f"테이블 {ci+1}")
        labels.append(raw.lstrip('* ').strip())

    if len(clusters) == 2:
        md += f"\n> **병렬 비교**: {labels[0]} ↔ {labels[1]}\n"
    elif len(clusters) >= 3:
        md += f"\n> **병렬 비교 테이블 세트**: {' | '.join(labels)}\n"

    for ci, cluster in enumerate(clusters):
        raw = group_map.get(ci, "")
        label = raw.lstrip('* ').strip() if raw else ""
        if label:
            md += f"\n**{label}**\n"

        sub_headers = [n for _, n in cluster_headers[ci]]
        md += f"\n| {' | '.join(sub_headers)} |\n"
        md += f"| {' | '.join(['---'] * len(sub_headers))} |\n"
        for vals in cluster_data[ci]:
            cleaned = [v.replace('\n', ' / ').replace('|', '∣')[:200] for v in vals]
            md += f"| {' | '.join(cleaned)} |\n"

    return md, end_row


def detect_and_extract_table(ws, start_row, merge_skips, merge_origins, merge_spans, theme_colors=None):
    """테이블 자동 감지 및 Markdown 테이블로 변환"""
    header_cells = get_row_data(ws, start_row, merge_skips)
    if len(header_cells) < 3:
        return None, start_row

    headers = [(c, clean(v)) for c, v in header_cells]
    header_cols = [c for c, _ in headers]
    header_names = [n for _, n in headers]

    # --- 그룹 헤더 감지 ---
    group_header_text = None
    group_cells = []  # (col, text) 원본 그룹 헤더
    original_start = start_row

    has_gap = any(header_cols[i+1] - header_cols[i] > 2
                  for i in range(len(header_cols) - 1))
    has_wide_merge = any(
        (start_row, c) in merge_spans and merge_spans[(start_row, c)][1] > c
        for c, _ in header_cells
    )

    if (has_gap or has_wide_merge) and start_row < ws.max_row:
        next_cells = get_row_data(ws, start_row + 1, merge_skips)
        if next_cells:
            min_hc = header_cols[0]
            max_hc = header_cols[-1]
            for c, _ in header_cells:
                if (start_row, c) in merge_spans:
                    max_hc = max(max_hc, merge_spans[(start_row, c)][1])
            next_in_range = [(c, v) for c, v in next_cells
                            if min_hc <= c <= max_hc + 6]
            if len(next_in_range) > len(header_cells):
                group_cells = headers[:]  # 원본 그룹 헤더 보존
                group_header_text = '\n'.join(f"* {n}" for _, n in headers)
                headers = [(c, clean(v)) for c, v in next_in_range]
                header_cols = [c for c, _ in headers]
                header_names = [n for _, n in headers]
                start_row += 1

    # --- 병렬 테이블 감지 ---
    # 조건: 2+ 클러스터, 각 클러스터 3+ 열, 클러스터 간 열 수 유사 (차이 ≤ 1)
    clusters = _find_column_clusters(headers, merge_spans, start_row)
    cluster_sizes = [len(cl) for cl in clusters]
    is_parallel = (len(clusters) > 1
                   and all(sz >= 3 for sz in cluster_sizes)
                   and max(cluster_sizes) - min(cluster_sizes) <= 1)
    if is_parallel:
        # 그룹 헤더가 없으면 이전 행에서 찾기 (look-back)
        if not group_cells and original_start > 1:
            prev_cells = get_row_data(ws, original_start - 1, merge_skips)
            if prev_cells and 1 <= len(prev_cells) <= len(clusters) + 1:
                test_match = _match_headers_to_clusters(
                    [(c, clean(v)) for c, v in prev_cells], clusters)
                if len(test_match) >= len(clusters) - 1:
                    group_cells = [(c, clean(v)) for c, v in prev_cells]
        return _extract_parallel_tables(
            ws, start_row, headers, clusters, group_cells,
            merge_skips, theme_colors, original_start)

    data_rows = []
    dr = start_row + 1
    min_header_col = min(header_cols)
    while dr <= ws.max_row:
        row_cells = get_row_data(ws, dr, merge_skips)

        if row_cells:
            first_col, first_val = row_cells[0]
            first_str = clean(first_val)
            # 섹션 마커 감지: 헤더 시작열보다 왼쪽에 있는 마커만 테이블 종료
            if first_col < min_header_col and (
                first_str.startswith('▶') or
                re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩]', first_str) or
                re.match(r'^\(\d+\)', first_str)):
                break

        vals = []
        any_val = False
        has_own_data = False  # 이 행에 고유 데이터(비병합)가 있는지
        for hc in header_cols:
            if (dr, hc) in merge_skips:
                origin_row, origin_col = merge_skips[(dr, hc)]
                origin_val = ws.cell(origin_row, origin_col).value
                if origin_val is not None:
                    vals.append(clean(origin_val))
                    any_val = True
                else:
                    vals.append('')
            else:
                cell = ws.cell(dr, hc)
                v = cell.value
                if v is not None:
                    text = clean(v)
                    # 글자색이 강조색이면 bold 처리
                    if text and get_font_color(cell, theme_colors):
                        text = f"**{text}**"
                    vals.append(text)
                    any_val = True
                    has_own_data = True
                else:
                    # 값이 없으면 셀 배경색 확인
                    color = get_cell_color(cell, theme_colors)
                    if color:
                        vals.append(color)
                        any_val = True
                        has_own_data = True
                    else:
                        vals.append('')

        if not any_val:
            all_skips = all((dr, hc) in merge_skips for hc in header_cols)
            if all_skips:
                dr += 1
                continue
            break

        # 모든 셀이 병합 continuation이면 중복행 → 건너뛰기
        if not has_own_data:
            dr += 1
            continue

        data_rows.append(vals)
        dr += 1

    if not data_rows:
        return None, original_start

    md = ""
    if group_header_text:
        md += f"\n{group_header_text}\n"
    md += f"\n| {' | '.join(header_names)} |\n"
    md += f"| {' | '.join(['---'] * len(header_names))} |\n"
    for vals in data_rows:
        cleaned = [v.replace('\n', ' / ').replace('|', '∣')[:200] for v in vals]
        md += f"| {' | '.join(cleaned)} |\n"

    return md, dr


def sheet_to_markdown(ws, sheet_name, source_filename, theme_colors=None):
    """단일 시트 → Markdown 변환"""
    merge_origins, merge_skips, merge_spans = build_merge_info(ws)

    # 이미지 위치 감지
    image_regions = _build_image_map(ws)
    emitted_markers = set()

    md = f"# {sheet_name}\n\n"
    md += f"> 원본: {source_filename} / {sheet_name}\n"
    md += f"> 행: {ws.max_row}, 열: {ws.max_column}, 병합셀: {len(ws.merged_cells.ranges)}, 이미지: {len(ws._images)}\n\n"

    r = 1
    while r <= ws.max_row:
        row_cells = get_row_data(ws, r, merge_skips)

        # 이미지 마커 삽입: 현재 행에 이미지가 걸쳐 있으면 마커 출력
        if image_regions:
            img_region = _find_image_at_row(image_regions, r, emitted_markers)
            if img_region:
                sr, er, sc, ec = img_region
                emitted_markers.add(img_region)
                md += f"\n> [IMAGE: 행 {sr}-{er}, 열 {sc}-{ec} 영역에 시각 자료 포함 — 스크린샷 참조 필요]\n\n"

        if not row_cells:
            r += 1
            continue

        if len(row_cells) >= 3:
            table_md, end_row = detect_and_extract_table(ws, r, merge_skips, merge_origins, merge_spans, theme_colors)
            if table_md:
                md += table_md + "\n"
                # 테이블 범위 내 건너뛴 이미지 마커도 삽입
                if image_regions:
                    for skip_r in range(r + 1, end_row):
                        img_region = _find_image_at_row(image_regions, skip_r, emitted_markers)
                        if img_region:
                            sr, er, sc, ec = img_region
                            emitted_markers.add(img_region)
                            md += f"\n> [IMAGE: 행 {sr}-{er}, 열 {sc}-{ec} 영역에 시각 자료 포함 — 스크린샷 참조 필요]\n\n"
                r = end_row
                continue

        for col, v in row_cells:
            text = clean(v)
            if not text or text == '-':
                continue

            # 셀 전체 글자색이 강조색이면 bold 래핑
            is_emphasis = bool(get_font_color(ws.cell(r, col), theme_colors))

            lines = text.split('\n')
            first = lines[0].strip()
            rest = [l.strip() for l in lines[1:] if l.strip()]

            if is_emphasis:
                first = f"**{first}**"

            if col <= 2 and first.lstrip('*').startswith('▶'):
                plain = first.replace('▶', '').replace('*', '').strip()
                md += f"\n# {plain}\n\n"
            elif col <= 2 and re.match(r'^\*{0,2}[①②③④⑤⑥⑦⑧⑨⑩]', first):
                md += f"\n## {first}\n\n"
            elif col <= 3 and re.match(r'^\*{0,2}\(\d+\)', first):
                md += f"### {first}\n"
            elif first.lstrip('*').startswith('→'):
                md += f"- {first}\n"
            elif 'ü' in first:
                md += f"  - {first}\n"
            elif col <= 3 and first.lstrip('*').startswith('*'):
                md += f"\n{first}\n"
            elif first.lstrip('*').startswith('"') or first.lstrip('*').startswith('\u201c'):
                md += f"  > {first}\n"
            else:
                md += f"{first}\n"

            for ln in rest:
                if is_emphasis:
                    ln = f"**{ln}**"
                if ln.lstrip('*').startswith('→'):
                    md += f"- {ln}\n"
                elif 'ü' in ln:
                    md += f"  - {ln}\n"
                elif ln.lstrip('*').startswith('-'):
                    md += f"  {ln}\n"
                else:
                    md += f"  {ln}\n"

        r += 1

    return md


# ============================================================
# Tier 1.5: OOXML 도형/플로우차트 → Markdown
# ============================================================

NS_XDR = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
NS_A = '{http://schemas.openxmlformats.org/drawingml/2006/main}'

SHAPE_TYPE_MAP = {
    'roundRect': '□', 'rect': '■', 'ellipse': '○',
    'flowChartDecision': '◇', 'diamond': '◇',
    'flowChartProcess': '□', 'flowChartTerminator': '◉',
}

# OOXML geo type → YAML node type 매핑
GEO_TO_NODE_TYPE = {
    'ellipse': 'terminal',
    'flowChartTerminator': 'terminal',
    'flowChartDecision': 'decision',
    'diamond': 'decision',
    'roundRect': 'process',
    'rect': 'annotation',
    'flowChartProcess': 'process',
}

def _sanitize_mermaid_id(sp_id, text):
    """Mermaid에서 안전한 노드 ID 생성"""
    # 짧은 텍스트가 있으면 그걸 기반으로, 없으면 sp_id 사용
    base = text if text else sp_id
    # 영문/숫자/한글만 남기고 _로 치환, 최대 20자
    safe = re.sub(r'[^\w가-힣]', '_', base).strip('_')[:20]
    if not safe or safe[0].isdigit():
        safe = f"n_{safe}"
    return safe


def extract_text_from_element(elem):
    texts = []
    for t in elem.iter(f'{NS_A}t'):
        if t.text:
            texts.append(t.text)
    return ''.join(texts).strip().replace('\n', ' ')


def _get_anchor_bounds(anchor):
    """앵커에서 도형의 from/to 좌표(col, row, colOff, rowOff)를 EMU 기반 중심점으로 반환"""
    fr = anchor.find(f'{NS_XDR}from')
    to = anchor.find(f'{NS_XDR}to')
    if fr is None or to is None:
        return None
    try:
        fc = int(fr.find(f'{NS_XDR}col').text)
        fr_ = int(fr.find(f'{NS_XDR}row').text)
        fco = int(fr.find(f'{NS_XDR}colOff').text)
        fro = int(fr.find(f'{NS_XDR}rowOff').text)
        tc = int(to.find(f'{NS_XDR}col').text)
        tr = int(to.find(f'{NS_XDR}row').text)
        tco = int(to.find(f'{NS_XDR}colOff').text)
        tro = int(to.find(f'{NS_XDR}rowOff').text)
        # 대략적 중심점 (col*default_width + offset)
        cx = (fc * 914400 + fco + tc * 914400 + tco) / 2
        cy = (fr_ * 190500 + fro + tr * 190500 + tro) / 2
        return (cx, cy)
    except:
        return None


def _process_shape(sp, shapes, row, anchor_bounds=None):
    sp_id = sp_name = ''
    for cNvPr in sp.iter(f'{NS_XDR}cNvPr'):
        sp_id = cNvPr.attrib.get('id', '')
        sp_name = cNvPr.attrib.get('name', '')
        break
    shape_type = 'unknown'
    for prstGeom in sp.iter(f'{NS_A}prstGeom'):
        shape_type = prstGeom.attrib.get('prst', 'unknown')
    text = extract_text_from_element(sp)
    symbol = SHAPE_TYPE_MAP.get(shape_type, '□')
    shapes[sp_id] = {
        'name': sp_name, 'text': text, 'type': shape_type,
        'symbol': symbol, 'row': row, 'pos': anchor_bounds
    }


def _process_connector(cxn, connectors, row, anchor_bounds=None):
    cxn_name = start_id = end_id = ''
    for cNvPr in cxn.iter(f'{NS_XDR}cNvPr'):
        cxn_name = cNvPr.attrib.get('name', '')
        break
    for cNvCxnSpPr in cxn.iter(f'{NS_XDR}cNvCxnSpPr'):
        stCxn = cNvCxnSpPr.find(f'{NS_A}stCxn')
        endCxn = cNvCxnSpPr.find(f'{NS_A}endCxn')
        if stCxn is not None:
            start_id = stCxn.attrib.get('id', '')
        if endCxn is not None:
            end_id = endCxn.attrib.get('id', '')
    connectors.append({'name': cxn_name, 'from': start_id, 'to': end_id, 'row': row, 'pos': anchor_bounds})


_DIAGRAM_GEOS = {'ellipse', 'flowChartTerminator', 'flowChartDecision', 'diamond',
                  'roundRect', 'flowChartProcess'}


def _process_group(grpSp, shapes, connectors, row, anchor_bounds=None):
    # 그룹 내 도형 처리 + ID 추적
    group_ids = []
    for sp in grpSp.iter(f'{NS_XDR}sp'):
        _process_shape(sp, shapes, row, anchor_bounds)
        for cNvPr in sp.iter(f'{NS_XDR}cNvPr'):
            group_ids.append(cNvPr.attrib.get('id', ''))
            break

    # 그룹 내 (다이어그램 도형 + TextBox 라벨) 쌍 병합
    # 패턴: diamond/ellipse(텍스트 없음) + rect/TextBox(텍스트 있음) → 텍스트를 도형에 복사
    textless = [(sid, shapes[sid]) for sid in group_ids
                if sid in shapes and not shapes[sid]['text']
                and shapes[sid]['type'] in _DIAGRAM_GEOS]
    text_labels = [(sid, shapes[sid]) for sid in group_ids
                   if sid in shapes and shapes[sid]['text']
                   and shapes[sid]['type'] == 'rect']

    if len(textless) == 1 and len(text_labels) >= 1:
        tl_id, tl_shape = textless[0]
        tr_id, tr_shape = text_labels[0]
        tl_shape['text'] = tr_shape['text']
        tl_shape['symbol'] = SHAPE_TYPE_MAP.get(tl_shape['type'], '□')
        del shapes[tr_id]  # TextBox는 라벨이므로 제거

    for cxn in grpSp.iter(f'{NS_XDR}cxnSp'):
        _process_connector(cxn, connectors, row, anchor_bounds)


def extract_shapes_from_container(container, row_offset=0):
    shapes = {}
    connectors = []
    images = []  # embedded picture info
    for anchor in container:
        atag = anchor.tag.replace(NS_XDR, '')
        if atag not in ('twoCellAnchor', 'oneCellAnchor'):
            if atag == 'sp':
                _process_shape(anchor, shapes, row_offset)
            elif atag == 'cxnSp':
                _process_connector(anchor, connectors, row_offset)
            elif atag == 'grpSp':
                _process_group(anchor, shapes, connectors, row_offset)
            continue

        from_elem = anchor.find(f'{NS_XDR}from')
        to_elem = anchor.find(f'{NS_XDR}to')
        row = row_offset
        if from_elem is not None:
            r = from_elem.find(f'{NS_XDR}row')
            if r is not None:
                row = int(r.text)

        bounds = _get_anchor_bounds(anchor)

        # 이미지(pic) 감지
        for pic in anchor.findall(f'{NS_XDR}pic'):
            pic_name = ''
            for cNvPr in pic.iter(f'{NS_XDR}cNvPr'):
                pic_name = cNvPr.attrib.get('name', '')
                break
            fr_row = fr_col = to_row = to_col = 0
            if from_elem is not None:
                fr_row = int(from_elem.find(f'{NS_XDR}row').text) + 1
                fr_col = int(from_elem.find(f'{NS_XDR}col').text) + 1
            if to_elem is not None:
                to_row = int(to_elem.find(f'{NS_XDR}row').text) + 1
                to_col = int(to_elem.find(f'{NS_XDR}col').text) + 1
            images.append({
                'name': pic_name,
                'from_row': fr_row, 'from_col': fr_col,
                'to_row': to_row or fr_row + 5, 'to_col': to_col or fr_col + 5,
            })

        for sp in anchor.findall(f'{NS_XDR}sp'):
            _process_shape(sp, shapes, row, bounds)
        for cxn in anchor.findall(f'{NS_XDR}cxnSp'):
            _process_connector(cxn, connectors, row, bounds)
        for grpSp in anchor.findall(f'{NS_XDR}grpSp'):
            _process_group(grpSp, shapes, connectors, row, bounds)

    return shapes, connectors, images


def get_sheet_drawing_map(z):
    mapping = {}
    for f in sorted(z.namelist()):
        if f.startswith('xl/worksheets/_rels/') and f.endswith('.xml.rels'):
            content = z.read(f).decode('utf-8')
            root = ET.fromstring(content)
            sheet_file = f.replace('_rels/', '').replace('.rels', '')
            for rel in root:
                target = rel.attrib.get('Target', '')
                if 'drawing' in target and target.endswith('.xml'):
                    drawing_file = target.replace('..', 'xl')
                    if not drawing_file.startswith('xl/'):
                        drawing_file = 'xl/' + drawing_file.lstrip('/')
                    drawing_file = drawing_file.replace('//', '/')
                    if drawing_file.startswith('xl/xl/'):
                        drawing_file = drawing_file[3:]
                    mapping[sheet_file] = drawing_file
    return mapping


def get_sheet_names_from_zip(z):
    wb_xml = z.read('xl/workbook.xml').decode('utf-8')
    root = ET.fromstring(wb_xml)
    ns_wb = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    names = []
    for sheet in root.iter(f'{ns_wb}sheet'):
        names.append(sheet.attrib.get('name', ''))
    return names


def _find_nearest_shape(pos, shapes, exclude_id=''):
    """위치 기반으로 가장 가까운 텍스트 도형을 찾는다"""
    if not pos:
        return None
    cx, cy = pos
    best = None
    best_dist = float('inf')
    for sid, s in shapes.items():
        if sid == exclude_id or not s.get('text') or not s.get('pos'):
            continue
        sx, sy = s['pos']
        dist = ((cx - sx) ** 2 + (cy - sy) ** 2) ** 0.5
        if dist < best_dist:
            best_dist = dist
            best = s
    # 너무 멀면 매칭 안함 (임계값: 대략 10열 * 5행 범위)
    if best and best_dist < 10000000:
        return best
    return None


def shapes_to_yaml_data(shapes, connectors):
    """도형/커넥터 → 구조화된 YAML 데이터 (dict) 변환

    Returns:
        dict with 'nodes' and 'edges' lists, or None if no text shapes
    """
    text_shapes = {k: v for k, v in shapes.items() if v['text']}
    if not text_shapes:
        return None

    # --- 노드 생성 ---
    # ID 중복 방지를 위한 카운터
    id_counter = defaultdict(int)
    id_map = {}  # sp_id → yaml_id

    nodes = []
    for sp_id, s in sorted(text_shapes.items(), key=lambda x: x[1]['row']):
        base_id = _sanitize_mermaid_id(sp_id, s['text'])
        id_counter[base_id] += 1
        if id_counter[base_id] > 1:
            yaml_id = f"{base_id}_{id_counter[base_id]}"
        else:
            yaml_id = base_id
        id_map[sp_id] = yaml_id

        node_type = GEO_TO_NODE_TYPE.get(s.get('type', ''), 'process')
        node = {
            'id': yaml_id,
            'type': node_type,
            'label': s['text'],
        }
        if s.get('type'):
            node['geo'] = s['type']
        nodes.append(node)

    # --- 엣지 생성 (커넥터 해석) ---
    edges = []
    for cxn in sorted(connectors, key=lambda x: x['row']):
        f = shapes.get(cxn['from'], {})
        t = shapes.get(cxn['to'], {})
        ft = f.get('text', '')
        tt = t.get('text', '')

        # 미매칭 시 위치 기반 근접 매칭 시도
        from_id = cxn.get('from', '')
        to_id = cxn.get('to', '')
        if not ft and cxn.get('pos'):
            nearest = _find_nearest_shape(cxn['pos'], shapes, to_id)
            if nearest:
                # nearest에서 sp_id를 찾아야 함
                for sid, sv in shapes.items():
                    if sv is nearest:
                        from_id = sid
                        ft = nearest.get('text', '')
                        break
        if not tt and cxn.get('pos'):
            nearest = _find_nearest_shape(cxn['pos'], shapes, from_id)
            if nearest:
                for sid, sv in shapes.items():
                    if sv is nearest:
                        to_id = sid
                        tt = nearest.get('text', '')
                        break

        from_yaml_id = id_map.get(from_id, '')
        to_yaml_id = id_map.get(to_id, '')

        if from_yaml_id and to_yaml_id:
            edge = {'from': from_yaml_id, 'to': to_yaml_id}
            # 연결선 자체에 텍스트가 있으면 condition으로
            cxn_text = ''
            for sp_id_chk, s_chk in text_shapes.items():
                if s_chk['text'].upper() in ('YES', 'NO', 'Y', 'N', 'TRUE', 'FALSE'):
                    # 분기 라벨 도형이 근처에 있는지 확인
                    pass
            edges.append(edge)
        elif from_yaml_id:
            edges.append({'from': from_yaml_id, 'to': '?', '_orphan': True})
        elif to_yaml_id:
            edges.append({'from': '?', 'to': to_yaml_id, '_orphan': True})

    return {'nodes': nodes, 'edges': edges}


def yaml_data_to_mermaid(data, direction='LR'):
    """YAML 데이터 → Mermaid flowchart 문자열 변환

    Args:
        data: shapes_to_yaml_data()의 반환값
        direction: LR (좌→우) or TD (위→아래)
    """
    if not data or not data.get('nodes'):
        return ""

    lines = [f"flowchart {direction}"]
    node_ids = set()

    # 노드 정의
    for node in data['nodes']:
        nid = node['id']
        label = node['label'].replace('"', "'")
        ntype = node.get('type', 'process')

        if nid in node_ids:
            continue
        node_ids.add(nid)

        if ntype == 'terminal':
            lines.append(f'  {nid}(["{label}"])')
        elif ntype == 'decision':
            lines.append(f'  {nid}{{"{label}"}}')
        elif ntype == 'annotation':
            lines.append(f'  {nid}["{label}"]')
        else:  # process
            lines.append(f'  {nid}["{label}"]')

    lines.append("")

    # 엣지 정의
    for edge in data['edges']:
        f = edge['from']
        t = edge['to']
        if f == '?' or t == '?':
            continue
        cond = edge.get('condition', '')
        if cond:
            lines.append(f'  {f} -->|{cond}| {t}')
        else:
            lines.append(f'  {f} --> {t}')

    return '\n'.join(lines)


def yaml_data_to_markdown(data, shapes, connectors):
    """YAML 데이터 + 원본 shapes → 하이브리드 Markdown (Mermaid + 원본 텍스트 리스트)"""
    if not data:
        return ""

    md = ""
    text_shapes = {k: v for k, v in shapes.items() if v['text']}

    # --- 섹션 1: YAML 구조 (AI 파싱용) ---
    md += "### 도형 구조 (YAML)\n\n"
    md += "```yaml\n"
    md += "nodes:\n"
    for node in data['nodes']:
        md += f"  - id: {node['id']}\n"
        md += f"    type: {node['type']}\n"
        md += f"    label: \"{node['label']}\"\n"
    md += "\nedges:\n"
    for edge in data['edges']:
        f = edge['from']
        t = edge['to']
        line = f"  - from: {f}\n    to: {t}\n"
        if edge.get('condition'):
            line = f"  - from: {f}\n    to: {t}\n    condition: \"{edge['condition']}\"\n"
        if edge.get('_orphan'):
            line = f"  - from: {f}\n    to: {t}  # orphan connector\n"
        md += line
    md += "```\n\n"

    # --- 섹션 2: Mermaid (렌더링용) ---
    mermaid = yaml_data_to_mermaid(data)
    if mermaid:
        md += "### 플로우차트 (Mermaid)\n\n"
        md += f"```mermaid\n{mermaid}\n```\n\n"

    # --- 섹션 3: 원본 도형 텍스트 리스트 (참조용) ---
    md += f"### 원본 도형 텍스트 ({len(text_shapes)}개)\n\n"
    by_row = defaultdict(list)
    for sid, s in text_shapes.items():
        by_row[s['row'] // 10 * 10].append(s)

    for row_group in sorted(by_row.keys()):
        for s in by_row[row_group]:
            md += f"- {s['symbol']} {s['text']}\n"

    # --- 섹션 4: 미해석 연결 (orphan) ---
    orphans = [e for e in data['edges'] if e.get('_orphan')]
    if orphans:
        md += f"\n### 미해석 연결 ({len(orphans)}개)\n\n"
        for e in orphans:
            md += f"- {e['from']} → {e['to']}\n"

    return md


# 하위 호환: 기존 함수명 유지
def flowchart_to_markdown(shapes, connectors):
    """레거시 호환 래퍼 — 새 YAML+Mermaid 포맷으로 출력"""
    data = shapes_to_yaml_data(shapes, connectors)
    if not data:
        return ""
    return yaml_data_to_markdown(data, shapes, connectors)


def extract_shapes_from_xlsx(xlsx_path, out_dir, source_filename):
    """Tier 1.5: XLSX 내 도형 추출 → YAML (AI용) + Markdown/Mermaid (렌더링용)"""
    results = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        sheet_names = get_sheet_names_from_zip(z)
        sheet_drawing_map = get_sheet_drawing_map(z)

        for sheet_file, drawing_file in sheet_drawing_map.items():
            sheet_num = sheet_file.split('sheet')[-1].replace('.xml', '')
            try:
                sheet_idx = int(sheet_num) - 1
                sheet_name = sheet_names[sheet_idx] if sheet_idx < len(sheet_names) else f"Sheet{sheet_num}"
            except:
                sheet_name = f"Sheet{sheet_num}"

            norm_path = drawing_file
            if not norm_path.startswith('xl/'):
                norm_path = 'xl/' + norm_path
            norm_path = norm_path.replace('//', '/')

            try:
                xml = z.read(norm_path).decode('utf-8')
            except KeyError:
                alt = norm_path.replace('xl/xl/', 'xl/')
                try:
                    xml = z.read(alt).decode('utf-8')
                except KeyError:
                    continue

            root = ET.fromstring(xml)
            shapes, connectors, images = extract_shapes_from_container(root)
            text_shapes = {k: v for k, v in shapes.items() if v['text']}

            if not text_shapes:
                continue

            # YAML 구조 데이터 생성
            yaml_data = shapes_to_yaml_data(shapes, connectors)

            # --- Markdown (Mermaid 포함) ---
            md = f"# {sheet_name} — 도형/플로우차트 추출\n\n"
            md += f"> 원본: {source_filename} / {sheet_name}\n"
            md += f"> 도형: {len(shapes)}개 (텍스트 포함: {len(text_shapes)}), 연결선: {len(connectors)}개, 이미지: {len(images)}개\n\n"
            if yaml_data:
                md += yaml_data_to_markdown(yaml_data, shapes, connectors)

            # 이미지 위치 정보
            if images:
                md += f"\n### 내장 이미지 ({len(images)}개)\n\n"
                for img_info in images:
                    md += (f"- {img_info['name']}: "
                           f"행 {img_info['from_row']}-{img_info['to_row']}, "
                           f"열 {img_info['from_col']}-{img_info['to_col']} "
                           f"[IMAGE: 스크린샷 참조 필요]\n")

            text_connectors = [c for c in connectors if c['from'] and c['to']]
            results[sheet_name] = {
                'shapes': len(text_shapes),
                'connectors': len(text_connectors),
                'images': len(images),
                'chars': len(md)
            }

            safe_name = sheet_name.replace('/', '_').replace(' ', '_')

            # Markdown 저장
            filepath_md = os.path.join(out_dir, f"{safe_name}_도형.md")
            with open(filepath_md, 'w', encoding='utf-8') as f:
                f.write(md)

            # YAML 저장 (AI 파싱용)
            if yaml_data:
                yaml_out = {
                    'source': source_filename,
                    'sheet': sheet_name,
                    'stats': {
                        'shapes': len(text_shapes),
                        'connectors': len(text_connectors),
                    },
                    'flowchart': yaml_data,
                }
                filepath_yaml = os.path.join(out_dir, f"{safe_name}_도형.yaml")
                with open(filepath_yaml, 'w', encoding='utf-8') as f:
                    # PyYAML 없이 수동 YAML 직렬화 (의존성 추가 안함)
                    _write_yaml(f, yaml_out)

    return results


def _write_yaml(f, data, indent=0):
    """PyYAML 없이 간단한 YAML 직렬화 (의존성 최소화)"""
    prefix = '  ' * indent
    if isinstance(data, dict):
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                f.write(f"{prefix}{k}:\n")
                _write_yaml(f, v, indent + 1)
            elif isinstance(v, bool):
                f.write(f"{prefix}{k}: {'true' if v else 'false'}\n")
            elif isinstance(v, (int, float)):
                f.write(f"{prefix}{k}: {v}\n")
            else:
                # 문자열: 특수문자 포함 시 따옴표
                sv = str(v)
                if any(c in sv for c in ':#{}[]|>&*!?,') or '\n' in sv:
                    sv = sv.replace('"', '\\"')
                    f.write(f'{prefix}{k}: "{sv}"\n')
                else:
                    f.write(f"{prefix}{k}: {sv}\n")
    elif isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                first = True
                for k, v in item.items():
                    if first:
                        f.write(f"{prefix}- {k}:")
                        first = False
                    else:
                        f.write(f"{prefix}  {k}:")
                    if isinstance(v, (dict, list)):
                        f.write("\n")
                        _write_yaml(f, v, indent + 2)
                    elif isinstance(v, bool):
                        f.write(f" {'true' if v else 'false'}\n")
                    elif isinstance(v, (int, float)):
                        f.write(f" {v}\n")
                    else:
                        sv = str(v)
                        if any(c in sv for c in ':#{}[]|>&*!?,') or '\n' in sv:
                            sv = sv.replace('"', '\\"')
                            f.write(f' "{sv}"\n')
                        else:
                            f.write(f" {sv}\n")
            else:
                f.write(f"{prefix}- {item}\n")


# ============================================================
# 메인: 단일 XLSX 파일 변환
# ============================================================

def _insert_shape_references(out_dir, sheet_summary, shape_results):
    """Tier 1 MD 파일 끝에 도형 데이터 참조 링크를 삽입"""
    for sheet_name in shape_results:
        if sheet_name not in sheet_summary:
            continue
        safe_name = sheet_name.replace('/', '_').replace(' ', '_')
        md_path = os.path.join(out_dir, f"{safe_name}.md")
        if not os.path.exists(md_path):
            continue

        info = shape_results[sheet_name]
        ref_block = (
            f"\n\n---\n\n"
            f"> **도형/플로우차트 데이터**: 이 시트에는 {info['shapes']}개 도형, "
            f"{info['connectors']}개 연결선이 포함되어 있습니다.\n"
            f"> - AI 파싱용 (YAML): `{safe_name}_도형.yaml`\n"
            f"> - 시각화 (Mermaid): `{safe_name}_도형.md`\n"
        )

        with open(md_path, 'a', encoding='utf-8') as f:
            f.write(ref_block)


def convert_xlsx(xlsx_path, out_dir=None):
    """XLSX 파일을 AI-Readable Markdown으로 변환"""
    xlsx_path = os.path.abspath(xlsx_path)
    filename = os.path.basename(xlsx_path)
    stem = Path(xlsx_path).stem

    # 출력 디렉토리 결정
    if out_dir is None:
        base = os.path.dirname(xlsx_path)
        kb_dir = os.path.join(base, '..', '_knowledge_base', 'sheets')
        out_dir = os.path.join(os.path.abspath(kb_dir), stem.replace(' ', '_'))

    os.makedirs(out_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"변환 시작: {filename}")
    print(f"출력 폴더: {out_dir}")
    print(f"{'='*60}\n")

    # --- Tier 1: 시트 데이터 ---
    print("[Tier 1] 시트 셀 데이터 추출...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    theme_colors = _extract_theme_colors(wb)
    if theme_colors:
        print(f"  테마 색상 {len(theme_colors)}개 로드")
    sheet_summary = {}

    for name in wb.sheetnames:
        ws = wb[name]
        md = sheet_to_markdown(ws, name, filename, theme_colors)
        safe_name = name.replace('/', '_').replace(' ', '_')
        filepath = os.path.join(out_dir, f"{safe_name}.md")
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(md)
        sheet_summary[name] = {
            'rows': ws.max_row, 'cols': ws.max_column,
            'merged': len(ws.merged_cells.ranges),
            'images': len(ws._images),
            'file': filepath, 'chars': len(md)
        }
        print(f"  [{name}] {len(md):,} chars")

    # --- Tier 1.5: 도형 추출 ---
    print("\n[Tier 1.5] 도형/플로우차트 추출...")
    shape_results = extract_shapes_from_xlsx(xlsx_path, out_dir, filename)
    for sname, info in shape_results.items():
        print(f"  [{sname}] 도형 {info['shapes']}개, 연결 {info['connectors']}개")

    if not shape_results:
        print("  (도형 없음)")

    # --- 후처리: Tier 1 MD에 도형 참조 삽입 ---
    if shape_results:
        _insert_shape_references(out_dir, sheet_summary, shape_results)

    # --- INDEX 생성 ---
    index = {
        'source': filename,
        'source_path': xlsx_path,
        'sheets': sheet_summary,
        'shapes': shape_results,
        'tier1_total_chars': sum(v['chars'] for v in sheet_summary.values()),
        'tier15_total_chars': sum(v['chars'] for v in shape_results.values()),
    }

    index_path = os.path.join(out_dir, "_INDEX.json")
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    total = index['tier1_total_chars'] + index['tier15_total_chars']
    n_files = len(sheet_summary) + len(shape_results)
    print(f"\n완료: {n_files}개 파일, {total:,} chars")
    return index


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='XLSX → AI-Readable Markdown 변환')
    parser.add_argument('input', nargs='?', help='XLSX 파일 경로')
    parser.add_argument('--out', '-o', help='출력 폴더 (기본: _knowledge_base/sheets/{파일명})')
    parser.add_argument('--batch', '-b', help='폴더 내 모든 xlsx 일괄 변환')

    args = parser.parse_args()

    if args.batch:
        # 배치 모드: 폴더 내 모든 xlsx
        batch_dir = os.path.abspath(args.batch)
        xlsx_files = []
        for root, dirs, files in os.walk(batch_dir):
            # _knowledge_base 폴더는 건너뜀
            dirs[:] = [d for d in dirs if d != '_knowledge_base' and d != '_tools']
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~$'):
                    xlsx_files.append(os.path.join(root, f))

        print(f"발견된 XLSX 파일: {len(xlsx_files)}개\n")

        results = {}
        for i, xlsx in enumerate(sorted(xlsx_files), 1):
            print(f"\n[{i}/{len(xlsx_files)}] {os.path.basename(xlsx)}")
            try:
                idx = convert_xlsx(xlsx)
                results[xlsx] = {'status': 'success', 'files': len(idx['sheets']) + len(idx['shapes'])}
            except Exception as e:
                print(f"  오류: {e}")
                results[xlsx] = {'status': 'error', 'error': str(e)}

        # 배치 결과 요약
        success = sum(1 for v in results.values() if v['status'] == 'success')
        print(f"\n{'='*60}")
        print(f"배치 완료: {success}/{len(results)} 성공")
        return results

    elif args.input:
        # 단일 파일 모드
        return convert_xlsx(args.input, args.out)

    else:
        parser.print_help()


if __name__ == '__main__':
    main()
