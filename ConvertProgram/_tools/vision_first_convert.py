#!/usr/bin/env python3
"""
vision_first_convert.py - Vision-First XLSX → Markdown 변환 파이프라인

전략:
  [1차] LibreOffice headless → 시트별 PDF → 페이지별 PNG 캡처
  [2차] Vision API (Claude Opus) → 스크린샷으로부터 구조화된 Markdown 생성
  [3차] openpyxl 데이터 파싱 → 정확한 수치/숨겨진 셀/수식 보강
  [합성] Vision MD + openpyxl 데이터 → 최종 Markdown

사용법:
    python vision_first_convert.py <input.xlsx> [--output-dir DIR] [--dry-run]
    python vision_first_convert.py --folder 7_System [--skip-existing]

환경변수:
    AWS_BEARER_TOKEN_BEDROCK: Bedrock API Bearer Token
    AWS_REGION: AWS 리전 (기본: us-east-1)
"""

import argparse
import base64
import json
import os
import sys
import time
import tempfile
import shutil
from pathlib import Path

# .env 파일에서 환경변수 로드 (ConvertProgram/.env)
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).parent.parent / ".env"
    if _env_path.exists():
        load_dotenv(_env_path)
except ImportError:
    pass

# ─── 설정 ────────────────────────────────────────────────────────
VISION_MODEL = "claude-opus-4-6"
VISION_MAX_TOKENS = 16384
IMAGE_DPI = 150  # PNG 렌더링 해상도
MAX_IMAGE_HEIGHT = 4000  # Vision API에 보낼 최대 이미지 높이 (px)
OVERLAP_PX = 200  # 분할 시 겹침
SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
LO_PYTHON = r"C:\Program Files\LibreOffice\program\python.exe"
SCRIPT_DIR = Path(__file__).parent


def _set_vision_model(model: str):
    global VISION_MODEL
    VISION_MODEL = model


# ═══════════════════════════════════════════════════════════════════
# 1단계: 스크린샷 캡처 (LibreOffice → PDF → PNG)
# ═══════════════════════════════════════════════════════════════════

def capture_sheet_images(xlsx_path: str, work_dir: str) -> list[dict]:
    """XLSX → 시트별 PDF → 페이지별 PNG 변환

    Returns:
        [{"sheet_name": str, "sheet_index": int,
          "images": [path, ...], "page_count": int}, ...]
    """
    import subprocess
    import fitz  # PyMuPDF

    pdf_dir = os.path.join(work_dir, "pdfs")
    img_dir = os.path.join(work_dir, "images")
    os.makedirs(pdf_dir, exist_ok=True)
    os.makedirs(img_dir, exist_ok=True)

    # lo_sheet_export.py 호출
    print("[1/4] LibreOffice로 시트별 PDF 내보내기...")
    lo_script = str(SCRIPT_DIR / "lo_sheet_export.py")
    result = subprocess.run(
        [sys.executable, lo_script, xlsx_path, pdf_dir],
        capture_output=True, text=True, timeout=600,
    )
    if result.returncode != 0:
        err_msg = result.stderr.strip() or result.stdout.strip()
        print(f"  [WARN] lo_sheet_export 종료코드 {result.returncode}")
        if err_msg:
            # cp949 안전 출력
            try:
                print(f"  {err_msg[:500]}")
            except UnicodeEncodeError:
                print(f"  (stderr contains non-encodable chars)")

    # 매니페스트 읽기
    manifest_path = os.path.join(pdf_dir, "_export_manifest.json")
    if not os.path.exists(manifest_path):
        raise RuntimeError("시트 내보내기 매니페스트가 없습니다")

    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    # PDF → 페이지별 PNG
    print("[2/4] PDF → PNG 변환...")
    sheet_images = []

    for sheet_info in manifest["sheets"]:
        if not sheet_info.get("success"):
            print(f"  [SKIP] {sheet_info['name']}: PDF 내보내기 실패")
            continue

        pdf_path = sheet_info["pdf_path"]
        if not os.path.exists(pdf_path):
            continue

        doc = fitz.open(pdf_path)
        images = []

        for page_idx in range(doc.page_count):
            page = doc.load_page(page_idx)
            pix = page.get_pixmap(dpi=IMAGE_DPI)

            # 빈 페이지 감지 (내용이 거의 없는 페이지 건너뛰기)
            if _is_mostly_blank(pix):
                continue

            # 콘텐츠 영역으로 자르기
            pix = _crop_to_content(pix)

            safe_name = sheet_info["safe_name"]
            img_path = os.path.join(img_dir, f"{sheet_info['index']:02d}_{safe_name}_p{page_idx+1}.png")
            pix.save(img_path)
            images.append(img_path)

        doc.close()

        # 너무 큰 이미지 분할
        final_images = []
        for img_path in images:
            final_images.extend(_split_tall_image(img_path))

        sheet_images.append({
            "sheet_name": sheet_info["name"],
            "sheet_index": sheet_info["index"],
            "images": final_images,
            "page_count": len(final_images),
        })

        print(f"  {sheet_info['name']}: {len(final_images)} 이미지")

    return sheet_images


def _is_mostly_blank(pix, threshold=0.98) -> bool:
    """이미지가 거의 비어있는지 확인 (흰색 비율이 threshold 이상)"""
    try:
        from PIL import Image
        import io
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        pixels = list(img.convert("RGB").getdata())
        white_count = sum(1 for r, g, b in pixels if r > 240 and g > 240 and b > 240)
        ratio = white_count / len(pixels)
        return ratio > threshold
    except Exception:
        return False


def _crop_to_content(pix, margin=20):
    """콘텐츠 영역으로 이미지 자르기 (여백 제거)"""
    try:
        from PIL import Image
        import io

        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        # 비-흰색 영역 찾기
        pixels = img.load()
        w, h = img.size

        top = 0
        for y in range(h):
            row_has_content = False
            for x in range(0, w, 5):  # 5px 간격으로 샘플링
                r, g, b = pixels[x, y]
                if r < 240 or g < 240 or b < 240:
                    row_has_content = True
                    break
            if row_has_content:
                top = max(0, y - margin)
                break

        bottom = h
        for y in range(h - 1, -1, -1):
            row_has_content = False
            for x in range(0, w, 5):
                r, g, b = pixels[x, y]
                if r < 240 or g < 240 or b < 240:
                    row_has_content = True
                    break
            if row_has_content:
                bottom = min(h, y + margin)
                break

        left = 0
        for x in range(w):
            col_has_content = False
            for y in range(0, h, 5):
                r, g, b = pixels[x, y]
                if r < 240 or g < 240 or b < 240:
                    col_has_content = True
                    break
            if col_has_content:
                left = max(0, x - margin)
                break

        right = w
        for x in range(w - 1, -1, -1):
            col_has_content = False
            for y in range(0, h, 5):
                r, g, b = pixels[x, y]
                if r < 240 or g < 240 or b < 240:
                    col_has_content = True
                    break
            if col_has_content:
                right = min(w, x + margin)
                break

        if right - left < 50 or bottom - top < 50:
            return pix  # 너무 작으면 원본 반환

        cropped = img.crop((left, top, right, bottom))

        # PIL → PyMuPDF Pixmap 변환
        import fitz
        img_bytes = io.BytesIO()
        cropped.save(img_bytes, format="PNG")
        img_bytes.seek(0)
        return fitz.Pixmap(img_bytes)

    except Exception:
        return pix


def _split_tall_image(image_path: str) -> list[str]:
    """세로로 긴 이미지를 분할"""
    try:
        from PIL import Image
    except ImportError:
        return [image_path]

    img = Image.open(image_path)
    w, h = img.size

    if h <= MAX_IMAGE_HEIGHT:
        return [image_path]

    parts = []
    y = 0
    idx = 0
    stem = Path(image_path).stem
    parent = Path(image_path).parent

    while y < h:
        y_end = min(y + MAX_IMAGE_HEIGHT, h)
        crop = img.crop((0, y, w, y_end))
        part_path = str(parent / f"{stem}_part{idx}.png")
        crop.save(part_path, "PNG")
        parts.append(part_path)
        idx += 1
        y = y_end - OVERLAP_PX if y_end < h else h

    # 원본 이미지 제거
    os.unlink(image_path)
    return parts


# ═══════════════════════════════════════════════════════════════════
# 2단계: Vision API 분석
# ═══════════════════════════════════════════════════════════════════

def build_vision_prompt(sheet_name: str, workbook_name: str,
                        page_idx: int, total_pages: int) -> str:
    """Vision API에 보낼 시스템 프롬프트 생성"""
    context = ""
    if total_pages > 1:
        context = f"\n이 이미지는 '{sheet_name}' 시트의 {page_idx+1}/{total_pages} 페이지입니다."

    return f"""당신은 모바일 MMORPG 게임 기획서 변환 전문가입니다.

아래 이미지는 Excel 기획서 "{workbook_name}"의 "{sheet_name}" 시트 스크린샷입니다.{context}

## 작업 지시

이미지에 보이는 **모든 정보**를 빠짐없이 구조화된 Markdown으로 변환하세요.

### 변환 규칙

1. **테이블**: Markdown 테이블로 변환. 헤더 행과 데이터 행을 정확히 구분.
   - 병합 셀은 내용을 상위 셀에 포함하고 하위 셀은 비움
   - 숫자/수치는 정확히 옮기기
   - 색상으로 구분된 행/열이 있으면 색상의 의미를 주석으로 표기

2. **플로우차트/도형**: Mermaid 다이어그램으로 변환
   - 모든 도형의 텍스트 내용 포함
   - 화살표/연결선의 레이블(Yes/No, 조건 등) 포함
   - 도형의 색상이 의미를 갖는 경우 주석으로 표기

3. **게임 UI 스크린샷**: 상세 설명으로 변환
   - 화면에 보이는 모든 UI 요소 나열
   - 번호 어노테이션이 있으면 번호별로 설명
   - 버튼, 아이콘, 텍스트 필드 등 구분

4. **섹션 구조**: 원본의 시각적 구분을 Markdown 헤딩으로 변환
   - 색상 배너/제목 → ## 헤딩
   - 하위 그룹 → ### 서브헤딩
   - 주석/코멘트 → > 인용 블록

5. **특수 표기**:
   - ▶ 마커가 있는 항목은 그대로 유지
   - 스펙아웃(삭제/보류) 표시가 있으면 ~~취소선~~ 사용
   - 강조 표시는 **볼드** 사용

### 출력 형식

순수 Markdown만 출력하세요 (코드 블록으로 감싸지 마세요).
시트의 논리적 구조를 최대한 보존하면서 AI가 이해하기 쉬운 형태로 작성하세요.
"""


def analyze_sheet_with_vision(sheet_info: dict, workbook_name: str) -> str:
    """시트의 모든 이미지를 Vision API로 분석하여 Markdown 생성"""
    images = sheet_info["images"]
    if not images:
        return f"# {sheet_info['sheet_name']}\n\n> (빈 시트)\n"

    sheet_name = sheet_info["sheet_name"]
    total_pages = len(images)

    # 모든 페이지를 하나의 API 호출로 보내기 (최대 20개 이미지)
    if total_pages <= 20:
        prompt = build_vision_prompt(sheet_name, workbook_name, 0, total_pages)
        if total_pages > 1:
            prompt += f"\n\n총 {total_pages}개 이미지가 연속된 하나의 시트입니다. 모든 이미지의 내용을 합쳐서 하나의 통합 Markdown으로 작성하세요."

        result = call_bedrock_vision(prompt, images)
        return f"# {sheet_name}\n\n{result}\n"
    else:
        # 20개 초과 시 분할 호출
        all_parts = []
        for batch_start in range(0, total_pages, 15):
            batch = images[batch_start:batch_start + 15]
            prompt = build_vision_prompt(
                sheet_name, workbook_name, batch_start, total_pages)
            prompt += f"\n\n이미지 {batch_start+1}~{batch_start+len(batch)} / 전체 {total_pages}개"
            result = call_bedrock_vision(prompt, batch)
            all_parts.append(result)

        return f"# {sheet_name}\n\n" + "\n\n---\n\n".join(all_parts) + "\n"


def call_bedrock_vision(prompt: str, image_paths: list[str]) -> str:
    """AWS Bedrock Claude Vision API 호출"""
    import requests

    token = os.environ.get("AWS_BEARER_TOKEN_BEDROCK")
    if not token:
        raise RuntimeError(
            "AWS_BEARER_TOKEN_BEDROCK 환경변수가 설정되지 않았습니다.\n"
            "설정: export AWS_BEARER_TOKEN_BEDROCK='your-token'"
        )
    region = os.environ.get("AWS_REGION", "us-east-1")

    # 모델 ID 매핑
    model_mapping = {
        "claude-opus-4-6": "us.anthropic.claude-opus-4-6-v1",
        "claude-sonnet-4-6": "us.anthropic.claude-sonnet-4-6-v1:0",
    }
    model_id = model_mapping.get(VISION_MODEL, f"us.anthropic.{VISION_MODEL}-v1:0")

    url = f"https://bedrock-runtime.{region}.amazonaws.com/model/{model_id}/invoke"

    # 이미지 콘텐츠 구성
    content = []
    for img_path in image_paths:
        ext = Path(img_path).suffix.lower()
        media_types = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
        media_type = media_types.get(ext, "image/png")

        with open(img_path, "rb") as f:
            img_data = base64.standard_b64encode(f.read()).decode("utf-8")

        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": img_data},
        })

    content.append({"type": "text", "text": prompt})

    body = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": VISION_MAX_TOKENS,
        "messages": [{"role": "user", "content": content}],
    }

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}",
    }

    print(f"    Vision API 호출 ({len(image_paths)} 이미지, 모델: {VISION_MODEL})...")
    resp = requests.post(url, json=body, headers=headers, timeout=600)

    if resp.status_code != 200:
        raise RuntimeError(f"Bedrock API 오류 {resp.status_code}: {resp.text[:500]}")

    result = resp.json()
    text = result["content"][0]["text"]

    # 코드 블록으로 감싸져 있으면 제거
    if text.startswith("```markdown"):
        text = text[len("```markdown"):].strip()
    if text.startswith("```"):
        text = text[3:].strip()
    if text.endswith("```"):
        text = text[:-3].strip()

    usage = result.get("usage", {})
    input_tokens = usage.get("input_tokens", 0)
    output_tokens = usage.get("output_tokens", 0)
    print(f"    완료 (입력: {input_tokens:,} / 출력: {output_tokens:,} 토큰)")

    return text


# ═══════════════════════════════════════════════════════════════════
# 3단계: openpyxl 데이터 보강
# ═══════════════════════════════════════════════════════════════════

def extract_openpyxl_data(xlsx_path: str) -> dict[str, dict]:
    """openpyxl로 시트별 정확한 데이터 추출 (Vision 보강용)

    Returns:
        {"시트명": {"tables": [...], "formulas": [...], "hidden_data": [...]}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    sheet_data = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        ws_f = wb_formula[ws_name]
        data = {"tables": [], "formulas": [], "hidden_rows": [], "hidden_cols": []}

        # 숨겨진 행/열 감지
        for row_idx in range(1, ws.max_row + 1):
            rd = ws.row_dimensions.get(row_idx)
            if rd and rd.hidden:
                cells = []
                for col_idx in range(1, ws.max_column + 1):
                    val = ws.cell(row_idx, col_idx).value
                    if val is not None:
                        cells.append({"col": col_idx, "value": str(val)})
                if cells:
                    data["hidden_rows"].append({"row": row_idx, "cells": cells})

        for col_idx in range(1, ws.max_column + 1):
            cd = ws.column_dimensions.get(
                openpyxl.utils.get_column_letter(col_idx))
            if cd and cd.hidden:
                data["hidden_cols"].append(col_idx)

        # 수식 감지 (data_only=False에서)
        for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row,
                                   min_col=1, max_col=ws_f.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    computed = ws.cell(cell.row, cell.column).value
                    data["formulas"].append({
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "computed_value": str(computed) if computed is not None else None,
                    })

        # 테이블 영역의 정밀 데이터 (셀 값 전체)
        table_data = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 500),
                                 min_col=1, max_col=min(ws.max_column or 1, 50),
                                 values_only=False):
            row_vals = []
            for cell in row:
                if cell.value is not None:
                    row_vals.append({
                        "col": cell.column,
                        "value": str(cell.value),
                        "is_bold": cell.font.bold if cell.font else False,
                    })
            if row_vals:
                table_data.append({"row": row[0].row, "cells": row_vals})

        if table_data:
            data["tables"] = table_data

        sheet_data[ws_name] = data

    wb.close()
    wb_formula.close()
    return sheet_data


def build_reinforcement_section(sheet_name: str, openpyxl_data: dict) -> str:
    """openpyxl 데이터를 보강 섹션 Markdown으로 변환"""
    parts = []

    # 숨겨진 행
    hidden_rows = openpyxl_data.get("hidden_rows", [])
    if hidden_rows:
        parts.append("## 숨겨진 행 데이터\n")
        parts.append("> Vision API 스크린샷에는 보이지 않는 숨겨진 데이터입니다.\n")
        for hr in hidden_rows[:20]:  # 최대 20개
            cells_str = ", ".join(f"{c['col']}열={c['value']}" for c in hr["cells"])
            parts.append(f"- **행 {hr['row']}**: {cells_str}")
        parts.append("")

    # 수식
    formulas = openpyxl_data.get("formulas", [])
    if formulas:
        parts.append("## 수식 정보\n")
        parts.append("> 셀에 포함된 계산 수식과 결과값입니다.\n")
        for fm in formulas[:30]:  # 최대 30개
            parts.append(f"- **{fm['cell']}**: `{fm['formula']}` → {fm['computed_value']}")
        parts.append("")

    # 숨겨진 열
    hidden_cols = openpyxl_data.get("hidden_cols", [])
    if hidden_cols:
        from openpyxl.utils import get_column_letter
        col_letters = [get_column_letter(c) for c in hidden_cols]
        parts.append(f"## 숨겨진 열\n")
        parts.append(f"다음 열이 숨겨져 있습니다: {', '.join(col_letters)}\n")

    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════
# 4단계: 합성 및 출력
# ═══════════════════════════════════════════════════════════════════

def convert_single_xlsx(xlsx_path: str, output_dir: str = None,
                         dry_run: bool = False) -> dict:
    """단일 XLSX 파일을 Vision-First 방식으로 변환"""
    xlsx_path = os.path.abspath(xlsx_path)
    workbook_name = Path(xlsx_path).stem

    if output_dir is None:
        kb_root = str(SCRIPT_DIR.parent.parent / "_knowledge_base" / "sheets")
        output_dir = os.path.join(kb_root, workbook_name)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Vision-First 변환: {workbook_name}")
    print(f"{'='*60}")

    if dry_run:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        print(f"  시트: {', '.join(wb.sheetnames)}")
        wb.close()
        return {"status": "dry_run", "file": xlsx_path}

    start_time = time.time()

    # 임시 작업 디렉토리
    work_dir = tempfile.mkdtemp(prefix="vf_convert_")

    try:
        # 1. 스크린샷 캡처
        sheet_images = capture_sheet_images(xlsx_path, work_dir)

        # 2. Vision API 분석
        print("[3/4] Vision API로 시트별 분석...")
        vision_results = {}
        for si in sheet_images:
            sheet_name = si["sheet_name"]
            if si["page_count"] == 0:
                vision_results[sheet_name] = f"# {sheet_name}\n\n> (빈 시트)\n"
                continue

            md = analyze_sheet_with_vision(si, workbook_name)
            vision_results[sheet_name] = md

            # 시트별 Vision MD 저장
            safe = si["sheet_name"].replace("/", "_").replace("\\", "_")
            vision_md_path = os.path.join(output_dir, f"{safe}_vision.md")
            with open(vision_md_path, "w", encoding="utf-8") as f:
                f.write(md)

        # 3. openpyxl 데이터 보강
        print("[4/4] openpyxl 데이터 보강...")
        openpyxl_data = extract_openpyxl_data(xlsx_path)

        # 4. 합성: Vision MD + openpyxl 보강 → 최종 MD
        final_parts = []
        final_parts.append(f"# {workbook_name}\n")
        final_parts.append(f"> Vision-First 변환 | 원본: {Path(xlsx_path).name}")
        final_parts.append(f"> 변환일: {time.strftime('%Y-%m-%d %H:%M')}\n")

        for sheet_name in vision_results:
            final_parts.append(vision_results[sheet_name])

            # openpyxl 보강 데이터 추가
            if sheet_name in openpyxl_data:
                reinforcement = build_reinforcement_section(
                    sheet_name, openpyxl_data[sheet_name])
                if reinforcement.strip():
                    final_parts.append(f"\n---\n### 📊 데이터 보강 ({sheet_name})\n")
                    final_parts.append(reinforcement)

            final_parts.append("\n---\n")

        # 최종 MD 저장
        final_md = "\n".join(final_parts)
        final_path = os.path.join(output_dir, f"{workbook_name}.md")
        with open(final_path, "w", encoding="utf-8") as f:
            f.write(final_md)

        # 인덱스 JSON 저장
        elapsed = time.time() - start_time
        index = {
            "source": Path(xlsx_path).name,
            "method": "vision-first",
            "model": VISION_MODEL,
            "sheets": [si["sheet_name"] for si in sheet_images],
            "sheet_count": len(sheet_images),
            "output_file": final_path,
            "elapsed_seconds": round(elapsed, 1),
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        }

        index_path = os.path.join(output_dir, "_INDEX.json")
        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False, indent=2)

        print(f"\n[OK] 변환 완료: {final_path}")
        print(f"  소요시간: {elapsed:.1f}초")
        print(f"  출력 크기: {len(final_md):,} bytes")

        return {"status": "success", "file": xlsx_path, "output": final_path,
                "elapsed": elapsed, "size": len(final_md)}

    finally:
        # 임시 디렉토리 정리
        try:
            shutil.rmtree(work_dir, ignore_errors=True)
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Vision-First XLSX → Markdown 변환 파이프라인")
    parser.add_argument("input", nargs="?",
                        help="변환할 XLSX 파일 또는 폴더 경로")
    parser.add_argument("--output-dir", "-o",
                        help="출력 디렉토리 (기본: _knowledge_base/sheets/{파일명})")
    parser.add_argument("--folder", "-f",
                        help="폴더 내 모든 XLSX 일괄 변환 (예: 7_System)")
    parser.add_argument("--dry-run", action="store_true",
                        help="실제 변환 없이 대상 확인만")
    parser.add_argument("--skip-existing", action="store_true",
                        help="이미 변환된 파일 건너뛰기")
    parser.add_argument("--model", default=VISION_MODEL,
                        help=f"Vision 모델 (기본: {VISION_MODEL})")
    args = parser.parse_args()

    if args.model != VISION_MODEL:
        _set_vision_model(args.model)

    if args.folder:
        # 폴더 일괄 변환
        project_root = str(SCRIPT_DIR.parent.parent)
        folder_path = os.path.join(project_root, args.folder)
        if not os.path.isdir(folder_path):
            print(f"ERROR: 폴더 없음: {folder_path}")
            sys.exit(1)

        xlsx_files = sorted(Path(folder_path).glob("*.xlsx"))
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

        print(f"대상 폴더: {args.folder}")
        print(f"XLSX 파일: {len(xlsx_files)}개")

        if args.dry_run:
            for f in xlsx_files:
                print(f"  - {f.name}")
            return

        results = []
        for xlsx_file in xlsx_files:
            if args.skip_existing:
                kb_dir = str(SCRIPT_DIR.parent.parent / "_knowledge_base" / "sheets" / xlsx_file.stem)
                index_path = os.path.join(kb_dir, "_INDEX.json")
                if os.path.exists(index_path):
                    print(f"\n[SKIP] {xlsx_file.name}: 이미 변환됨")
                    continue

            result = convert_single_xlsx(str(xlsx_file), args.output_dir, args.dry_run)
            results.append(result)

        # 요약
        success = sum(1 for r in results if r["status"] == "success")
        print(f"\n{'='*60}")
        print(f"일괄 변환 완료: {success}/{len(results)} 성공")

    elif args.input:
        # 단일 파일 변환
        if not os.path.exists(args.input):
            print(f"ERROR: 파일 없음: {args.input}")
            sys.exit(1)
        convert_single_xlsx(args.input, args.output_dir, args.dry_run)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
