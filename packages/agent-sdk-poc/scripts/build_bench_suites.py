"""
벤치마크 질문 suite 생성기 — Full + Sample 두 종류 관리.

Full suite (`benchmark_full.json`):
  - DataSheet 73 xlsx × 2 질문 = 146
  - 추가: preset_prompts.py 의 기존 24 프리셋 (회귀 커버)
  - 총 ~170 질문, full 1회 ~3시간 / ~$80
  - 위클리/배포 전 thorough 회귀 용도

Sample suite (`benchmark_sample.json`):
  - KV 7 시트 × 2 질문 = 14 (DataSheet 인덱스 효과 직접 측정)
  - Instance 카테고리별 random 5 (회귀 분광)
  - 기존 preset 5 (overview/cross/spec 분광)
  - 총 24 질문, ~30분 / ~$15
  - 매 코드 변경 후 빠른 검증 용도

각 질문 객체:
  {
    "id": "kv-ContentSetting-MetamorphSwitchCoolTime",
    "category": "kv-shortform" | "instance" | "preset",
    "question": "메타모프 변경 쿨타임의 정확한 값과 상수명은?",
    "expected": {                       # answer key (선택, 자동 채점용 시드)
       "must_contain": ["MetamorphSwitchCoolTime", "3000"],
       "source_hint": "ContentSetting"
    },
    "compare_mode": false,
    "source": "ContentSetting.xlsx#row42"
  }

Usage:
    python scripts/build_bench_suites.py            # 두 suite 모두 생성
    python scripts/build_bench_suites.py --sample   # sample 만
    python scripts/build_bench_suites.py --full     # full 만
"""
from __future__ import annotations

import argparse
import json
import os
import random
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import glob

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
DESIGN = Path(os.environ.get("PROJK_DESIGN_ROOT", "D:/ProjectK/Resource/design"))
OUT_DIR = ROOT / "scripts" / "bench_out"

# 질문 생성 시드 — 같은 seed 면 같은 질문이 나와서 before/after 비교 안정.
SAMPLE_SEED = 4242
FULL_SEED = 7

# KV 시트 (build_datasheet_index.py 와 동일)
KV_SHEETS = [
    ("ContentSetting.xlsx", "ContentSetting"),
    ("Karma.xlsx", "Karma"),
    ("Mail.xlsx", "MailClass"),
    ("Quest.xlsx", "QuestBase"),
    ("SupportMode.xlsx", "SupportMode"),
    ("SupportMode.xlsx", "SupportModeTimeDungeon"),
    ("TableAttribute.xlsx", "Attribute"),
]


@dataclass
class BenchQuestion:
    id: str
    category: str
    question: str
    expected: dict = field(default_factory=dict)
    compare_mode: bool = False
    source: str = ""


def make_kv_question(file_stem: str, sheet: str, header: list[str], row: dict) -> BenchQuestion | None:
    """KV 시트의 한 행 → 단답형 질문.

    예시 row: {ContentSettingEnum=MetamorphSwitchCoolTime, Value=3000, Comment="메타모프 변경 쿨타임(ms)"}
    → 질문: "메타모프 변경 쿨타임의 정확한 값과 상수명은? (DataSheet 우선)"
    → expected: must_contain=["MetamorphSwitchCoolTime", "3000"]
    """
    key = row.get(header[0], "").strip() if header else ""
    if not key:
        return None
    # 값 컬럼 후보: header[1] 또는 첫 숫자/단위 가진 컬럼
    value = ""
    if len(header) > 1:
        value = (row.get(header[1], "") or "").strip()
    # 한국어 코멘트
    comment = ""
    for h in header[1:]:
        v = (row.get(h, "") or "").strip()
        if v and any("가" <= ch <= "힣" for ch in v):
            comment = v
            break
    if not comment:
        # 한국어 코멘트가 없으면 키 자체로 질문 (영문)
        comment = key
    # 질문 형식
    q = f"{comment[:60]} 의 정확한 값과 상수명은? (DataSheet 우선)"
    expected_must = [key]
    if value and len(value) < 25:
        expected_must.append(value)
    return BenchQuestion(
        id=f"kv-{file_stem}-{key}",
        category="kv-shortform",
        question=q,
        expected={
            "must_contain": expected_must,
            "source_hint": f"DataSheet / {sheet}",
        },
        source=f"{file_stem}.xlsx / {sheet} → {key}={value}",
    )


def collect_kv_rows(file_stem: str, sheet: str) -> tuple[list[str], list[dict]]:
    p = DESIGN / f"{file_stem}.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return [], []
    ws = wb[sheet]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header = [h for h in header if h]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or r[0] is None:
            continue
        rec = {h: (str(v).strip() if v is not None else "") for h, v in zip(header, r)}
        if rec.get(header[0], ""):
            rows.append(rec)
    return header, rows


def make_instance_question(file_stem: str, sheet: str, sample_rows: list[dict], header: list[str]) -> BenchQuestion | None:
    """Instance 시트 → query_game_table 질문.

    행 1개 골라서 'Id=N 행의 주요 컬럼은?' 또는 'Comment 가 X 인 행의 ID 와 핵심 수치는?' 형식.
    """
    if not sample_rows or not header:
        return None
    row = sample_rows[0]
    id_val = row.get("Id", "") or row.get(header[0], "")
    if not id_val:
        return None
    # Comment 컬럼이 있으면 더 풍부한 질문
    comment = row.get("Comment", "") or ""
    name_key = ""
    for k in ("NameKey", "TextkeyTitle", "Name", "TextKey"):
        if k in row and row[k]:
            name_key = row[k]
            break
    if comment:
        q = f"DataSheet {sheet} 테이블에서 Id={id_val} ({comment[:30]}) 행의 핵심 컬럼 (Name/Id/주요 수치) 을 알려줘."
    else:
        q = f"DataSheet {sheet} 테이블의 Id={id_val} 행의 핵심 컬럼을 알려줘."
    must = [str(id_val)]
    if name_key:
        must.append(name_key[:20])
    return BenchQuestion(
        id=f"inst-{file_stem}-{id_val}",
        category="instance",
        question=q,
        expected={"must_contain": must, "source_hint": f"DataSheet / {sheet}"},
        source=f"{file_stem}.xlsx / {sheet} → Id={id_val}",
    )


def collect_instance_rows(file_stem: str, sheet: str, n_rows: int = 3) -> tuple[list[str], list[dict]]:
    p = DESIGN / f"{file_stem}.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return [], []
    ws = wb[sheet]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header = [h for h in header if h]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or r[0] is None:
            continue
        rec = {h: ("" if v is None else str(v).strip()) for h, v in zip(header, r)}
        if not rec.get(header[0], ""):
            continue
        rows.append(rec)
        if len(rows) >= 50:
            break
    rows = rows[:n_rows] if rows else []
    return header, rows


def build_sample(rng: random.Random) -> list[BenchQuestion]:
    qs: list[BenchQuestion] = []

    # === Group A: KV 14 (각 KV 시트당 2 random) ===
    for fname, sheet in KV_SHEETS:
        stem = Path(fname).stem
        header, rows = collect_kv_rows(stem, sheet)
        if not rows:
            continue
        n = min(2, len(rows))
        picked = rng.sample(rows, n)
        for r in picked:
            q = make_kv_question(stem, sheet, header, r)
            if q:
                qs.append(q)

    # === Group B: Instance random 5 — 카테고리별 분광 ===
    instance_files = [
        ("Buff.xlsx", "BuffClass"),
        ("Skill.xlsx", "Skill"),
        ("MonsterClass.xlsx", "MonsterClass"),
        ("ItemEquipClass.xlsx", "ItemEquipClass"),
        ("Achievement.xlsx", "Achievement"),
    ]
    for fname, sheet in instance_files:
        stem = Path(fname).stem
        try:
            header, rows = collect_instance_rows(stem, sheet, n_rows=2)
        except Exception:
            continue
        if not rows or not header:
            continue
        # rng 로 1개 픽
        picked = [rng.choice(rows)] if rows else []
        for r in picked:
            q = make_instance_question(stem, sheet, [r], header)
            if q:
                qs.append(q)

    # === Group C: 기존 preset 5 (overview/cross/spec/system regression) ===
    preset_path = ROOT / "src" / "preset_prompts.py"
    if preset_path.exists():
        # ad-hoc 추출 — eval 위험하지만 trusted source
        ns = {}
        exec(preset_path.read_text(encoding="utf-8"), ns)
        all_presets = [p for p in ns.get("PRESETS", []) if not p.get("compare_mode")]
        # 카테고리 다양성: system/spec/cross/overview/content 한 개씩
        seen_cat = set()
        regression = []
        rng.shuffle(all_presets)
        for p in all_presets:
            cat = p.get("category", "")
            if cat in seen_cat or cat == "datasheet":
                continue
            seen_cat.add(cat)
            regression.append(p)
            if len(regression) >= 5:
                break
        for p in regression:
            qs.append(BenchQuestion(
                id=f"preset-{p['label'][:20].replace(' ','_')}",
                category=f"preset-{p.get('category','?')}",
                question=p["prompt"],
                compare_mode=False,
                source=f"preset: {p['label']}",
            ))
    return qs


def build_full(rng: random.Random) -> list[BenchQuestion]:
    """Full = 73 xlsx × 2 질문 + 모든 preset."""
    qs: list[BenchQuestion] = []
    seen = set()

    # 73 xlsx 모두 — KV 와 Instance 자동 분류해서 질문 만듦
    for f in sorted(glob.glob(str(DESIGN / "*.xlsx"))):
        stem = Path(f).stem
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        for sheet in wb.sheetnames:
            if sheet.lower() == "info":
                continue
            ws = wb[sheet]
            try:
                hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            if not hdr_row or hdr_row[0] is None:
                continue
            header = [str(c).strip() if c is not None else "" for c in hdr_row]
            header = [h for h in header if h]
            if not header:
                continue
            # 첫 데이터 행
            try:
                first_data = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            except StopIteration:
                continue
            if not first_data or first_data[0] is None:
                continue
            # KV 패턴 — 첫 컬럼 enum-like 문자열
            is_kv = isinstance(first_data[0], str) and not header[0].lower().startswith("id")
            try:
                if is_kv:
                    h, rows = collect_kv_rows(stem, sheet)
                    if not rows:
                        continue
                    n = min(2, len(rows))
                    picked = rng.sample(rows, n)
                    for r in picked:
                        q = make_kv_question(stem, sheet, h, r)
                        if q and q.id not in seen:
                            qs.append(q)
                            seen.add(q.id)
                else:
                    h, rows = collect_instance_rows(stem, sheet, n_rows=20)
                    if not rows:
                        continue
                    n = min(2, len(rows))
                    picked = rng.sample(rows, n)
                    for r in picked:
                        q = make_instance_question(stem, sheet, [r], h)
                        if q and q.id not in seen:
                            qs.append(q)
                            seen.add(q.id)
            except Exception:
                continue
            break  # 한 xlsx 의 첫 시트만

    # 모든 preset 추가
    preset_path = ROOT / "src" / "preset_prompts.py"
    if preset_path.exists():
        ns = {}
        exec(preset_path.read_text(encoding="utf-8"), ns)
        for p in ns.get("PRESETS", []):
            qs.append(BenchQuestion(
                id=f"preset-{p['label'][:20].replace(' ','_')}",
                category=f"preset-{p.get('category','?')}",
                question=p["prompt"],
                compare_mode=p.get("compare_mode", False),
                source=f"preset: {p['label']}",
            ))
    return qs


def serialize(qs: list[BenchQuestion]) -> list[dict]:
    return [
        {
            "id": q.id,
            "category": q.category,
            "question": q.question,
            "expected": q.expected,
            "compare_mode": q.compare_mode,
            "source": q.source,
        }
        for q in qs
    ]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sample", action="store_true")
    ap.add_argument("--full", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    do_sample = args.sample or not args.full
    do_full = args.full or not args.sample

    if do_sample:
        rng = random.Random(SAMPLE_SEED)
        qs = build_sample(rng)
        out = OUT_DIR / "benchmark_sample.json"
        out.write_text(json.dumps({"seed": SAMPLE_SEED, "questions": serialize(qs)}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[sample] {len(qs)} 질문 → {out.relative_to(ROOT)}")
        # 카테고리 breakdown
        from collections import Counter
        cnt = Counter(q.category for q in qs)
        for c, n in cnt.items():
            print(f"           {c:<30} {n}")

    if do_full:
        rng = random.Random(FULL_SEED)
        qs = build_full(rng)
        out = OUT_DIR / "benchmark_full.json"
        out.write_text(json.dumps({"seed": FULL_SEED, "questions": serialize(qs)}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[full]   {len(qs)} 질문 → {out.relative_to(ROOT)}")
        from collections import Counter
        cnt = Counter(q.category for q in qs)
        for c, n in cnt.items():
            print(f"           {c:<30} {n}")


if __name__ == "__main__":
    main()
