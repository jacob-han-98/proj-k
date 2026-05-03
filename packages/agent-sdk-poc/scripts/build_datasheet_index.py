"""
Step 2 — DataSheet KV 인덱스 빌더 (LLM-free, 결정적 추출).

목적:
  D:/ProjectK/Resource/design/ 의 KV (key-value) 패턴 xlsx 들을 읽어
  index/summaries/datasheet/<파일명>.md 로 저장.
  Agent 의 표준 워크플로우 1단계 "Grep -i 키워드 index/summaries/" 가
  새 datasheet/ 디렉토리도 자연스럽게 hit 함.

KV 패턴 시트 (현재 7개):
  ContentSetting / Karma / Mail / Quest / SupportMode (2시트) / TableAttribute

Instance 패턴은 별도 빌더 (data-pipeline 의 컬럼 메타) 에서 처리.

Usage:
    python scripts/build_datasheet_index.py            # 기본 (전체 KV)
    python scripts/build_datasheet_index.py --only ContentSetting
    python scripts/build_datasheet_index.py --dry-run  # 출력 경로만 확인
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# 동의어 사전 — 한국어 어휘 mismatch 해결 (ex. 메타모프↔변신)
sys.path.insert(0, str(Path(__file__).resolve().parent / "lib"))
try:
    from synonyms import expand_synonyms
except ImportError:
    def expand_synonyms(s):
        return s

ROOT = Path(__file__).resolve().parent.parent  # packages/agent-sdk-poc/
OUT_DIR = ROOT / "index" / "summaries" / "datasheet"

# Windows P4 동기화 위치 — 환경변수로 override 가능.
DESIGN_ROOT = Path(os.environ.get("PROJK_DESIGN_ROOT", "D:/ProjectK/Resource/design"))

# DataSheet 인용 prefix (CLAUDE.md 와 일치)
P4_PREFIX = "//main/ProjectK/Resource/design"


@dataclass
class KvSpec:
    """KV 시트 명세."""
    file: str         # xlsx 파일명 (확장자 포함)
    sheet: str        # 시트명
    title_label: str  # 사람이 읽는 제목


# 분류 결과(scan) 에서 KV 로 잡힌 7 시트.
KV_SHEETS: list[KvSpec] = [
    KvSpec("ContentSetting.xlsx", "ContentSetting", "글로벌 게임 상수 (쿨타임/제한/배율/비용/임계값)"),
    KvSpec("Karma.xlsx",          "Karma",          "성향 (Karma) 등급 정의"),
    KvSpec("Mail.xlsx",           "MailClass",      "메일 분류 (캐릭터/계정/월드)"),
    KvSpec("Quest.xlsx",          "QuestBase",      "직업별 시작 퀘스트 ID"),
    KvSpec("SupportMode.xlsx",    "SupportMode",    "지원 모드 분류"),
    KvSpec("SupportMode.xlsx",    "SupportModeTimeDungeon", "타임던전별 지원 모드 매핑"),
    KvSpec("TableAttribute.xlsx", "Attribute",      "테이블 속성 메타 (cs/s/c 도메인)"),
]


def first_meaningful_korean_label(text: str) -> str | None:
    """문자열에서 한국어 라벨이 있으면 추출 (라벨/제목용 Comment 셀)."""
    if not text:
        return None
    s = str(text).strip()
    if any("가" <= ch <= "힣" for ch in s):
        return s
    return None


def render_kv_sheet(spec: KvSpec) -> tuple[str, dict]:
    """KV 시트를 markdown 으로 렌더 + 메타 통계 리턴."""
    p = DESIGN_ROOT / spec.file
    if not p.exists():
        return "", {"error": f"file missing: {p}"}
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if spec.sheet not in wb.sheetnames:
        return "", {"error": f"sheet missing: {spec.sheet} in {spec.file}"}
    ws = wb[spec.sheet]

    # 헤더 (1행) — 컬럼명 추출
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    # 1열 = key (enum/Type), 2열 = value 또는 Comment 일반적 패턴.
    # 일부 시트는 Value 가 없고 다른 컬럼 (TimeDungeonId) 가 값임. 동적 처리.

    # 데이터 행 (3행부터; 2행은 domain meta)
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or r[0] is None:
            continue
        rec = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            rec[h] = "" if v is None else str(v).strip()
        if rec.get(header[0], "") == "":
            continue
        rows.append(rec)

    # 빈 컬럼 제거 (전 행이 비어있으면 표기에서 빠짐)
    nonempty_cols = [h for h in header if h and any((row.get(h, "") != "") for row in rows)]
    # 너무 길어지면 첫 4개만
    display_cols = nonempty_cols[:4]

    # markdown 렌더
    title = f"# DataSheet / {spec.sheet} (요약)"
    src_url = f"{P4_PREFIX}/{spec.file}"
    out: list[str] = []
    out.append(title)
    out.append(f"")
    out.append(f"> 출처: DataSheet / {spec.sheet}")
    out.append(f"> 원본: {src_url}")
    out.append(f"> 목적: {spec.title_label}")
    out.append(f"")
    out.append(f"## 한 줄 설명")
    out.append(f"{spec.title_label}. 총 {len(rows)} 항목.")
    out.append(f"")
    out.append(f"## 항목 ({len(rows)})")
    out.append(f"")
    # 각 행: `KEY` = `VALUE` — 한국어 코멘트
    for row in rows:
        key = row.get(header[0], "")
        # value 후보: header[1] (대부분 Value 또는 다음 의미 컬럼)
        value = ""
        if len(display_cols) >= 2:
            value = row.get(display_cols[1], "")
        # 한국어 코멘트 — Comment 컬럼이 있으면 그걸, 없으면 라벨다운 컬럼
        comment = ""
        for h in display_cols[1:]:
            v = row.get(h, "")
            kor = first_meaningful_korean_label(v)
            if kor:
                comment = kor
                break
        # 출력 라인 — 코멘트에 동의어 확장 적용
        line = f"- `{key}`"
        if value and value != comment:
            line += f" = `{value}`"
        if comment:
            line += f" — {expand_synonyms(comment)}"
        out.append(line)
    out.append("")
    out.append("## 인용 형식")
    out.append(f"`(출처: DataSheet / {spec.sheet} § <KEY>)`")
    out.append("")
    md = "\n".join(out)

    return md, {
        "rows": len(rows),
        "displayed_cols": display_cols,
        "src": src_url,
        "out_chars": len(md),
    }


def output_path(spec: KvSpec) -> Path:
    """저장 경로. SupportMode 처럼 한 파일에 두 시트면 시트명 suffix."""
    base = spec.file.replace(".xlsx", "")
    if spec.sheet != base:
        # 같은 xlsx 안의 보조 시트 — `<File>_<Sheet>.md`
        return OUT_DIR / f"{base}_{spec.sheet}.md"
    return OUT_DIR / f"{base}.md"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", default=None, help="특정 시트만 (e.g. ContentSetting)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[build] DESIGN_ROOT = {DESIGN_ROOT}")
    print(f"[build] OUT_DIR     = {OUT_DIR}")
    print()

    specs = KV_SHEETS
    if args.only:
        specs = [s for s in specs if s.sheet == args.only or s.file.startswith(args.only)]
        if not specs:
            print(f"[build] no match for --only {args.only!r}")
            return 1

    summary = []
    for spec in specs:
        out_p = output_path(spec)
        if args.dry_run:
            print(f"  [dry] {spec.file} / {spec.sheet}  →  {out_p}")
            continue
        md, meta = render_kv_sheet(spec)
        if "error" in meta:
            print(f"  [skip] {spec.file} / {spec.sheet}: {meta['error']}")
            continue
        out_p.write_text(md, encoding="utf-8")
        print(f"  [ok ] {spec.file} / {spec.sheet}  →  {out_p.relative_to(ROOT)}  ({meta['rows']} rows, {meta['out_chars']:,} chars)")
        summary.append((spec.sheet, meta["rows"], meta["out_chars"]))

    if not args.dry_run:
        total_rows = sum(s[1] for s in summary)
        total_chars = sum(s[2] for s in summary)
        print()
        print(f"=== 인덱스 빌드 완료 ===")
        print(f"   {len(summary)} 파일 · {total_rows} 항목 · {total_chars:,} 자")
        print()
        for sn, rows, chars in summary:
            print(f"   {sn:<30}  {rows:>4} rows   {chars:>6,} chars")

    return 0


if __name__ == "__main__":
    sys.exit(main())
