#!/usr/bin/env python3
"""capture_failure_report.py — 8_Contents 캡처 실패(OnlyOffice error -10 등) 리포트.

입력 xlsx 73개 vs 성공 캡처(_capture_manifest.json OK) 비교 → 실패 목록.
각 실패 파일은 openpyxl 로 텍스트 추출(렌더 실패해도 셀 데이터는 읽힘) → 조회 가능.
산출: data/capture_failures.json + data/capture_fail_text/<stem>.md
"""
import json, sys, re
from pathlib import Path
import openpyxl

P4 = Path("/home/jacob/p4sync/Design/8_Contents")
OUT = Path("/home/jacob/proj-k-data/xlsx-extractor/output/8_Contents")
ROOT = Path("/home/jacob/repos/proj-k/packages/agent-sdk-poc")
DATA = ROOT / "data"
TEXTDIR = DATA / "capture_fail_text"
TEXTDIR.mkdir(parents=True, exist_ok=True)
LOG = Path("/home/jacob/proj-k-data/_logs/8c_capture.log")

files = sorted(p for p in P4.rglob("*.xlsx") if "아이콘" not in str(p))

def captured_ok(stem):
    mf = OUT / stem / "_capture_manifest.json"
    if not mf.exists():
        return False
    try:
        m = json.loads(mf.read_text())
        sheets = m.get("sheets", [])
        return any(s.get("split_success") or s.get("blank") for s in sheets)
    except Exception:
        return False

# 로그에서 파일별 에러코드 추출
errmap = {}
if LOG.exists():
    cur = None
    for line in LOG.read_text(errors="ignore").splitlines():
        m = re.search(r"\[capture\] File: (.+\.xlsx)", line)
        if m: cur = m.group(1)
        e = re.search(r"error code=(-?\d+)", line)
        if e and cur: errmap[cur] = f"OnlyOffice error {e.group(1)}"
        if "page/sheet mismatch" in line and cur: errmap[cur] = "page/sheet mismatch"

def extract_text(xlsx):
    """openpyxl 로 시트별 비어있지 않은 셀 텍스트 추출 (렌더 실패해도 데이터 접근)."""
    out = []
    try:
        wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    except Exception as e:
        return f"(openpyxl 로드 실패: {e})"
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        rows = []
        for r in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in r if c is not None and str(c).strip()]
            if cells:
                rows.append(" | ".join(cells))
            if len(rows) > 400:
                rows.append("…(이하 생략)")
                break
        out.append(f"## 시트: {ws.title}  ({len(rows)} 행)\n" + "\n".join(rows[:400]))
    wb.close()
    return "\n\n".join(out) if out else "(텍스트 없음 — 이미지 위주)"

failures = []
for f in files:
    stem = f.stem
    if captured_ok(stem):
        continue
    size_mb = round(f.stat().st_size / 1048576, 1)
    try:
        wb = openpyxl.load_workbook(str(f), read_only=True)
        sheets = [s for s in wb.sheetnames]
        wb.close()
    except Exception:
        sheets = []
    # 텍스트 추출 저장
    txt = extract_text(f)
    safe = re.sub(r'[/\\:*?"<>|]', "_", stem)
    (TEXTDIR / f"{safe}.md").write_text(f"# {stem}\n> {f}\n> 크기 {size_mb}MB · {len(sheets)} 시트 · {errmap.get(f.name,'캡처 실패')}\n\n{txt}", encoding="utf-8")
    failures.append({"file": f.name, "stem": stem, "path": str(f), "size_mb": size_mb,
                     "sheets": sheets, "sheet_count": len(sheets),
                     "error": errmap.get(f.name, "캡처 실패(미상)"),
                     "text_md": str(TEXTDIR / f"{safe}.md")})

report = {"generated": None, "total_input": len(files),
          "captured_ok": len(files) - len(failures), "failed": len(failures),
          "failures": sorted(failures, key=lambda x: -x["size_mb"])}
(DATA / "capture_failures.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"입력 {len(files)} / 성공 {report['captured_ok']} / 실패 {len(failures)}")
print(f"리포트: {DATA/'capture_failures.json'}")
for x in report["failures"]:
    print(f"  ✗ {x['size_mb']:>6}MB  {x['sheet_count']}시트  {x['error']:<22} {x['file']}")
