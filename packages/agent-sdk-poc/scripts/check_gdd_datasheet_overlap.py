"""
사용자 가설 검증:
  "기획서에 정의된 식별자는 DataSheet 에도 모두 있어야 한다."

Step 1 의 GDD content.md 식별자 스캔 결과 (2,347 unique) 와
DataSheet 73 xlsx 의 모든 식별자 (테이블명 + 컬럼명 + KV enum + Instance Comment 인근 식별자)
를 비교해서 overlap % 를 보고함.

출력:
  - overlap %, missing %, missing 식별자 sample
  - missing 식별자 카테고리 분류 (UI 키 / 메시지 키 / 그 외)
  - 결정 가이드: 95%+ overlap → DataSheet 인덱스만으로 충분, 그 미만이면 GDD 식별자도 별도 인덱스 필요
"""
from __future__ import annotations

import glob
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

DESIGN_ROOT = Path("D:/ProjectK/Resource/design")
SCAN_RESULT = Path(__file__).resolve().parent / "bench_out" / "identifier_scan_full.json"


# 식별자 패턴 (scan_identifiers.py 와 동일)
PAT_CAMEL = re.compile(r"\b[A-Z][a-z]{1,}(?:[A-Z][a-z0-9]+){1,}\b")
PAT_SNAKE = re.compile(r"\b[A-Z][a-zA-Z0-9]+(?:_[A-Z][a-zA-Z0-9]+)+\b")
PAT_UPPER_SNAKE = re.compile(r"\b[A-Z]{3,}(?:_[A-Z0-9]{2,})+\b")
STOPWORDS = {"TRUE","FALSE","NULL","NONE","YES","NO","AND","OR","NOT","START","END","BEGIN","STOP","MIN","MAX","AVG","SUM","COUNT",
             "TD","TR","TH","BR","BM","LR","UL","OL","LI","OOXML","PDF","PPT","DOC","ID","UI","UX","FX","OK","IO",
             "PK","PK_","NPC","PC","PVP","PVE","DPS","HP","MP","EXP","SP","AP"}
MIN_LEN = 4


def is_meaningful(token: str) -> bool:
    if len(token) < MIN_LEN: return False
    if token in STOPWORDS: return False
    if token.isupper() and "_" not in token and len(token) <= 5: return False
    return True


def extract_identifiers_from_string(s: str) -> set[str]:
    if not s: return set()
    found = set()
    for pat in (PAT_SNAKE, PAT_CAMEL, PAT_UPPER_SNAKE):
        for m in pat.finditer(s):
            t = m.group(0)
            if is_meaningful(t):
                found.add(t)
    return found


def collect_datasheet_identifiers() -> tuple[set[str], dict]:
    """73 xlsx 의 모든 식별자 모음.
    추출 소스:
      1. xlsx 파일명 (스템) — Buff, Skill, MonsterClass 등
      2. 시트명 — BuffClass, BuffOverlapRule 등
      3. 컬럼명 (1행) — Id, MaxHp, AttackType 등
      4. 데이터 셀 안의 식별자 — Comment 셀의 영문 식별자 포함
    """
    files = sorted(glob.glob(str(DESIGN_ROOT / "*.xlsx")))
    ids: set[str] = set()
    file_count = len(files)
    sheet_count = 0
    cell_count = 0
    for f in files:
        stem = Path(f).stem
        if is_meaningful(stem):
            ids.add(stem)
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            sheet_count += 1
            if is_meaningful(sn):
                ids.add(sn)
            ws = wb[sn]
            # 헤더 (1행)
            try:
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            if hdr:
                for c in hdr:
                    if c is None: continue
                    s = str(c).strip()
                    if is_meaningful(s):
                        ids.add(s)
                    # 셀 자체에 한국어 + 영문 식별자가 섞여있을 수 있음
                    ids.update(extract_identifiers_from_string(s))
            # 데이터 셀 (3행 이후, 첫 5,000 셀까지만 — 식별자는 보통 enum/comment/이름 컬럼에 몰림)
            cells_scanned = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row is None: continue
                for v in row:
                    if v is None: continue
                    if isinstance(v, (int, float)): continue
                    s = str(v).strip()
                    if not s: continue
                    cells_scanned += 1
                    cell_count += 1
                    # 짧은 식별자만 보고 (긴 코멘트는 한국어가 대부분 — extract 패턴이 어차피 영문식별자만 잡음)
                    ids.update(extract_identifiers_from_string(s))
                    if cells_scanned >= 5000:
                        break
                if cells_scanned >= 5000:
                    break
    return ids, {"files": file_count, "sheets": sheet_count, "cells_scanned": cell_count}


def categorize_missing(token: str) -> str:
    """GDD 에만 있는 식별자의 성격 분류."""
    if token.startswith(("Meta_", "Sys_", "Hud_", "UIString_", "TextKey_", "Quest_", "Mon_", "Npc_")):
        return "UI/메시지 키"
    if token.startswith(("Sprite", "Cinematic", "Effect", "FX_", "BGM", "SE_")):
        return "에셋명"
    if "_" in token and token.split("_")[0] in ("Buff", "Skill", "Item", "Quest", "Mon", "Npc", "Char", "Class"):
        return "데이터 ID"
    if any(s in token for s in ("Type", "Class", "Enum")):
        return "타입/클래스"
    return "기타"


def main():
    if not SCAN_RESULT.exists():
        print(f"[err] Step 1 결과 파일 없음: {SCAN_RESULT}")
        print("      먼저 scripts/scan_identifiers.py 실행 필요")
        return 1

    print("[load] Step 1 GDD scan 결과...")
    scan = json.loads(SCAN_RESULT.read_text(encoding="utf-8"))
    occurrences = scan["occurrences"]
    gdd_ids: set[str] = set()
    for o in occurrences:
        gdd_ids.add(o["id"])
    print(f"  GDD 식별자 (unique): {len(gdd_ids):,}")
    print()

    print("[scan] DataSheet 73 xlsx ...")
    ds_ids, meta = collect_datasheet_identifiers()
    print(f"  파일 {meta['files']} / 시트 {meta['sheets']} / 셀 sample {meta['cells_scanned']:,}")
    print(f"  DataSheet 식별자 (unique): {len(ds_ids):,}")
    print()

    overlap = gdd_ids & ds_ids
    only_gdd = gdd_ids - ds_ids
    only_ds = ds_ids - gdd_ids

    print("=== 결과 ===")
    print(f"교집합 (양쪽 모두):       {len(overlap):>5,}")
    print(f"GDD 에만 있음:           {len(only_gdd):>5,}")
    print(f"DataSheet 에만 있음:     {len(only_ds):>5,}")
    print()
    pct_gdd_in_ds = 100 * len(overlap) / max(len(gdd_ids), 1)
    print(f"GDD 식별자가 DataSheet 에 있을 확률: {pct_gdd_in_ds:.1f}%")
    print(f"  → 95%+ 면 'DataSheet 인덱스만으로 충분' 가설 확정")
    print(f"  → 미만이면 GDD 전용 인덱스도 필요")
    print()

    # only_gdd 카테고리 분포 (얼마나 의미있는 누락인가)
    cat_counter = Counter()
    for t in only_gdd:
        cat_counter[categorize_missing(t)] += 1

    print("--- GDD 에만 있는 식별자 카테고리 분포 ---")
    for cat, cnt in cat_counter.most_common():
        pct = 100 * cnt / max(len(only_gdd), 1)
        print(f"   {cat:<20} {cnt:>5}  ({pct:.1f}%)")
    print()

    print("--- GDD 에만 있는 식별자 sample 50 (= DataSheet 인덱스가 못 잡는 것) ---")
    sample = sorted(only_gdd)[:50]
    for t in sample:
        print(f"   {t}")
    print(f"   ... (총 {len(only_gdd)})")
    print()

    # 자주 묻는 패턴에 영향이 있을 누락? — 대표적인 "쿨타임/제한/배율/레벨" 키워드를 포함한 한글 코멘트와
    # 매칭되는 GDD-only 식별자 추출
    # (간단 휴리스틱: token 자체가 *Cool*/*Time*/*Max* 등을 포함하면 단답형 질문 후보)
    important_pats = ["Cool", "Time", "Max", "Min", "Level", "Price", "Rate", "Limit", "Count", "Delay"]
    important = [t for t in only_gdd if any(p in t for p in important_pats)]
    print(f"--- GDD-only 중 단답형 질문 후보 키워드 포함: {len(important)}건 ---")
    for t in sorted(important)[:30]:
        print(f"   {t}")
    if len(important) > 30:
        print(f"   ... (+{len(important) - 30})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
