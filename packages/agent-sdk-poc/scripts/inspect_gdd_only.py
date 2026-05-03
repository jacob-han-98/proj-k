"""
GDD-only 식별자 1,597 의 실제 등장 라인을 보여줘서
"진짜 파라미터 vs 노이즈" 판정에 visibility 제공.
"""
from __future__ import annotations

import json
import random
import sys
from pathlib import Path

import openpyxl  # noqa: F401  (lint)
import glob

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
SCAN = ROOT / "scripts" / "bench_out" / "identifier_scan_full.json"
DESIGN = Path("D:/ProjectK/Resource/design")


def collect_ds_ids() -> set[str]:
    """DataSheet 식별자 (간단 — 컬럼명 + 파일명/시트명만, 셀 스캔 생략).
    overlap check 와 다르게 더 엄격 (== "선언된 식별자" 만)."""
    ids: set[str] = set()
    import re
    pat_camel = re.compile(r"\b[A-Z][a-z]{1,}(?:[A-Z][a-z0-9]+){1,}\b")
    pat_snake = re.compile(r"\b[A-Z][a-zA-Z0-9]+(?:_[A-Z][a-zA-Z0-9]+)+\b")
    for f in sorted(glob.glob(str(DESIGN / "*.xlsx"))):
        ids.add(Path(f).stem)
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ids.add(sn)
            ws = wb[sn]
            try:
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            if hdr:
                for c in hdr:
                    if c is None:
                        continue
                    s = str(c).strip()
                    for pat in (pat_snake, pat_camel):
                        for m in pat.finditer(s):
                            ids.add(m.group(0))
            # KV 시트 1열 (enum) 도 식별자로 등록
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row is None or row[0] is None:
                    break  # 한 행만 보고 KV 패턴 추정 — 너무 빡셈, 그냥 continue
                v = row[0]
                if isinstance(v, str):
                    s = v.strip()
                    if 4 <= len(s) <= 50:
                        for pat in (pat_snake, pat_camel):
                            for m in pat.finditer(s):
                                ids.add(m.group(0))
    return ids


def main():
    scan = json.loads(SCAN.read_text(encoding="utf-8"))
    occ = scan["occurrences"]

    # 빠른 lookup: id → list of occurrences
    by_id: dict[str, list] = {}
    for o in occ:
        by_id.setdefault(o["id"], []).append(o)
    print(f"GDD 식별자 종류: {len(by_id):,}")

    print("DataSheet 식별자 모으는 중...")
    ds_ids = collect_ds_ids()
    print(f"DataSheet 식별자: {len(ds_ids):,}")

    only_gdd = sorted(set(by_id.keys()) - ds_ids)
    print(f"GDD-only: {len(only_gdd):,}\n")

    # 단답형 후보 키워드 (Cool/Time/Max/Min/Level/Rate/Limit/Count/Delay/Ratio/Price)
    HOT = ["Cool", "Time", "Max", "Min", "Level", "Rate", "Limit", "Count", "Delay", "Ratio", "Price", "Speed", "Damage"]
    hot = [t for t in only_gdd if any(h in t for h in HOT)]
    print(f"단답형 후보 키워드 포함 GDD-only: {len(hot)}\n")

    # 30 random sample, 등장 라인 함께
    random.seed(7)
    sample = random.sample(hot, min(30, len(hot)))

    print("=" * 100)
    print(f"{'식별자':<35}  {'시트':<35}  컨텍스트 (등장 라인)")
    print("=" * 100)
    for tok in sample:
        for o in by_id[tok][:1]:  # 첫 등장만
            sheet = o["sheet"][:34]
            ctx = (o["ctx"] or "").replace("\n", " ")[:80]
            print(f"  {tok:<33}  {sheet:<35}  {ctx}")
    print()
    print(f"※ 컨텍스트가 마크다운 표 행 (`| ... |`) 또는 ` = `, `:` 패턴이면 진짜 파라미터")
    print(f"※ mermaid 다이어그램 (`-->`, `==>`, `[X]`) 이면 노이즈")
    print(f"※ 일반 산문 문장이면 키워드만 매칭된 노이즈")


if __name__ == "__main__":
    main()
