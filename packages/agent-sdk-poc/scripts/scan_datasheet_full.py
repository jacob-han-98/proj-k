"""
DataSheet 전수 스캐너 — LLM 미사용, 모든 셀 검사, 누락 금지.

목적:
  GDD↔DataSheet 비교의 신뢰도를 위해 DataSheet 의 모든 식별자를 빠짐없이 추출.
  이전 overlap check 의 5,000 cells/sheet limit 가 누락의 원인이었음.

추출 정의 (식별자 = 데이터시트에 "선언" 된 토큰):
  1. xlsx 파일명 stem
  2. 시트명 모두
  3. 컬럼명 (1행, 모든 셀)
  4. KV 시트 1열 (enum/Type 모든 행)
  5. domain meta (2행) 의 type=<Enum> 명
  6. Comment/Name 컬럼이 아닌 영문 식별자 셀 — 보수적 매칭

각 토큰마다 "어디서 발견됐나" 메타데이터 보존:
  {token: [{file, sheet, kind: 'filename'|'sheet'|'header'|'kv-key'|'enum'|'cell',
            row, col}]}

Usage:
    python scripts/scan_datasheet_full.py
    → bench_out/datasheet_full_index.json
"""
from __future__ import annotations

import glob
import json
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
DESIGN = Path(os.environ.get("PROJK_DESIGN_ROOT", "D:/ProjectK/Resource/design"))
OUT = ROOT / "scripts" / "bench_out" / "datasheet_full_index.json"

PAT_CAMEL = re.compile(r"\b[A-Z][a-z]{1,}(?:[A-Z][a-z0-9]+){1,}\b")
PAT_SNAKE = re.compile(r"\b[A-Z][a-zA-Z0-9]+(?:_[A-Z][a-zA-Z0-9]+)+\b")
PAT_UPPER_SNAKE = re.compile(r"\b[A-Z]{3,}(?:_[A-Z0-9]{2,})+\b")
STOPWORDS = {"TRUE","FALSE","NULL","NONE","YES","NO","AND","OR","NOT","START","END","BEGIN","STOP","MIN","MAX","AVG","SUM","COUNT",
             "TD","TR","TH","BR","BM","LR","UL","OL","LI","OOXML","PDF","PPT","DOC","ID","UI","UX","FX","OK","IO",
             "PK","PK_","NPC","PC","PVP","PVE","DPS","HP","MP","EXP","SP","AP"}


def is_meaningful(token: str) -> bool:
    if not token or len(token) < 4:
        return False
    if token in STOPWORDS:
        return False
    if token.isupper() and "_" not in token and len(token) <= 5:
        return False
    return True


def extract_ids(s) -> list[str]:
    if s is None:
        return []
    if isinstance(s, (int, float)):
        return []
    s = str(s)
    out = []
    for pat in (PAT_SNAKE, PAT_CAMEL, PAT_UPPER_SNAKE):
        for m in pat.finditer(s):
            t = m.group(0)
            if is_meaningful(t):
                out.append(t)
    # dedup
    seen = set()
    res = []
    for t in out:
        if t not in seen:
            seen.add(t)
            res.append(t)
    return res


def main():
    files = sorted(glob.glob(str(DESIGN / "*.xlsx")))
    print(f"[scan] {len(files)} xlsx 파일")
    print(f"[scan] DESIGN_ROOT = {DESIGN}")
    print()

    # token → [list of provenance dicts]
    occurrences: defaultdict[str, list[dict]] = defaultdict(list)

    t0 = time.time()
    total_cells = 0
    for fi, f in enumerate(files, 1):
        stem = Path(f).stem
        if is_meaningful(stem):
            occurrences[stem].append({"file": stem, "sheet": "", "kind": "filename"})

        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [err] {stem}: {e}")
            continue

        for sn in wb.sheetnames:
            if is_meaningful(sn):
                occurrences[sn].append({"file": stem, "sheet": sn, "kind": "sheet"})
            ws = wb[sn]

            # 헤더 (1행)
            try:
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            for ci, c in enumerate(hdr):
                if c is None:
                    continue
                ids = extract_ids(c)
                for t in ids:
                    occurrences[t].append({"file": stem, "sheet": sn, "kind": "header", "row": 1, "col": ci+1})

            # domain row (2행) — type=<Enum> 등에서 enum 명 추출
            try:
                drow = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            except StopIteration:
                drow = None
            if drow:
                for ci, c in enumerate(drow):
                    if c is None: continue
                    ids = extract_ids(c)
                    for t in ids:
                        occurrences[t].append({"file": stem, "sheet": sn, "kind": "domain-meta", "row": 2, "col": ci+1})

            # 데이터 셀 — **전수 검사** (limit 없음)
            for ri, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if row is None:
                    continue
                for ci, c in enumerate(row):
                    if c is None: continue
                    total_cells += 1
                    if isinstance(c, (int, float)):
                        continue
                    ids = extract_ids(c)
                    for t in ids:
                        # KV 첫 컬럼 (enum/Type) 은 "kv-key" 로 별도 분류
                        kind = "kv-key" if ci == 0 else "cell"
                        occurrences[t].append({"file": stem, "sheet": sn, "kind": kind, "row": ri, "col": ci+1})

        if fi % 10 == 0 or fi == len(files):
            elapsed = time.time() - t0
            print(f"  [{fi:>3}/{len(files)}] {stem:<35}  {len(occurrences):,} unique tokens, {total_cells:,} cells, {elapsed:.1f}s")

    elapsed = time.time() - t0
    print()
    print(f"[scan] DONE. {len(occurrences):,} unique tokens / {total_cells:,} cells / {elapsed:.1f}s")

    # 가장 많이 등장 / kind 별 분포
    kinds: defaultdict[str, int] = defaultdict(int)
    for tok, occs in occurrences.items():
        kinds_for_tok = set(o["kind"] for o in occs)
        for k in kinds_for_tok:
            kinds[k] += 1
    print()
    print("--- 토큰을 발견한 위치 종류 (한 토큰이 여러 kind 에 있을 수 있음) ---")
    for k, v in sorted(kinds.items(), key=lambda x: -x[1]):
        print(f"   {k:<18} {v:,}")

    # 저장 (occurrences 는 큼 — 토큰 + 첫 occurrence 만 보존하면 6MB→1MB)
    compact = {}
    for tok, occs in occurrences.items():
        compact[tok] = {
            "count": len(occs),
            "kinds": sorted(set(o["kind"] for o in occs)),
            "first": occs[0],
        }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "stats": {
            "files": len(files),
            "total_cells": total_cells,
            "unique_tokens": len(occurrences),
            "elapsed_s": round(elapsed, 1),
        },
        "tokens": compact,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print()
    print(f"[saved] {OUT.relative_to(ROOT)}  ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
