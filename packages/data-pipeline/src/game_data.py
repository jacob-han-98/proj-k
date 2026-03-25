"""
game_data.py — 게임 데이터시트 인제스트 + 구조화 쿼리 엔진

게임 클라/서버가 실제 사용하는 Excel 데이터시트를 SQLite로 인제스트하고,
Agent의 tool calling으로 안전하게 쿼리할 수 있는 인터페이스를 제공한다.

데이터시트 스키마 규칙:
  - Row 1: 컬럼 헤더
  - Row 2: 타입 메타데이터 (domain=cs\\ntype=int32\\ndefault=0)
  - Row 3+: 실제 데이터
  - "Info" 시트: 문서 시트 (스킵)
  - "#Common#" 시트: Perforce 내부 (스킵)

사용법:
  from src.game_data import ingest_all, execute_game_query, get_schema_summary

  # 인제스트
  report = ingest_all("/mnt/d/projectk/resource/design")

  # 쿼리
  result = execute_game_query({"action": "query", "table": "MonsterClass",
                                "filters": [{"column": "Type", "op": "=", "value": "Boss"}]})
"""

import hashlib
import json
import logging
import os
import re
import sqlite3
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl

log = logging.getLogger("game_data")

# ── 상수 ──────────────────────────────────────────────────

DEFAULT_DB_PATH = Path.home() / ".qna-poc-gamedata" / "game_data.db"

SKIP_SHEETS = {"Info", "#Common#"}

# Row 2 type → SQL type 매핑
TYPE_MAP = {
    "int32": "INTEGER",
    "int64": "INTEGER",
    "float": "REAL",
    "float32": "REAL",
    "bool": "INTEGER",
    "string": "TEXT",
}

# 쿼리 안전성
ALLOWED_OPS = {"=", "!=", "<", ">", "<=", ">=", "LIKE", "IN", "IS NULL", "IS NOT NULL"}
MAX_LIMIT = 500
QUERY_TIMEOUT_SEC = 5


# ── 데이터 구조 ──────────────────────────────────────────

@dataclass
class ColumnDef:
    name: str
    sql_type: str = "TEXT"
    domain: str = ""       # c, s, cs, sc
    raw_type: str = ""     # int32, string, SkillTypeEnum, ...
    default: str = ""
    is_array: bool = False
    is_enum: bool = False
    enum_name: str = ""


@dataclass
class TableDef:
    table_name: str        # TableAttribute 정식명
    sheet_name: str        # xlsx 시트명
    source_file: str       # xlsx 파일명
    columns: list[ColumnDef] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)


@dataclass
class QueryResult:
    columns: list[str]
    rows: list[list[Any]]
    total_matched: int
    sql: str
    execution_ms: float
    error: str = ""


# ── 타입 메타데이터 파싱 ─────────────────────────────────

def _parse_type_meta(cell_value: Any) -> dict:
    """Row 2 셀의 타입 메타데이터 파싱.

    예: 'domain=cs\\ntype=int32\\ndefault=0' → {'domain': 'cs', 'type': 'int32', 'default': '0'}
    """
    if not cell_value:
        return {}
    text = str(cell_value)
    meta = {}
    for line in text.split("\n"):
        line = line.strip()
        if "=" in line:
            k, _, v = line.partition("=")
            meta[k.strip().lower()] = v.strip()
    return meta


def _meta_to_column_def(col_name: str, meta: dict) -> ColumnDef:
    """타입 메타데이터 → ColumnDef."""
    raw_type = meta.get("type", "string")
    domain = meta.get("domain", "")
    default = meta.get("default", "")
    is_array = meta.get("array", "").upper() == "TRUE"

    # Enum 판별: TYPE_MAP에 없으면 Enum으로 간주
    sql_type = TYPE_MAP.get(raw_type.lower(), "TEXT")
    is_enum = raw_type.lower() not in TYPE_MAP and raw_type != ""

    return ColumnDef(
        name=col_name,
        sql_type="TEXT" if is_array else sql_type,  # array는 TEXT로 저장
        domain=domain,
        raw_type=raw_type,
        default=default,
        is_array=is_array,
        is_enum=is_enum,
        enum_name=raw_type if is_enum else "",
    )


# ── xlsx 파싱 ────────────────────────────────────────────

def parse_xlsx_table(file_path: str | Path, sheet_name: str) -> Optional[TableDef]:
    """xlsx의 특정 시트를 파싱하여 TableDef 반환."""
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(all_rows) < 2:
        return None

    # Row 1: 헤더
    raw_headers = list(all_rows[0])
    # 빈 컬럼 제거 (trailing None)
    while raw_headers and raw_headers[-1] is None:
        raw_headers.pop()

    if not raw_headers:
        return None

    col_count = len(raw_headers)
    headers = [str(h) if h else f"_col{i}" for i, h in enumerate(raw_headers)]

    # Row 2: 타입 메타데이터
    meta_row = list(all_rows[1])[:col_count] if len(all_rows) > 1 else [None] * col_count
    columns = []
    for i, hdr in enumerate(headers):
        meta = _parse_type_meta(meta_row[i] if i < len(meta_row) else None)
        columns.append(_meta_to_column_def(hdr, meta))

    # Row 3+: 데이터
    data_rows = []
    for row in all_rows[2:]:
        vals = list(row)[:col_count]
        # 완전 빈 행 스킵
        if all(v is None for v in vals):
            continue
        # Disable 컬럼이 있고 값이 truthy면 스킵 (비활성 데이터)
        disable_idx = next((i for i, c in enumerate(columns) if c.name == "Disable"), -1)
        if disable_idx >= 0 and disable_idx < len(vals) and vals[disable_idx]:
            continue
        data_rows.append(vals)

    source_file = Path(file_path).name
    return TableDef(
        table_name=sheet_name,  # 기본값; 외부에서 정식명으로 오버라이드
        sheet_name=sheet_name,
        source_file=source_file,
        columns=columns,
        rows=data_rows,
    )


def load_table_attribute(design_dir: str | Path) -> dict[str, dict]:
    """TableAttribute.xlsx에서 정식 테이블명 레지스트리 로드.

    Returns: {table_name: {"cs": "cs", "only_schema": False}}
    """
    file_path = Path(design_dir) / "TableAttribute.xlsx"
    if not file_path.exists():
        log.warning(f"TableAttribute.xlsx not found: {file_path}")
        return {}

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    registry = {}
    for row in rows[1:]:  # 헤더 스킵
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        cs = str(row[1]).strip() if len(row) > 1 and row[1] else "cs"
        only_schema = bool(row[2]) if len(row) > 2 else False
        registry[name] = {"cs": cs, "only_schema": only_schema}

    return registry


def parse_enums(enum_dir: str | Path) -> list[tuple[str, str, int | None, str]]:
    """Enum 디렉토리의 모든 xlsx에서 (enum_type, name, value, comment) 튜플 목록 반환."""
    enum_dir = Path(enum_dir)
    if not enum_dir.exists():
        return []

    results = []
    for xlsx_file in sorted(enum_dir.glob("*.xlsx")):
        enum_type = xlsx_file.stem  # e.g., "SkillTypeEnum"
        wb = openpyxl.load_workbook(str(xlsx_file), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if len(rows) < 2:
            continue

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            value = int(row[1]) if len(row) > 1 and row[1] is not None else None
            comment = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            results.append((enum_type, name, value, comment))

    return results


def _file_hash(file_path: Path) -> str:
    """파일의 SHA256 해시 (변경 감지용)."""
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


# ── SQLite 인제스트 ──────────────────────────────────────

def _sanitize_table_name(name: str) -> str:
    """SQL 안전 테이블명 (알파벳/숫자/밑줄만)."""
    return re.sub(r"[^a-zA-Z0-9_]", "_", name)


def _create_table_sql(table_name: str, columns: list[ColumnDef]) -> str:
    """CREATE TABLE SQL 생성."""
    safe_name = _sanitize_table_name(table_name)
    col_defs = []
    for col in columns:
        safe_col = _sanitize_table_name(col.name)
        col_defs.append(f'  "{safe_col}" {col.sql_type}')
    cols_str = ",\n".join(col_defs)
    return f'CREATE TABLE IF NOT EXISTS "{safe_name}" (\n{cols_str}\n);'


def ingest_all(
    design_dir: str | Path,
    db_path: str | Path = None,
    dry_run: bool = False,
) -> dict:
    """전체 데이터시트를 SQLite로 인제스트.

    Args:
        design_dir: 데이터시트 디렉토리 (xlsx + Enum/)
        db_path: SQLite DB 경로 (기본: ~/.qna-poc-gamedata/game_data.db)
        dry_run: True면 실제 DB 생성 없이 분석만

    Returns:
        인제스트 리포트 dict
    """
    design_dir = Path(design_dir)
    db_path = Path(db_path) if db_path else DEFAULT_DB_PATH

    t_start = time.time()

    # TableAttribute 레지스트리 로드
    registry = load_table_attribute(design_dir)
    registered_names = set(registry.keys())
    log.info(f"TableAttribute: {len(registry)}개 테이블 등록됨")

    # xlsx 파일 스캔
    xlsx_files = sorted(design_dir.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if f.name != "TableAttribute.xlsx"]

    # 시트명 → 정식 테이블명 매핑 빌드
    sheet_to_table: dict[str, str] = {}
    for name in registered_names:
        sheet_to_table[name] = name  # 대부분 시트명 = 정식명

    # Enum 파싱
    enum_dir = design_dir / "Enum"
    enums = parse_enums(enum_dir)
    log.info(f"Enum: {len(enums)}개 값 파싱 ({len(set(e[0] for e in enums))}개 타입)")

    # 모든 xlsx 파싱
    tables: list[TableDef] = []
    skipped_sheets = []
    for xlsx_file in xlsx_files:
        wb = openpyxl.load_workbook(str(xlsx_file), read_only=True, data_only=True)
        sheet_names = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
        wb.close()

        for sheet_name in sheet_names:
            table_def = parse_xlsx_table(xlsx_file, sheet_name)
            if not table_def or not table_def.columns:
                continue

            # 정식 테이블명 매핑
            if sheet_name in registered_names:
                table_def.table_name = sheet_name
            else:
                # 등록 안 된 시트 → 그래도 인제스트 (시트명 그대로)
                skipped_sheets.append(f"{xlsx_file.name}/{sheet_name}")
                table_def.table_name = sheet_name

            tables.append(table_def)

    log.info(f"테이블: {len(tables)}개 파싱 완료 (등록 외 시트: {len(skipped_sheets)}개)")

    if dry_run:
        return {
            "dry_run": True,
            "registered_tables": len(registry),
            "parsed_tables": len(tables),
            "enum_values": len(enums),
            "unregistered_sheets": skipped_sheets[:20],
            "tables": [
                {"name": t.table_name, "file": t.source_file, "sheet": t.sheet_name,
                 "columns": len(t.columns), "rows": len(t.rows)}
                for t in tables
            ],
        }

    # DB 생성
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    try:
        # ── 메타 테이블 ──
        conn.execute("CREATE TABLE _meta (key TEXT PRIMARY KEY, value TEXT)")
        conn.execute("INSERT INTO _meta VALUES (?, ?)",
                     ("ingested_at", time.strftime("%Y-%m-%dT%H:%M:%S")))
        conn.execute("INSERT INTO _meta VALUES (?, ?)",
                     ("design_dir", str(design_dir)))

        # ── 테이블 카탈로그 ──
        conn.execute("""
            CREATE TABLE _table_catalog (
                table_name TEXT PRIMARY KEY,
                source_file TEXT,
                sheet_name TEXT,
                column_count INTEGER,
                row_count INTEGER,
                columns_json TEXT,
                cs TEXT DEFAULT 'cs'
            )
        """)

        # ── Enum 테이블 ──
        conn.execute("""
            CREATE TABLE _enums (
                enum_type TEXT NOT NULL,
                name TEXT NOT NULL,
                value INTEGER,
                comment TEXT DEFAULT '',
                PRIMARY KEY (enum_type, name)
            )
        """)
        if enums:
            conn.executemany(
                "INSERT OR IGNORE INTO _enums (enum_type, name, value, comment) VALUES (?, ?, ?, ?)",
                enums,
            )

        # ── FK 관계 테이블 ──
        conn.execute("""
            CREATE TABLE _fk_relationships (
                source_table TEXT,
                source_column TEXT,
                target_table TEXT,
                target_column TEXT DEFAULT 'Id',
                relation_type TEXT DEFAULT 'id_ref'
            )
        """)

        # ── 데이터 테이블 생성 & 인제스트 ──
        total_rows = 0
        created_tables = []

        for table_def in tables:
            safe_name = _sanitize_table_name(table_def.table_name)

            # CREATE TABLE
            create_sql = _create_table_sql(table_def.table_name, table_def.columns)
            try:
                conn.execute(create_sql)
            except sqlite3.OperationalError as e:
                # 중복 테이블명 (다른 파일에서 같은 시트명)
                log.warning(f"테이블 생성 실패 {safe_name}: {e}")
                continue

            # INSERT 데이터
            if table_def.rows:
                placeholders = ", ".join(["?"] * len(table_def.columns))
                insert_sql = f'INSERT INTO "{safe_name}" VALUES ({placeholders})'
                cleaned_rows = []
                for row in table_def.rows:
                    # 행 길이 맞추기
                    padded = list(row) + [None] * (len(table_def.columns) - len(row))
                    cleaned_rows.append(padded[:len(table_def.columns)])
                try:
                    conn.executemany(insert_sql, cleaned_rows)
                except sqlite3.Error as e:
                    log.warning(f"INSERT 실패 {safe_name}: {e}")
                    continue

            # Id 컬럼 인덱스
            id_cols = [c for c in table_def.columns if c.name == "Id"]
            if id_cols:
                try:
                    conn.execute(f'CREATE INDEX IF NOT EXISTS "idx_{safe_name}_Id" ON "{safe_name}" ("Id")')
                except sqlite3.Error:
                    pass

            # 카탈로그 등록
            columns_json = json.dumps([
                {"name": c.name, "type": c.raw_type, "sql_type": c.sql_type,
                 "domain": c.domain, "is_enum": c.is_enum, "enum_name": c.enum_name,
                 "is_array": c.is_array}
                for c in table_def.columns
            ], ensure_ascii=False)

            cs = registry.get(table_def.table_name, {}).get("cs", "cs")
            conn.execute(
                "INSERT OR REPLACE INTO _table_catalog VALUES (?, ?, ?, ?, ?, ?, ?)",
                (table_def.table_name, table_def.source_file, table_def.sheet_name,
                 len(table_def.columns), len(table_def.rows), columns_json, cs),
            )

            total_rows += len(table_def.rows)
            created_tables.append(table_def.table_name)

            # FK 관계 감지
            all_table_names = {_sanitize_table_name(t.table_name) for t in tables}
            for col in table_def.columns:
                # *Id 패턴
                if col.name.endswith("Id") and col.name != "Id":
                    # 대상 테이블 추정: BuffId → BuffClass, MonsterId → MonsterClass, WorldId → WorldClass
                    base = col.name[:-2]  # "Monster"
                    candidates = [
                        f"{base}Class", base, f"{base}Group",
                    ]
                    for cand in candidates:
                        if cand in registered_names:
                            conn.execute(
                                "INSERT INTO _fk_relationships VALUES (?, ?, ?, ?, ?)",
                                (table_def.table_name, col.name, cand, "Id", "id_ref"),
                            )
                            break

                # Enum 참조
                if col.is_enum and col.enum_name:
                    conn.execute(
                        "INSERT INTO _fk_relationships VALUES (?, ?, ?, ?, ?)",
                        (table_def.table_name, col.name, "_enums", col.enum_name, "enum_ref"),
                    )

        conn.commit()

    finally:
        conn.close()

    elapsed = time.time() - t_start
    report = {
        "db_path": str(db_path),
        "tables_created": len(created_tables),
        "total_rows": total_rows,
        "enum_values": len(enums),
        "enum_types": len(set(e[0] for e in enums)),
        "db_size_mb": round(db_path.stat().st_size / 1024 / 1024, 2),
        "elapsed_sec": round(elapsed, 1),
        "unregistered_sheets": skipped_sheets[:10],
    }
    log.info(f"인제스트 완료: {report['tables_created']}개 테이블, {total_rows}행, "
             f"{report['db_size_mb']}MB, {elapsed:.1f}초")
    return report


# ── 스키마 요약 (Planning LLM 주입용) ────────────────────

def get_schema_summary(db_path: str | Path = None) -> str:
    """Planning LLM에 주입할 컴팩트 스키마 요약 생성."""
    db_path = Path(db_path) if db_path else DEFAULT_DB_PATH
    if not db_path.exists():
        return ""

    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        # 테이블 카탈로그
        tables = conn.execute(
            "SELECT table_name, row_count, columns_json, cs FROM _table_catalog ORDER BY table_name"
        ).fetchall()

        if not tables:
            return ""

        # FK 관계
        fk_rows = conn.execute(
            "SELECT source_table, source_column, target_table, target_column, relation_type "
            "FROM _fk_relationships"
        ).fetchall()
        fk_map: dict[str, list] = {}
        for fk in fk_rows:
            key = (fk["source_table"], fk["source_column"])
            fk_map[key] = fk

        # Enum 타입 목록
        enum_types = conn.execute(
            "SELECT enum_type, COUNT(*) as cnt FROM _enums GROUP BY enum_type ORDER BY enum_type"
        ).fetchall()

        lines = [f"## 게임 데이터 테이블 ({len(tables)}개 테이블)"]
        lines.append("")

        for t in tables:
            cols = json.loads(t["columns_json"])
            col_strs = []
            for c in cols:
                if c["name"] in ("Disable", "Comment"):
                    continue
                s = c["name"]
                if c["is_enum"]:
                    s += f'[{c["enum_name"]}]'
                # FK 표시
                fk = fk_map.get((t["table_name"], c["name"]))
                if fk and fk["relation_type"] == "id_ref":
                    s += f'→{fk["target_table"]}'
                col_strs.append(s)

            cols_str = ", ".join(col_strs[:15])
            if len(col_strs) > 15:
                cols_str += f" (+{len(col_strs)-15}개)"
            lines.append(f"- **{t['table_name']}** ({t['row_count']}행): {cols_str}")

        # Enum 요약
        lines.append("")
        lines.append(f"## Enum ({len(enum_types)}개 타입)")
        for et in enum_types:
            lines.append(f"- {et['enum_type']} ({et['cnt']}개)")

        return "\n".join(lines)

    finally:
        conn.close()


def is_db_ready(db_path: str | Path = None) -> bool:
    """게임 데이터 DB가 사용 가능한지 확인."""
    db_path = Path(db_path) if db_path else DEFAULT_DB_PATH
    if not db_path.exists():
        return False
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        count = conn.execute("SELECT COUNT(*) FROM _table_catalog").fetchone()[0]
        conn.close()
        return count > 0
    except Exception:
        return False


# ── 쿼리 엔진 ───────────────────────────────────────────

def _get_table_columns(conn: sqlite3.Connection, table_name: str) -> dict[str, dict]:
    """테이블의 컬럼 정보 반환."""
    row = conn.execute(
        "SELECT columns_json FROM _table_catalog WHERE table_name = ?", (table_name,)
    ).fetchone()
    if not row:
        return {}
    cols = json.loads(row[0])
    return {c["name"]: c for c in cols}


def _validate_table(conn: sqlite3.Connection, table_name: str) -> str:
    """테이블명 검증. 유효하면 sanitized 이름 반환."""
    row = conn.execute(
        "SELECT table_name FROM _table_catalog WHERE table_name = ?", (table_name,)
    ).fetchone()
    if not row:
        raise ValueError(f"테이블 '{table_name}'이 존재하지 않습니다")
    return _sanitize_table_name(row[0])


def execute_game_query(
    query_spec: dict,
    db_path: str | Path = None,
) -> QueryResult:
    """구조화 쿼리 스펙을 안전한 SQL로 변환하여 실행.

    Args:
        query_spec: {"action": "query"|"list_tables"|"describe"|"lookup_enum",
                     "table": "MonsterClass", "columns": [...], "filters": [...], ...}
        db_path: SQLite DB 경로

    Returns:
        QueryResult
    """
    db_path = Path(db_path) if db_path else DEFAULT_DB_PATH
    if not db_path.exists():
        return QueryResult([], [], 0, "", 0, error="게임 데이터 DB가 없습니다")

    t_start = time.time()
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row

    try:
        action = query_spec.get("action", "query")

        if action == "list_tables":
            return _action_list_tables(conn, t_start)
        elif action == "describe":
            return _action_describe(conn, query_spec, t_start)
        elif action == "lookup_enum":
            return _action_lookup_enum(conn, query_spec, t_start)
        elif action == "query":
            return _action_query(conn, query_spec, t_start)
        else:
            return QueryResult([], [], 0, "", 0, error=f"알 수 없는 action: {action}")
    except Exception as e:
        elapsed = (time.time() - t_start) * 1000
        return QueryResult([], [], 0, "", elapsed, error=str(e))
    finally:
        conn.close()


def _action_list_tables(conn, t_start) -> QueryResult:
    rows = conn.execute(
        "SELECT table_name, source_file, row_count, column_count, cs FROM _table_catalog ORDER BY table_name"
    ).fetchall()
    columns = ["table_name", "source_file", "rows", "columns", "cs"]
    data = [[r[0], r[1], r[2], r[3], r[4]] for r in rows]
    elapsed = (time.time() - t_start) * 1000
    return QueryResult(columns, data, len(data), "SELECT * FROM _table_catalog", elapsed)


def _action_describe(conn, spec, t_start) -> QueryResult:
    table = spec.get("table", "")
    _validate_table(conn, table)
    col_info = _get_table_columns(conn, table)
    columns = ["name", "type", "sql_type", "domain", "is_enum", "enum_name"]
    data = [[c["name"], c["type"], c["sql_type"], c["domain"],
             c.get("is_enum", False), c.get("enum_name", "")]
            for c in col_info.values()]
    elapsed = (time.time() - t_start) * 1000
    return QueryResult(columns, data, len(data), f"DESCRIBE {table}", elapsed)


def _action_lookup_enum(conn, spec, t_start) -> QueryResult:
    enum_type = spec.get("enum_type", spec.get("table", ""))
    if not enum_type:
        return QueryResult([], [], 0, "", 0, error="enum_type 필요")

    rows = conn.execute(
        "SELECT name, value, comment FROM _enums WHERE enum_type = ? ORDER BY value",
        (enum_type,)
    ).fetchall()
    columns = ["name", "value", "comment"]
    data = [[r[0], r[1], r[2]] for r in rows]
    elapsed = (time.time() - t_start) * 1000
    return QueryResult(columns, data, len(data),
                       f"SELECT * FROM _enums WHERE enum_type='{enum_type}'", elapsed)


def _action_query(conn, spec, t_start) -> QueryResult:
    table = spec.get("table", "")
    safe_table = _validate_table(conn, table)
    valid_cols = _get_table_columns(conn, table)

    # SELECT 컬럼
    select_cols = spec.get("columns")
    if select_cols:
        for c in select_cols:
            if c != "*" and c not in valid_cols:
                raise ValueError(f"컬럼 '{c}'가 테이블 '{table}'에 없습니다")
        select_str = ", ".join(f'"{_sanitize_table_name(c)}"' for c in select_cols)
    else:
        select_str = "*"

    sql_parts = [f'SELECT {select_str} FROM "{safe_table}"']
    params = []

    # JOIN
    joins = spec.get("joins", [])
    for j in joins:
        join_table = j.get("table", "")
        safe_join = _validate_table(conn, join_table)
        on_cols = j.get("on", [])
        if len(on_cols) == 2:
            # on: ["MonsterId", "Id"] → safe_table.MonsterId = safe_join.Id
            sql_parts.append(
                f'JOIN "{safe_join}" ON "{safe_table}"."{_sanitize_table_name(on_cols[0])}" '
                f'= "{safe_join}"."{_sanitize_table_name(on_cols[1])}"'
            )

    # WHERE
    filters = spec.get("filters", [])
    if filters:
        where_clauses = []
        for f in filters:
            col = f.get("column", "")
            op = f.get("op", "=").upper()
            val = f.get("value")

            if col not in valid_cols:
                raise ValueError(f"필터 컬럼 '{col}'가 테이블 '{table}'에 없습니다")
            if op not in ALLOWED_OPS:
                raise ValueError(f"허용되지 않는 연산자: {op}")

            safe_col = _sanitize_table_name(col)

            if op in ("IS NULL", "IS NOT NULL"):
                where_clauses.append(f'"{safe_table}"."{safe_col}" {op}')
            elif op == "IN":
                if isinstance(val, list):
                    placeholders = ", ".join(["?"] * len(val))
                    where_clauses.append(f'"{safe_table}"."{safe_col}" IN ({placeholders})')
                    params.extend(val)
                else:
                    where_clauses.append(f'"{safe_table}"."{safe_col}" = ?')
                    params.append(val)
            else:
                where_clauses.append(f'"{safe_table}"."{safe_col}" {op} ?')
                params.append(val)

        sql_parts.append("WHERE " + " AND ".join(where_clauses))

    # GROUP BY + aggregation
    agg = spec.get("aggregation")
    if agg:
        group_by = agg.get("group_by", [])
        measures = agg.get("measures", [])
        if group_by:
            gb_str = ", ".join(f'"{_sanitize_table_name(g)}"' for g in group_by)
            # Rebuild SELECT
            agg_cols = [f'"{_sanitize_table_name(g)}"' for g in group_by]
            for m in measures:
                func = m.get("func", "COUNT").upper()
                if func not in ("COUNT", "SUM", "AVG", "MIN", "MAX"):
                    raise ValueError(f"허용되지 않는 집계 함수: {func}")
                mcol = m.get("column", "*")
                alias = m.get("alias", f"{func}_{mcol}")
                if mcol == "*":
                    agg_cols.append(f'{func}(*) AS "{alias}"')
                else:
                    agg_cols.append(f'{func}("{_sanitize_table_name(mcol)}") AS "{alias}"')
            sql_parts[0] = f'SELECT {", ".join(agg_cols)} FROM "{safe_table}"'
            sql_parts.append(f"GROUP BY {gb_str}")

    # ORDER BY
    order_by = spec.get("order_by", [])
    if order_by:
        ob_parts = []
        for o in order_by:
            col = o.get("column", "")
            direction = o.get("direction", "ASC").upper()
            if direction not in ("ASC", "DESC"):
                direction = "ASC"
            ob_parts.append(f'"{_sanitize_table_name(col)}" {direction}')
        sql_parts.append("ORDER BY " + ", ".join(ob_parts))

    # LIMIT
    limit = min(int(spec.get("limit", 100)), MAX_LIMIT)
    sql_parts.append(f"LIMIT {limit}")

    sql = " ".join(sql_parts)

    # 실행
    try:
        cursor = conn.execute(sql, params)
        result_rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description] if cursor.description else []
        data = [list(r) for r in result_rows]

        # 총 매칭 행수 (LIMIT 없이)
        count_sql = f'SELECT COUNT(*) FROM "{safe_table}"'
        if filters:
            count_parts = [count_sql]
            count_parts.append("WHERE " + " AND ".join(
                [c for c in sql.split("WHERE")[1].split("ORDER BY")[0].split("GROUP BY")[0].strip().split("LIMIT")[0].strip()]
                if "WHERE" in sql else []
            ))
            # 간단하게: 필터가 있으면 반환된 행 수를 total로 사용
        total = len(data)

        elapsed = (time.time() - t_start) * 1000
        return QueryResult(columns, data, total, sql, elapsed)

    except sqlite3.Error as e:
        elapsed = (time.time() - t_start) * 1000
        return QueryResult([], [], 0, sql, elapsed, error=f"SQL 실행 오류: {e}")


# ── 결과 포맷터 ──────────────────────────────────────────

def format_game_data_result(result: QueryResult, max_display: int = 50) -> str:
    """QueryResult를 Markdown 테이블로 변환."""
    if result.error:
        return f"**쿼리 오류**: {result.error}"

    if not result.rows:
        return f"**결과 없음** (쿼리: `{result.sql[:200]}`)"

    lines = []
    lines.append(f"**{result.total_matched}건 조회** ({result.execution_ms:.0f}ms)")
    lines.append("")

    # 헤더
    lines.append("| " + " | ".join(str(c) for c in result.columns) + " |")
    lines.append("| " + " | ".join("---" for _ in result.columns) + " |")

    # 데이터
    display_rows = result.rows[:max_display]
    for row in display_rows:
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                s = str(v)
                if len(s) > 60:
                    s = s[:57] + "..."
                cells.append(s)
        lines.append("| " + " | ".join(cells) + " |")

    if len(result.rows) > max_display:
        lines.append(f"\n*... 외 {len(result.rows) - max_display}건 생략*")

    return "\n".join(lines)


def get_db_path() -> Path:
    """환경변수 또는 기본 경로에서 DB 경로 반환."""
    return Path(os.getenv("GAME_DATA_DB_PATH", str(DEFAULT_DB_PATH)))
