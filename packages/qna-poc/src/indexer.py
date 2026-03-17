"""
indexer.py — content.md 파일들을 청킹 → 임베딩 → ChromaDB 저장

Sources:
    1. xlsx-extractor output (packages/xlsx-extractor/output/*/_final/content.md)
    2. confluence-downloader output (packages/confluence-downloader/output/**/content.md)
       - content_enriched.md 우선 사용 (이미지 설명 포함)

Usage:
    python -m src.indexer                           # 전체 인덱싱 (Excel + Confluence)
    python -m src.indexer --workbook "PK_변신 및 스킬 시스템"  # 단일 Excel 워크북
    python -m src.indexer --source excel            # Excel만 인덱싱
    python -m src.indexer --source confluence       # Confluence만 인덱싱
    python -m src.indexer --stats                   # 통계만 출력
"""

import json
import os
import re
import sys
import time
from pathlib import Path

import chromadb
import requests
from dotenv import load_dotenv

# ── 설정 ──

EXTRACTOR_OUTPUT = Path(__file__).resolve().parent.parent.parent / "xlsx-extractor" / "output"
CONFLUENCE_OUTPUT = Path(__file__).resolve().parent.parent.parent / "confluence-downloader" / "output"
CHROMA_DIR = Path.home() / ".qna-poc-chroma"
COLLECTION_NAME = "project_k"

MAX_CHUNK_TOKENS = 2000  # 청크 최대 토큰 (대략 1000 한글 글자)
MIN_CHUNK_TOKENS = 100   # 너무 작은 청크 방지
APPROX_TOKENS_PER_CHAR = 0.5  # 한국어 대략 추정 (실제는 ~0.4-0.6)
_OCR_CORRECT_ENABLED = False  # --correct-ocr 플래그로 활성화

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

# ── XLSX 소스 경로 (구조적 OOXML 추출용) ──
XLSX_SOURCE_DIRS = [
    Path("D:/ProjectK/Design/3_Base"),
    Path("D:/ProjectK/Design/7_System"),
    Path("D:/ProjectK/Design/9_MileStone"),
    Path("D:/ProjectK/Design/2_Development"),
]


def _find_xlsx_file(workbook: str) -> Path | None:
    """워크북 이름으로 소스 XLSX 파일 찾기."""
    for d in XLSX_SOURCE_DIRS:
        if not d.exists():
            continue
        xlsx_path = d / f"{workbook}.xlsx"
        if xlsx_path.exists():
            return xlsx_path
    return None


def _extract_structured_ooxml(workbook: str, sheet: str, min_len: int = 3) -> str | None:
    """XLSX 파일에서 셀 텍스트를 행/열 위치와 함께 추출.

    반환 형식: 행별로 그룹핑된 셀 데이터.
    예: R5: C3:기본형 | C4:베리1 | C5:베리2
        R6: C3:스켈라 병사 | C4:로바르스 왕 호위병사 | C5:소르브 국경 수비대

    같은 열에 위치한 셀들의 관계를 LLM이 파악할 수 있게 해줌.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    xlsx_path = _find_xlsx_file(workbook)
    if not xlsx_path:
        return None

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        # 정확한 시트명 매칭
        target_sheet = None
        for name in wb.sheetnames:
            if name == sheet:
                target_sheet = name
                break
        if not target_sheet:
            # 공백/언더스코어 차이 허용
            for name in wb.sheetnames:
                if name.replace(" ", "_") == sheet.replace(" ", "_"):
                    target_sheet = name
                    break
        if not target_sheet:
            wb.close()
            return None

        ws = wb[target_sheet]

        from collections import defaultdict
        rows = defaultdict(list)
        cell_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if len(val) >= min_len:
                        rows[cell.row].append((cell.column, val))
                        cell_count += 1
        wb.close()

        if not rows or cell_count == 0:
            return None

        # 행별로 정렬, 포맷팅
        lines = []
        for row_num in sorted(rows):
            cells = sorted(rows[row_num], key=lambda x: x[0])
            cell_texts = [f"C{col}:{text}" for col, text in cells]
            lines.append(f"R{row_num}: " + " | ".join(cell_texts))

        return "\n".join(lines)

    except Exception:
        return None


# ── OCR 보정 (LLM) ──

OCR_CORRECT_SYSTEM = """당신은 OCR 보정 전문가입니다. 게임 기획서를 스크린샷으로 캡처한 뒤 Vision AI로 OCR 변환한 텍스트에 문자 인식 오류가 있습니다.
원본 Excel 셀 데이터(OOXML)를 참고하여 OCR 텍스트를 교정하세요.

규칙:
1. OCR 텍스트의 마크다운 구조(제목, 표, 목록, 들여쓰기, 이미지 링크)를 그대로 유지하세요.
2. 깨진 한글/영문만 OOXML의 정확한 텍스트로 교체하세요.
3. OOXML에 대응하는 원본이 없는 부분(이미지 설명, Mermaid 다이어그램 등)은 OCR 텍스트를 그대로 유지하세요.
4. 교정된 텍스트만 출력하세요. 설명이나 주석은 추가하지 마세요.
5. 표(테이블)의 셀 값이 깨져있으면 OOXML에서 같은 행/열의 값을 찾아 교정하세요."""


def _correct_ocr_section(section_text: str, ooxml_text: str) -> str:
    """OCR 변환 텍스트를 OOXML 원본으로 교정 (Haiku 사용).

    Args:
        section_text: OCR 변환된 마크다운 섹션 텍스트
        ooxml_text: 같은 시트의 OOXML 구조적 셀 데이터

    Returns:
        교정된 텍스트. 실패 시 원본 반환.
    """
    from src.generator import call_bedrock

    # 너무 짧은 섹션은 교정 불필요
    if len(section_text.strip()) < 50:
        return section_text

    # OOXML이 없으면 교정 불가
    if not ooxml_text or len(ooxml_text.strip()) < 10:
        return section_text

    # OOXML 참조는 최대 4000자 (Haiku 입력 절약)
    ooxml_ref = ooxml_text[:4000]

    user_msg = f"""## OCR 변환 텍스트 (교정 대상)

{section_text[:3000]}

## OOXML 원본 셀 데이터 (참고용 — 정확한 문자)

{ooxml_ref}

위 OCR 텍스트를 OOXML을 참고하여 교정하세요. 마크다운 구조는 유지하고 깨진 문자만 수정하세요."""

    try:
        result = call_bedrock(
            messages=[{"role": "user", "content": user_msg}],
            system=OCR_CORRECT_SYSTEM,
            model="claude-haiku-4-5",
            max_tokens=4096,
            temperature=0,
        )
        corrected = result["text"].strip()
        # 교정 결과가 너무 짧으면 (잘림) 원본 유지
        if len(corrected) < len(section_text) * 0.3:
            return section_text
        return corrected
    except Exception as e:
        print(f"    [WARN] OCR correction failed: {e}")
        return section_text


# ── 청킹 ──

def estimate_tokens(text: str) -> int:
    """한국어/영어 혼합 텍스트의 토큰 수 대략 추정."""
    return max(1, int(len(text) * APPROX_TOKENS_PER_CHAR))


def _parse_yaml_frontmatter(lines: list[str]) -> tuple[dict, int]:
    """YAML frontmatter (--- 구분) 파싱. (metadata_dict, content_start_line) 반환."""
    meta = {}
    if not lines or lines[0].strip() != "---":
        return meta, 0
    for i, line in enumerate(lines[1:], start=1):
        if line.strip() == "---":
            return meta, i + 1
        if ":" in line:
            key, _, val = line.partition(":")
            meta[key.strip()] = val.strip().strip('"').strip("'")
    return meta, 0


def parse_content_md(filepath: Path, source_type: str = "excel") -> dict:
    """content.md 파일에서 메타데이터와 섹션을 추출.

    source_type: "excel" | "confluence"
    """
    text = filepath.read_text(encoding="utf-8")
    lines = text.split("\n")

    workbook = ""
    sheet = ""
    content_start = 0

    source_url = ""  # Confluence 원본 페이지 URL 또는 Excel depot 경로

    if source_type == "confluence":
        # Confluence: YAML frontmatter 파싱
        meta, content_start = _parse_yaml_frontmatter(lines)
        title = meta.get("title", filepath.parent.name)
        source_url = meta.get("source", "")
        # Confluence 경로에서 카테고리 추출 (Design/시스템/하위시스템)
        try:
            rel = filepath.parent.relative_to(CONFLUENCE_OUTPUT)
            path_parts = list(rel.parts)
        except ValueError:
            path_parts = [filepath.parent.name]
        workbook = "Confluence"
        sheet = title
        # 카테고리 정보를 workbook에 포함 (검색 시 활용)
        if len(path_parts) > 1:
            workbook = f"Confluence/{'/'.join(path_parts[:-1])}"
    else:
        # Excel: blockquote 메타데이터
        for line in lines[:10]:
            if line.startswith("> 원본:"):
                parts = line.replace("> 원본:", "").strip().split(" / 시트: ")
                workbook = parts[0].strip()
                sheet = parts[1].strip() if len(parts) > 1 else ""
                source_url = f"//main/ProjectK/Design/7_System/{workbook}.xlsx"
                break
        # 메타데이터 블록(--- 구분선 이전) 건너뛰기
        # 주의: --- 구분선은 파일 앞부분(메타데이터 영역)에서만 찾음
        # 본문 중간의 --- (섹션 구분선)을 메타데이터로 오인하는 문제 방지
        for i, line in enumerate(lines[:20]):  # 앞 20줄 이내에서만 탐색
            if line.strip() == "---":
                content_start = i + 1
                break

    # H1 제목 추출
    h1_title = ""
    for line in lines[content_start:]:
        if line.startswith("# ") and not line.startswith("## "):
            h1_title = line[2:].strip()
            break

    # 섹션 분할 (H2 기준)
    sections = []
    current_heading = h1_title or sheet
    current_lines = []
    heading_level = 1

    for line in lines[content_start:]:
        # H2 헤딩 감지
        if line.startswith("## ") and not line.startswith("### "):
            # 이전 섹션 저장
            if current_lines:
                section_text = "\n".join(current_lines).strip()
                if section_text:
                    sections.append({
                        "heading": current_heading,
                        "level": heading_level,
                        "text": section_text,
                    })
            current_heading = line[3:].strip()
            heading_level = 2
            current_lines = []
        else:
            current_lines.append(line)

    # 마지막 섹션
    if current_lines:
        section_text = "\n".join(current_lines).strip()
        if section_text:
            sections.append({
                "heading": current_heading,
                "level": heading_level,
                "text": section_text,
            })

    # ── OOXML 기반 처리: OCR 보정 + 원본 텍스트 추가
    if source_type == "excel":
        # 1차: XLSX 파일에서 구조적 셀 데이터 추출
        structured = _extract_structured_ooxml(workbook, sheet)

        # ── LLM OCR 보정: OOXML 참조로 깨진 한글 교정 (--correct-ocr 플래그)
        if structured and _OCR_CORRECT_ENABLED:
            corrected_count = 0
            for sec in sections:
                # OOXML 섹션 자체나 이미 교정된 섹션은 스킵
                if "OOXML" in sec.get("heading", ""):
                    continue
                original = sec["text"]
                corrected = _correct_ocr_section(original, structured)
                if corrected != original:
                    sec["text"] = corrected
                    corrected_count += 1
            if corrected_count > 0:
                print(f"    OCR corrected {corrected_count} sections")

        # ── OOXML 원본 텍스트를 별도 섹션으로도 추가 (셀 위치 참조용)
        ooxml_added = False
        if structured:
            sections.append({
                "heading": "OOXML 원본 텍스트 (OCR 보정, 셀 위치 포함)",
                "level": 2,
                "text": structured,
            })
            ooxml_added = True
        # 2차 폴백: text_corpus.json (플랫 리스트)
        if not ooxml_added:
            ooxml_path = filepath.parent.parent / "_parse_ooxml_output" / "text_corpus.json"
            if ooxml_path.exists():
                try:
                    ooxml_texts = json.loads(ooxml_path.read_text(encoding="utf-8"))
                    if isinstance(ooxml_texts, list) and ooxml_texts:
                        meaningful = [t for t in ooxml_texts if isinstance(t, str) and len(t.strip()) >= 3]
                        if meaningful:
                            ooxml_section = "\n\n".join(meaningful)
                            sections.append({
                                "heading": "OOXML 원본 텍스트 (OCR 보정)",
                                "level": 2,
                                "text": ooxml_section,
                            })
                except Exception:
                    pass

    return {
        "workbook": workbook,
        "sheet": sheet,
        "h1_title": h1_title,
        "sections": sections,
        "filepath": str(filepath),
        "source_url": source_url,
    }


def chunk_section(section: dict, parent_context: str, max_tokens: int = MAX_CHUNK_TOKENS) -> list[dict]:
    """단일 섹션을 토큰 제한에 맞게 청크로 분할.

    Mermaid 블록은 절대 분할하지 않는다.
    """
    text = section["text"]
    heading = section["heading"]
    prefix = f"[{parent_context}]\n## {heading}\n\n" if parent_context else f"## {heading}\n\n"

    tokens = estimate_tokens(prefix + text)

    # 토큰 제한 이내면 그대로 반환
    if tokens <= max_tokens:
        return [{
            "heading": heading,
            "text": prefix + text,
            "tokens": tokens,
        }]

    # H3 기준으로 하위 분할
    chunks = []
    h3_parts = re.split(r'(?=^### )', text, flags=re.MULTILINE)

    current_chunk_lines = []
    current_chunk_heading = heading

    for part in h3_parts:
        part = part.strip()
        if not part:
            continue

        # H3 제목 추출
        h3_match = re.match(r'^### (.+)', part)
        sub_heading = h3_match.group(1).strip() if h3_match else ""

        part_with_prefix = prefix + part if not current_chunk_lines else part
        part_tokens = estimate_tokens(part_with_prefix)

        if current_chunk_lines:
            combined = "\n\n".join(current_chunk_lines + [part])
            combined_tokens = estimate_tokens(prefix + combined)

            if combined_tokens <= max_tokens:
                current_chunk_lines.append(part)
                continue

            # 현재 청크 저장
            chunk_text = prefix + "\n\n".join(current_chunk_lines)
            chunks.append({
                "heading": current_chunk_heading,
                "text": chunk_text,
                "tokens": estimate_tokens(chunk_text),
            })
            current_chunk_lines = [part]
            current_chunk_heading = f"{heading} > {sub_heading}" if sub_heading else heading
        else:
            current_chunk_lines = [part]
            current_chunk_heading = f"{heading} > {sub_heading}" if sub_heading else heading

    # 마지막 청크
    if current_chunk_lines:
        chunk_text = prefix + "\n\n".join(current_chunk_lines)
        chunks.append({
            "heading": current_chunk_heading,
            "text": chunk_text,
            "tokens": estimate_tokens(chunk_text),
        })

    return chunks if chunks else [{"heading": heading, "text": prefix + text, "tokens": tokens}]


def chunk_file(filepath: Path, source_type: str = "excel") -> list[dict]:
    """content.md 파일 하나를 청크 리스트로 변환."""
    parsed = parse_content_md(filepath, source_type=source_type)
    workbook = parsed["workbook"]
    sheet = parsed["sheet"]
    source_url = parsed.get("source_url", "")
    parent_context = f"{workbook} / {sheet}" if workbook else sheet

    all_chunks = []
    small_chunks = []  # MIN_CHUNK_TOKENS 미달 청크 임시 보관
    for section in parsed["sections"]:
        chunks = chunk_section(section, parent_context)
        for chunk in chunks:
            # 메타데이터 부착
            chunk["workbook"] = workbook
            chunk["sheet"] = sheet
            chunk["section_path"] = chunk["heading"]
            chunk["has_mermaid"] = "```mermaid" in chunk["text"]
            chunk["has_table"] = bool(re.search(r'\|.*\|.*\|', chunk["text"]))
            chunk["has_images"] = "![" in chunk["text"]
            chunk["source_path"] = str(filepath)
            chunk["source_url"] = source_url

            # 최소 토큰 필터
            if chunk["tokens"] >= MIN_CHUNK_TOKENS:
                all_chunks.append(chunk)
            else:
                small_chunks.append(chunk)

    # 모든 청크가 MIN_CHUNK_TOKENS 미달이면 페이지 전체를 1개 청크로 합쳐서 보존
    # (짧은 페이지가 인덱싱에서 누락되는 문제 방지)
    if not all_chunks and small_chunks:
        merged_text = "\n\n".join(c["text"] for c in small_chunks)
        merged = {
            "heading": small_chunks[0]["heading"],
            "text": merged_text,
            "tokens": estimate_tokens(merged_text),
            "workbook": workbook,
            "sheet": sheet,
            "section_path": small_chunks[0]["heading"],
            "has_mermaid": "```mermaid" in merged_text,
            "has_table": bool(re.search(r'\|.*\|.*\|', merged_text)),
            "has_images": "![" in merged_text,
            "source_path": str(filepath),
            "source_url": source_url,
        }
        all_chunks.append(merged)

    return all_chunks


def discover_content_files(workbook_filter: str = None) -> list[Path]:
    """output/ 아래의 모든 _final/content.md 파일을 탐색."""
    if not EXTRACTOR_OUTPUT.exists():
        print(f"[ERROR] Output directory not found: {EXTRACTOR_OUTPUT}")
        return []

    files = []
    for content_md in sorted(EXTRACTOR_OUTPUT.rglob("_final/content.md")):
        if workbook_filter:
            # 워크북 디렉토리명으로 필터
            workbook_dir = content_md.parent.parent.parent.name
            if workbook_filter not in workbook_dir:
                continue
        files.append(content_md)

    return files


def discover_confluence_files(title_filter: str = None) -> list[Path]:
    """Confluence output 아래의 content_enriched.md (또는 content.md) 파일을 탐색.

    content_enriched.md가 있으면 우선 사용 (이미지 설명 포함).
    """
    if not CONFLUENCE_OUTPUT.exists():
        print(f"[WARN] Confluence output directory not found: {CONFLUENCE_OUTPUT}")
        return []

    files = []
    # 모든 content.md를 찾고, enriched 버전이 있으면 대체
    for content_md in sorted(CONFLUENCE_OUTPUT.rglob("content.md")):
        # _manifest.json, _tree.md 등 최상위 파일 건너뛰기
        if content_md.parent == CONFLUENCE_OUTPUT:
            continue

        if title_filter:
            if title_filter.lower() not in content_md.parent.name.lower():
                continue

        # content_enriched.md 우선 사용
        enriched = content_md.parent / "content_enriched.md"
        target = enriched if enriched.exists() else content_md

        # 빈 파일 건너뛰기
        try:
            if target.stat().st_size < 50:
                continue
        except OSError:
            continue

        files.append(target)

    return files


# ── 임베딩 ──

def _embed_single(text: str, url: str, headers: dict) -> list[float]:
    """단일 텍스트 임베딩. 병렬 호출용."""
    if not text or not text.strip():
        return [0.0] * 1024
    truncated = text[:8000]
    body = {"inputText": truncated, "dimensions": 1024, "normalize": True}
    try:
        resp = requests.post(url, json=body, headers=headers, timeout=30)
        if resp.status_code != 200:
            print(f"[WARN] Embedding API error {resp.status_code}: {resp.text[:200]}")
            return [0.0] * 1024
        return resp.json()["embedding"]
    except Exception as e:
        print(f"[WARN] Embedding failed: {e}")
        return [0.0] * 1024


def embed_texts(texts: list[str], batch_size: int = 20, max_workers: int = 10) -> list[list[float]]:
    """Bedrock Titan Embeddings v2로 텍스트 목록을 병렬 임베딩."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    token = os.environ.get("AWS_BEARER_TOKEN_BEDROCK")
    if not token:
        raise RuntimeError("AWS_BEARER_TOKEN_BEDROCK 환경변수 미설정")
    region = os.environ.get("AWS_REGION", "us-east-1")
    model_id = os.environ.get("EMBEDDING_MODEL", "amazon.titan-embed-text-v2:0")

    url = f"https://bedrock-runtime.{region}.amazonaws.com/model/{model_id}/invoke"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}",
    }

    # 결과를 인덱스 순서대로 저장
    all_embeddings = [None] * len(texts)
    completed = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 배치 단위로 제출 (진행률 표시용)
        for batch_start in range(0, len(texts), batch_size):
            batch_end = min(batch_start + batch_size, len(texts))
            futures = {}
            for idx in range(batch_start, batch_end):
                future = executor.submit(_embed_single, texts[idx], url, headers)
                futures[future] = idx

            for future in as_completed(futures):
                idx = futures[future]
                all_embeddings[idx] = future.result()
                completed += 1

            if batch_end < len(texts):
                print(f"  Embedded {batch_end}/{len(texts)} chunks...")

    # None 방어 (혹시 빠진 항목)
    for i, emb in enumerate(all_embeddings):
        if emb is None:
            all_embeddings[i] = [0.0] * 1024

    return all_embeddings


# ── ChromaDB 저장 ──

def index_chunks(chunks: list[dict], reset: bool = False):
    """청크를 ChromaDB에 저장."""
    # 1. 임베딩 먼저 (오래 걸림)
    texts = [c["text"] for c in chunks]
    print(f"\n[INFO] Embedding {len(texts)} chunks...")

    t_start = time.time()
    embeddings = embed_texts(texts)
    t_embed = time.time() - t_start
    print(f"[INFO] Embedding complete: {t_embed:.1f}s ({len(texts) / max(t_embed, 0.1):.1f} chunks/s)")

    # 2. ChromaDB 컬렉션 생성 (임베딩 완료 후 — stale handle 방지)
    client = chromadb.PersistentClient(path=str(CHROMA_DIR))

    if reset:
        try:
            client.delete_collection(COLLECTION_NAME)
            print(f"[INFO] Collection '{COLLECTION_NAME}' deleted.")
        except Exception:
            pass

    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"},
    )

    # 기존 데이터 확인
    existing = collection.count()
    if existing > 0 and not reset:
        print(f"[INFO] Collection has {existing} existing documents. Adding new ones...")

    # 3. 메타데이터 구성
    ids = []
    metadatas = []
    documents = []

    for i, chunk in enumerate(chunks):
        chunk_id = f"{chunk['workbook']}__{chunk['sheet']}__{chunk['section_path']}__{i}"
        chunk_id = re.sub(r'[^a-zA-Z0-9가-힣_]', '_', chunk_id)[:512]

        ids.append(chunk_id)
        documents.append(chunk["text"])
        metadatas.append({
            "workbook": chunk["workbook"],
            "sheet": chunk["sheet"],
            "section_path": chunk["section_path"],
            "has_mermaid": chunk["has_mermaid"],
            "has_table": chunk["has_table"],
            "has_images": chunk["has_images"],
            "tokens": chunk["tokens"],
            "source_path": chunk["source_path"],
            "source_url": chunk.get("source_url", ""),
        })

    # 4. 배치 추가 (ChromaDB 제한: 5461개씩)
    batch_size = 5000
    for i in range(0, len(ids), batch_size):
        end = min(i + batch_size, len(ids))
        collection.add(
            ids=ids[i:end],
            embeddings=embeddings[i:end],
            documents=documents[i:end],
            metadatas=metadatas[i:end],
        )

    print(f"[INFO] Indexed {len(ids)} chunks into ChromaDB.")
    print(f"[INFO] Total collection size: {collection.count()}")

    return collection


# ── 통계 ──

def print_stats(chunks: list[dict]):
    """청크 통계 출력."""
    total_tokens = sum(c["tokens"] for c in chunks)
    total_chars = sum(len(c["text"]) for c in chunks)
    workbooks = set(c["workbook"] for c in chunks)
    sheets = set(f"{c['workbook']}/{c['sheet']}" for c in chunks)

    mermaid_count = sum(1 for c in chunks if c.get("has_mermaid"))
    table_count = sum(1 for c in chunks if c.get("has_table"))
    image_count = sum(1 for c in chunks if c.get("has_images"))

    # 소스별 통계
    excel_chunks = [c for c in chunks if not c["workbook"].startswith("Confluence")]
    conf_chunks = [c for c in chunks if c["workbook"].startswith("Confluence")]

    print(f"\n{'='*60}")
    print(f"Chunking Statistics")
    print(f"{'='*60}")
    print(f"  Total chunks:    {len(chunks)}")
    print(f"  Total tokens:    {total_tokens:,} (estimated)")
    print(f"  Total chars:     {total_chars:,}")
    print(f"  Avg tokens/chunk:{total_tokens / max(len(chunks), 1):.0f}")
    if excel_chunks:
        print(f"  Excel chunks:    {len(excel_chunks)} ({sum(c['tokens'] for c in excel_chunks):,} tokens)")
    if conf_chunks:
        print(f"  Confluence chunks:{len(conf_chunks)} ({sum(c['tokens'] for c in conf_chunks):,} tokens)")
    print(f"  Workbooks:       {len(workbooks)}")
    print(f"  Sheets:          {len(sheets)}")
    print(f"  With Mermaid:    {mermaid_count}")
    print(f"  With Tables:     {table_count}")
    print(f"  With Images:     {image_count}")
    print(f"{'='*60}")

    # 토큰 분포
    token_ranges = [(0, 200), (200, 500), (500, 1000), (1000, 1500), (1500, 2000), (2000, 99999)]
    print(f"\n  Token Distribution:")
    for low, high in token_ranges:
        count = sum(1 for c in chunks if low <= c["tokens"] < high)
        label = f"  {low}-{high}" if high < 99999 else f"  {low}+"
        bar = "#" * (count * 40 // max(len(chunks), 1))
        print(f"    {label:>10}: {count:>4} {bar}")


# ── 메인 ──

def main():
    import argparse

    parser = argparse.ArgumentParser(description="QnA PoC Indexer")
    parser.add_argument("--workbook", help="특정 워크북만 인덱싱 (부분 일치)")
    parser.add_argument("--source", choices=["all", "excel", "confluence"], default="all",
                        help="데이터 소스 선택 (default: all)")
    parser.add_argument("--stats", action="store_true", help="통계만 출력 (인덱싱 안 함)")
    parser.add_argument("--reset", action="store_true", help="기존 인덱스 삭제 후 재생성")
    parser.add_argument("--correct-ocr", action="store_true",
                        help="LLM(Haiku)으로 OCR 텍스트를 OOXML 기반 교정 (비용: ~$2)")
    args = parser.parse_args()

    # OCR 보정 플래그 활성화
    global _OCR_CORRECT_ENABLED
    if args.correct_ocr:
        _OCR_CORRECT_ENABLED = True
        print("[INFO] OCR correction enabled (Haiku LLM)")


    all_chunks = []

    # Excel 소스
    if args.source in ("all", "excel"):
        excel_files = discover_content_files(args.workbook)
        print(f"[INFO] Excel: {len(excel_files)} content.md files found.")
        t_start = time.time()
        for f in excel_files:
            chunks = chunk_file(f, source_type="excel")
            all_chunks.extend(chunks)
        t_excel = time.time() - t_start
        print(f"[INFO] Excel chunking: {t_excel:.1f}s, {len(all_chunks)} chunks.")

    # Confluence 소스
    if args.source in ("all", "confluence"):
        confluence_files = discover_confluence_files(args.workbook if args.source == "confluence" else None)
        print(f"[INFO] Confluence: {len(confluence_files)} files found "
              f"({sum(1 for f in confluence_files if f.name == 'content_enriched.md')} enriched).")
        t_start = time.time()
        n_before = len(all_chunks)
        for f in confluence_files:
            chunks = chunk_file(f, source_type="confluence")
            all_chunks.extend(chunks)
        t_conf = time.time() - t_start
        print(f"[INFO] Confluence chunking: {t_conf:.1f}s, {len(all_chunks) - n_before} chunks.")

    if not all_chunks:
        print("[ERROR] No content files found.")
        sys.exit(1)

    print(f"\n[INFO] Total: {len(all_chunks)} chunks from {args.source} source(s).")

    # 통계 출력
    print_stats(all_chunks)

    if args.stats:
        return

    # 인덱싱
    index_chunks(all_chunks, reset=args.reset)

    # 임베딩 비용 추정
    total_tokens = sum(c["tokens"] for c in all_chunks)
    cost = total_tokens * 0.00002 / 1000  # Titan v2 가격 대략
    print(f"\n[INFO] Estimated embedding cost: ${cost:.4f}")


if __name__ == "__main__":
    main()
