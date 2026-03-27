#!/usr/bin/env python3
"""
parse_ooxml.py - Stage 3: OOXML 커넥터 검증 + Vision AI Mermaid 보정

Vision AI가 생성한 Mermaid 플로우차트를 OOXML drawing XML의 커넥터 데이터로
검증하고 보정한다. 결과는 _parse_ooxml_output/ 에 저장한다.

입력: _vision_output/merged.md + 원본 XLSX
출력: _parse_ooxml_output/
  - merged.md         (Mermaid 보정이 있을 때만 생성)
  - parse_meta.json   (검증 결과 메타데이터, 항상 생성)
  - grade_colors.json (등급 색상 매핑, 데이터 있을 때만)
  - text_corpus.json  (OOXML 텍스트 코퍼스, OCR 교정용)

사용법:
    python parse_ooxml.py <xlsx_path> <output_base_dir>
    python parse_ooxml.py <xlsx_path> <output_base_dir> --sheet "시트이름"
"""

import sys
import os
import re
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


# ── OOXML 네임스페이스 ──

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    's': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}


# ── OOXML 파싱 ──

def get_sheet_drawing_map(xlsx_path):
    """시트명 -> drawing XML 파일명 매핑을 반환한다."""
    mapping = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # workbook.xml에서 시트 목록
        wb = ET.fromstring(z.read('xl/workbook.xml'))
        sheets = []
        for s in wb.findall('.//s:sheet', NS):
            rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheets.append({'name': s.get('name'), 'rId': rid})

        # workbook.xml.rels에서 rId -> sheet 파일
        rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid_to_target = {rel.get('Id'): rel.get('Target') for rel in rels}

        for s in sheets:
            sheet_file = rid_to_target.get(s['rId'], '')
            sheet_filename = sheet_file.split('/')[-1]
            sheet_rels_path = f'xl/worksheets/_rels/{sheet_filename}.rels'
            try:
                srels = ET.fromstring(z.read(sheet_rels_path))
                for rel in srels:
                    if 'drawing' in rel.get('Type', ''):
                        drawing_target = rel.get('Target', '')
                        # ../drawings/drawing1.xml -> xl/drawings/drawing1.xml
                        drawing_path = 'xl/drawings/' + drawing_target.split('/')[-1]
                        mapping[s['name']] = drawing_path
            except (KeyError, ET.ParseError):
                pass

    return mapping


def _parse_single_shape(sp, xdr_ns, a_ns):
    """단일 도형(sp) 요소에서 ID, 텍스트, 도형 타입을 추출한다."""
    cNvPr = sp.find(f'.//{{{xdr_ns}}}nvSpPr/{{{xdr_ns}}}cNvPr')
    if cNvPr is None:
        return None
    sid = cNvPr.get('id')
    # 텍스트 추출
    texts = []
    for t in sp.iter(f'{{{a_ns}}}t'):
        if t.text:
            texts.append(t.text)
    text = ''.join(texts).strip()
    text = re.sub(r'\s+', ' ', text)
    # 도형 타입
    geom = sp.find(f'.//{{{a_ns}}}prstGeom')
    geom_type = geom.get('prst', '?') if geom is not None else '?'
    return {
        'id': sid,
        'text': text,
        'geom': geom_type,
    }


def extract_shapes_and_connectors(xlsx_path, drawing_path):
    """drawing XML에서 도형과 커넥터를 추출한다.

    그룹 도형(grpSp) 처리:
    Excel에서 마름모+TextBox 등을 그룹으로 묶는 경우가 많다.
    커넥터는 그룹 내 도형 도형(예: 마름모 id=325)에 연결되지만,
    텍스트는 같은 그룹의 TextBox(id=326)에 있다.
    → 그룹 내 텍스트 없는 도형에 같은 그룹의 텍스트를 병합한다.
    """
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        tree = ET.fromstring(z.read(drawing_path))

    xdr_ns = NS['xdr']
    a_ns = NS['a']

    shapes = {}
    group_id_map = {}  # grpSp ID → 병합 텍스트 (커넥터가 그룹 ID를 참조할 때 대비)

    # 1. 그룹 도형(grpSp) 처리 — 그룹 내부 도형들을 수집하고 텍스트 병합
    for grp in tree.iter(f'{{{xdr_ns}}}grpSp'):
        grp_nvPr = grp.find(f'{{{xdr_ns}}}nvGrpSpPr/{{{xdr_ns}}}cNvPr')
        grp_id = grp_nvPr.get('id') if grp_nvPr is not None else None

        inner_shapes = []
        for sp in grp.findall(f'{{{xdr_ns}}}sp'):
            parsed = _parse_single_shape(sp, xdr_ns, a_ns)
            if parsed:
                inner_shapes.append(parsed)

        if not inner_shapes:
            continue

        # 그룹 내 텍스트 병합: 도형+TextBox 쌍(2~3개)만 처리
        # 큰 그룹(플로우차트 전체)의 텍스트를 합치면 안 됨
        group_texts = [s['text'] for s in inner_shapes if s['text']]
        merged_text = ' '.join(group_texts) if group_texts else ''

        if len(inner_shapes) <= 3 and merged_text:
            # 소규모 그룹: 도형+TextBox 쌍 → 텍스트 없는 도형에 병합
            for s in inner_shapes:
                if not s['text']:
                    s['text'] = merged_text
                    s['text_from_group'] = True
                shapes[s['id']] = s

            # 그룹 ID도 shapes에 등록
            if grp_id and grp_id not in shapes:
                main_geom = '?'
                for s in inner_shapes:
                    if s.get('geom') not in ('rect', '?'):
                        main_geom = s['geom']
                        break
                shapes[grp_id] = {
                    'id': grp_id,
                    'text': merged_text,
                    'geom': main_geom,
                    'is_group': True,
                    'text_from_group': True,
                }
                group_id_map[grp_id] = merged_text
        else:
            # 대규모 그룹: 텍스트 병합 없이 개별 도형만 등록
            for s in inner_shapes:
                shapes[s['id']] = s

    # 2. 비그룹 도형(sp) 수집 — 이미 그룹에서 추출된 것은 건너뜀
    for anchor in tree:
        for sp in anchor.findall(f'{{{xdr_ns}}}sp'):
            parsed = _parse_single_shape(sp, xdr_ns, a_ns)
            if parsed and parsed['id'] not in shapes:
                shapes[parsed['id']] = parsed

    # 3. 커넥터(cxnSp) 수집
    connectors = []
    for cxn in tree.iter(f'{{{xdr_ns}}}cxnSp'):
        cNvPr = cxn.find(f'.//{{{xdr_ns}}}nvCxnSpPr/{{{xdr_ns}}}cNvPr')
        cNvCxnSpPr = cxn.find(f'.//{{{xdr_ns}}}nvCxnSpPr/{{{xdr_ns}}}cNvCxnSpPr')

        stCxn = cNvCxnSpPr.find(f'{{{a_ns}}}stCxn') if cNvCxnSpPr is not None else None
        endCxn = cNvCxnSpPr.find(f'{{{a_ns}}}endCxn') if cNvCxnSpPr is not None else None

        st_id = stCxn.get('id') if stCxn is not None else None
        end_id = endCxn.get('id') if endCxn is not None else None

        if st_id and end_id:
            connectors.append({
                'start_id': st_id,
                'end_id': end_id,
                'start_text': shapes.get(st_id, {}).get('text', ''),
                'end_text': shapes.get(end_id, {}).get('text', ''),
            })

    # 4. 깨진 참조 복구 — 커넥터가 참조하지만 shapes에 없는 ID
    #    Excel에서 도형 수정 시 orphan reference가 발생할 수 있음.
    #    전략: 같은 플로우차트 그룹 내에서, orphan과 이웃이 완전히 겹치지 않고
    #    병합 시 in+out 모두 2개 이상이 되는 노드를 찾아 매핑.
    orphan_ids = set()
    for c in connectors:
        if c['start_id'] not in shapes:
            orphan_ids.add(c['start_id'])
        if c['end_id'] not in shapes:
            orphan_ids.add(c['end_id'])

    if orphan_ids:
        # orphan을 임시로 shapes에 등록 (그룹핑에 참여하도록)
        for oid in orphan_ids:
            shapes[oid] = {'id': oid, 'text': '', 'geom': '?', 'unresolved': True}

        # in/out 엣지 맵 구축
        in_edges = {}
        out_edges = {}
        for c in connectors:
            out_edges.setdefault(c['start_id'], set()).add(c['end_id'])
            in_edges.setdefault(c['end_id'], set()).add(c['start_id'])

        for oid in orphan_ids:
            o_in = in_edges.get(oid, set())
            o_out = out_edges.get(oid, set())
            o_neighbors = o_in | o_out

            best = None
            best_score = -1
            for sid, s in shapes.items():
                if sid == oid or s.get('unresolved') or not s.get('text'):
                    continue
                s_in = in_edges.get(sid, set())
                s_out = out_edges.get(sid, set())
                s_neighbors = s_in | s_out

                # 이웃이 완전히 겹치지 않아야 함 (보완적 관계)
                if o_neighbors & s_neighbors:
                    continue

                # 병합 시 in, out 모두 2개 이상이어야 함 (플로우차트 분기점)
                merged_in = len(o_in | s_in)
                merged_out = len(o_out | s_out)
                if merged_in >= 2 and merged_out >= 2:
                    score = merged_in + merged_out
                    if score > best_score:
                        best_score = score
                        best = sid

            if best:
                src = shapes[best]
                resolved_text = src['text']

                # 같은 텍스트를 가진 다른 노드가 커넥터에 이미 존재하면
                # orphan을 그 노드로 리다이렉트 (중복 노드 제거)
                redirect_to = None
                for sid, s in shapes.items():
                    if sid != oid and sid != best and not s.get('unresolved') \
                       and s.get('text') == resolved_text \
                       and sid in (in_edges.keys() | out_edges.keys()):
                        redirect_to = sid
                        break

                if redirect_to:
                    # orphan의 커넥터를 기존 노드로 리다이렉트
                    redirect_shape = shapes[redirect_to]
                    for c in connectors:
                        if c['start_id'] == oid:
                            c['start_id'] = redirect_to
                            c['start_text'] = redirect_shape['text']
                        if c['end_id'] == oid:
                            c['end_id'] = redirect_to
                            c['end_text'] = redirect_shape['text']
                    # orphan은 shapes에서 제거
                    del shapes[oid]
                else:
                    # 리다이렉트 대상 없으면 텍스트만 복사
                    shapes[oid] = {
                        'id': oid,
                        'text': resolved_text,
                        'geom': src.get('geom', '?'),
                        'resolved_from': best,
                        'text_from_group': True,
                    }
                    for c in connectors:
                        if c['start_id'] == oid:
                            c['start_text'] = resolved_text
                        if c['end_id'] == oid:
                            c['end_text'] = resolved_text

    return shapes, connectors


# ── 플로우차트 그룹핑 ──

def group_flowcharts(shapes, connectors):
    """커넥터로 연결된 도형들을 플로우차트 그룹으로 묶는다.
    하나의 drawing에 여러 플로우차트가 있을 수 있다."""
    # 커넥터에 참여하는 도형 ID 수집
    connected_ids = set()
    adj = {}  # adjacency list (undirected)
    for c in connectors:
        sid, eid = c['start_id'], c['end_id']
        connected_ids.add(sid)
        connected_ids.add(eid)
        adj.setdefault(sid, set()).add(eid)
        adj.setdefault(eid, set()).add(sid)

    # BFS로 연결 컴포넌트 찾기
    visited = set()
    groups = []
    for start_id in connected_ids:
        if start_id in visited:
            continue
        group = set()
        queue = [start_id]
        while queue:
            nid = queue.pop(0)
            if nid in visited:
                continue
            visited.add(nid)
            group.add(nid)
            for neighbor in adj.get(nid, []):
                if neighbor not in visited:
                    queue.append(neighbor)
        groups.append(group)

    # 주석/라벨 도형 필터링: 플로우차트 노드가 아닌 텍스트 전용 도형은 제외
    # (커넥터에 참여하는 도형만 남김)
    result = []
    for group_ids in groups:
        group_shapes = {sid: shapes[sid] for sid in group_ids if sid in shapes}
        group_connectors = [
            c for c in connectors
            if c['start_id'] in group_ids and c['end_id'] in group_ids
        ]
        if len(group_connectors) >= 2:  # 최소 2개 커넥터가 있어야 플로우차트
            result.append({
                'shapes': group_shapes,
                'connectors': group_connectors,
            })

    # 크기 순 정렬 (가장 큰 그룹부터)
    result.sort(key=lambda g: len(g['connectors']), reverse=True)
    return result


# ── Mermaid 파싱 ──

def extract_mermaid_blocks(md_text):
    """Markdown에서 mermaid 코드 블록을 추출한다."""
    pattern = re.compile(r'```mermaid\n(.*?)```', re.DOTALL)
    blocks = []
    for m in pattern.finditer(md_text):
        blocks.append({
            'code': m.group(1).strip(),
            'start': m.start(),
            'end': m.end(),
            'full_match': m.group(0),
        })
    return blocks


def parse_mermaid_edges(mermaid_code):
    """Mermaid 코드에서 엣지(연결)를 추출한다."""
    edges = []
    # 노드 정의 부분(괄호+텍스트)을 건너뛰고 화살표만 추적
    # A((시작)) --> B["텍스트"], C -->|No| D 등
    edge_pattern = re.compile(
        r'(\w+)'                                    # source node id
        r'(?:\s*[\[\(\{][^\n]*?[\]\)\}])*'          # skip node definition
        r'\s*(?:-->|---|-\.->|==>)'                  # arrow types
        r'\s*(?:\|([^|]*)\|)?'                       # optional |label|
        r'\s*(\w+)'                                  # target node id
    )
    # -- label --> 패턴
    label_arrow_pattern = re.compile(
        r'(\w+)'
        r'(?:\s*[\[\(\{][^\n]*?[\]\)\}])*'
        r'\s+--\s+([^-]+?)\s+-->\s*'
        r'(\w+)'
    )
    for line in mermaid_code.split('\n'):
        line = line.strip()
        if line.startswith('%') or line.startswith('flowchart') or not line:
            continue
        # -- label --> 먼저 시도
        for m in label_arrow_pattern.finditer(line):
            edges.append({
                'source': m.group(1),
                'target': m.group(3),
                'label': m.group(2).strip(),
            })
        # 일반 화살표
        for m in edge_pattern.finditer(line):
            src, lbl, tgt = m.group(1), (m.group(2) or '').strip(), m.group(3)
            # 중복 방지
            if not any(e['source'] == src and e['target'] == tgt for e in edges):
                edges.append({'source': src, 'target': tgt, 'label': lbl})
    return edges


def parse_mermaid_nodes(mermaid_code):
    """Mermaid 코드에서 노드 ID와 텍스트를 추출한다."""
    nodes = {}
    # 다양한 괄호 패턴: A["text"], A(["text"]), A{{"text"}}, A((text)) 등
    # 내부 따옴표/특수문자가 있을 수 있으므로 가장 바깥 괄호 기준으로 추출
    node_pattern = re.compile(
        r'(\w+)\s*'                               # node id
        r'([\[\(\{]{1,3})\s*'                     # opening brackets (1-3개)
        r'(.*?)\s*'                               # text (greedy하지 않게)
        r'[\]\)\}]{1,3}'                          # closing brackets
    )
    for line in mermaid_code.split('\n'):
        line = line.strip()
        if line.startswith('%') or line.startswith('flowchart') or not line:
            continue
        for m in node_pattern.finditer(line):
            nid = m.group(1)
            text = m.group(3).strip()
            # 따옴표 제거
            text = re.sub(r'^["\']|["\']$', '', text)
            # HTML <br> 태그 제거
            text = re.sub(r'<br\s*/?>', ' ', text)
            text = re.sub(r'&quot;', '"', text)
            text = re.sub(r'\s+', ' ', text).strip()
            if nid not in nodes and text:
                nodes[nid] = text
    return nodes


# ── 텍스트 매칭 ──

def normalize_text(text):
    """텍스트를 정규화하여 비교 가능하게 만든다."""
    text = re.sub(r'[\s\n\r]+', '', text)  # 모든 공백 제거
    text = re.sub(r'["\'""]', '', text)     # 따옴표 제거
    return text.lower()


def find_best_match(mermaid_text, ooxml_shapes):
    """Mermaid 노드 텍스트와 가장 잘 매칭되는 OOXML 도형을 찾는다."""
    norm_m = normalize_text(mermaid_text)
    if not norm_m:
        return None

    best_match = None
    best_score = 0

    for sid, shape in ooxml_shapes.items():
        norm_s = normalize_text(shape['text'])
        if not norm_s:
            continue
        # 정확 일치
        if norm_m == norm_s:
            return sid
        # 부분 일치 (짧은 쪽이 긴 쪽에 포함)
        if norm_m in norm_s or norm_s in norm_m:
            score = min(len(norm_m), len(norm_s)) / max(len(norm_m), len(norm_s))
            if score > best_score:
                best_score = score
                best_match = sid

    if best_score > 0.5:
        return best_match
    return None


# ── 핵심: Mermaid 보정 ──

def verify_and_correct_mermaid(mermaid_code, ooxml_group):
    """OOXML 커넥터로 Mermaid 엣지를 검증하고 보정한다.

    Returns: dict with corrections, missing_edges, extra_edges
    """
    ooxml_shapes = ooxml_group['shapes']
    ooxml_connectors = ooxml_group['connectors']

    # Mermaid 노드/엣지 파싱
    m_nodes = parse_mermaid_nodes(mermaid_code)
    m_edges = parse_mermaid_edges(mermaid_code)

    # Mermaid 노드 ID -> OOXML shape ID 매핑
    mermaid_to_ooxml = {}
    for mid, mtext in m_nodes.items():
        ooxml_id = find_best_match(mtext, ooxml_shapes)
        if ooxml_id:
            mermaid_to_ooxml[mid] = ooxml_id

    # 텍스트 없는 diamond 노드 매핑 (구조 기반)
    # OOXML에서 텍스트가 없는 diamond가 있고, Mermaid에서 {decision} 노드가 미매핑이면
    # 연결 구조(이웃 노드)를 비교하여 매칭
    unmapped_diamonds = [
        sid for sid, s in ooxml_shapes.items()
        if s.get('geom') == 'diamond' and not s.get('text')
        and sid not in mermaid_to_ooxml.values()
    ]
    unmapped_mermaid_decisions = [
        mid for mid, mtext in m_nodes.items()
        if mid not in mermaid_to_ooxml
    ]
    # Mermaid에서 {} 구문으로 정의된 decision 노드 찾기
    decision_re = re.compile(r'(\w+)\s*\{')
    mermaid_decision_ids = set(m.group(1) for m in decision_re.finditer(mermaid_code))
    unmapped_decisions = [mid for mid in unmapped_mermaid_decisions if mid in mermaid_decision_ids]

    if unmapped_diamonds and unmapped_decisions:
        # 각 diamond의 OOXML 이웃 노드 vs 각 decision의 Mermaid 이웃 노드 비교
        for diamond_id in unmapped_diamonds:
            diamond_neighbors = set()
            for c in ooxml_connectors:
                if c['start_id'] == diamond_id:
                    diamond_neighbors.add(c['end_id'])
                elif c['end_id'] == diamond_id:
                    diamond_neighbors.add(c['start_id'])

            best_mid = None
            best_overlap = 0
            for mid in unmapped_decisions:
                # Mermaid에서 이 decision 노드의 이웃
                m_neighbors_ooxml = set()
                for e in m_edges:
                    if e['source'] == mid:
                        oid = mermaid_to_ooxml.get(e['target'])
                        if oid:
                            m_neighbors_ooxml.add(oid)
                    elif e['target'] == mid:
                        oid = mermaid_to_ooxml.get(e['source'])
                        if oid:
                            m_neighbors_ooxml.add(oid)
                overlap = len(diamond_neighbors & m_neighbors_ooxml)
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_mid = mid
            if best_mid and best_overlap >= 2:
                mermaid_to_ooxml[best_mid] = diamond_id

    # 역매핑: OOXML ID -> Mermaid ID
    ooxml_to_mermaid = {v: k for k, v in mermaid_to_ooxml.items()}

    # OOXML 커넥터를 엣지 셋으로 변환
    ooxml_edges = set()
    for c in ooxml_connectors:
        ooxml_edges.add((c['start_id'], c['end_id']))

    # Mermaid 엣지를 OOXML ID 기준으로 변환
    mermaid_edges_ooxml = set()
    for e in m_edges:
        s_ooxml = mermaid_to_ooxml.get(e['source'])
        t_ooxml = mermaid_to_ooxml.get(e['target'])
        if s_ooxml and t_ooxml:
            mermaid_edges_ooxml.add((s_ooxml, t_ooxml))

    # 비교
    missing = ooxml_edges - mermaid_edges_ooxml  # OOXML에는 있지만 Mermaid에 없음
    extra = mermaid_edges_ooxml - ooxml_edges     # Mermaid에는 있지만 OOXML에 없음

    # 누락된 엣지를 Mermaid에 추가할 수 있는 형태로 변환
    missing_edges = []
    for st, en in missing:
        st_text = ooxml_shapes.get(st, {}).get('text', f'id={st}')
        en_text = ooxml_shapes.get(en, {}).get('text', f'id={en}')
        st_mid = ooxml_to_mermaid.get(st)
        en_mid = ooxml_to_mermaid.get(en)
        missing_edges.append({
            'ooxml_start': st, 'ooxml_end': en,
            'start_text': st_text, 'end_text': en_text,
            'mermaid_start': st_mid, 'mermaid_end': en_mid,
        })

    extra_edges = []
    for st, en in extra:
        st_text = ooxml_shapes.get(st, {}).get('text', f'id={st}')
        en_text = ooxml_shapes.get(en, {}).get('text', f'id={en}')
        extra_edges.append({
            'ooxml_start': st, 'ooxml_end': en,
            'start_text': st_text, 'end_text': en_text,
        })

    return {
        'node_mapping': {mid: {'ooxml_id': oid, 'mermaid_text': m_nodes.get(mid, ''),
                                'ooxml_text': ooxml_shapes.get(oid, {}).get('text', '')}
                         for mid, oid in mermaid_to_ooxml.items()},
        'missing_edges': missing_edges,
        'extra_edges': extra_edges,
        'ooxml_edge_count': len(ooxml_edges),
        'mermaid_edge_count': len(m_edges),
        'match_count': len(ooxml_edges & mermaid_edges_ooxml),
    }


def apply_corrections(mermaid_code, corrections):
    """누락된 엣지를 추가하고, 오판된(extra) 엣지를 제거한다.

    Returns: (corrected_code, added_list, removed_list)
    """
    added = []
    removed = []
    lines = mermaid_code.split('\n')

    # 1. 오판된 엣지 제거 (OOXML에 없지만 Mermaid에 있는 엣지)
    # extra_edges의 mermaid ID를 알아야 함 → node_mapping 역참조
    if corrections.get('extra_edges'):
        node_map = corrections.get('node_mapping', {})
        # ooxml_id -> mermaid_id 역매핑
        ooxml_to_mid = {}
        for mid, info in node_map.items():
            ooxml_to_mid[info['ooxml_id']] = mid

        for extra in corrections['extra_edges']:
            st_ooxml = extra['ooxml_start']
            en_ooxml = extra['ooxml_end']
            st_mid = ooxml_to_mid.get(st_ooxml)
            en_mid = ooxml_to_mid.get(en_ooxml)
            if not st_mid or not en_mid:
                continue

            # 해당 엣지를 포함하는 Mermaid 줄을 찾아 제거
            new_lines = []
            edge_removed = False
            for line in lines:
                stripped = line.strip()
                # 엣지 패턴: "A --> B" 또는 "A -->|label| B"
                edge_re = re.compile(
                    rf'^{re.escape(st_mid)}\b'
                    r'(?:\s*[\[\(\{].*?[\]\)\}])*'
                    r'\s*(?:-->|---|-\.->|==>)'
                    r'(?:\s*\|[^|]*\|)?'
                    rf'\s*{re.escape(en_mid)}\b'
                )
                if edge_re.search(stripped) and not edge_removed:
                    edge_removed = True
                    removed.append(f"{extra['start_text']} -> {extra['end_text']}")
                    continue  # 이 줄 제거
                new_lines.append(line)
            lines = new_lines

    # 2. 누락된 엣지 추가
    for edge in corrections.get('missing_edges', []):
        st_mid = edge['mermaid_start']
        en_mid = edge['mermaid_end']
        if st_mid and en_mid:
            new_line = f'    {st_mid} --> {en_mid}'
            lines.append(new_line)
            added.append(f"{edge['start_text']} -> {edge['end_text']}")

    return '\n'.join(lines), added, removed


# ── 전체 매칭: Mermaid 블록 ↔ OOXML 그룹 ──

def match_mermaid_to_ooxml(mermaid_blocks, ooxml_groups, all_shapes):
    """각 Mermaid 블록을 가장 적합한 OOXML 플로우차트 그룹에 매칭한다."""
    matches = []

    for block in mermaid_blocks:
        m_nodes = parse_mermaid_nodes(block['code'])
        best_group = None
        best_overlap = 0

        for gi, group in enumerate(ooxml_groups):
            overlap = 0
            for mtext in m_nodes.values():
                if find_best_match(mtext, group['shapes']):
                    overlap += 1
            if overlap > best_overlap:
                best_overlap = overlap
                best_group = gi

        if best_group is not None and best_overlap >= 2:
            matches.append({
                'block': block,
                'group_index': best_group,
                'overlap': best_overlap,
            })

    return matches


# ── MD 파일 보정 ──

def correct_md_file(md_path, ooxml_groups, all_shapes):
    """MD 파일의 Mermaid 블록들을 OOXML 데이터로 보정한다."""
    with open(md_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    mermaid_blocks = extract_mermaid_blocks(md_text)
    if not mermaid_blocks:
        return {'path': md_path, 'corrections': [], 'message': 'No mermaid blocks found'}

    matches = match_mermaid_to_ooxml(mermaid_blocks, ooxml_groups, all_shapes)

    all_corrections = []
    corrected_text = md_text

    # 뒤에서부터 교체 (offset 유지)
    for match in sorted(matches, key=lambda m: m['block']['start'], reverse=True):
        block = match['block']
        group = ooxml_groups[match['group_index']]

        result = verify_and_correct_mermaid(block['code'], group)

        if result['missing_edges'] or result['extra_edges']:
            corrected_code, added, removed = apply_corrections(block['code'], result)
            new_block = f"```mermaid\n{corrected_code}\n```"
            corrected_text = (
                corrected_text[:block['start']] +
                new_block +
                corrected_text[block['end']:]
            )
            all_corrections.append({
                'group_index': match['group_index'],
                'overlap': match['overlap'],
                'added_edges': added,
                'removed_edges': removed,
                'result': result,
            })
        else:
            all_corrections.append({
                'group_index': match['group_index'],
                'overlap': match['overlap'],
                'added_edges': [],
                'removed_edges': [],
                'result': result,
                'message': 'All edges verified - no corrections needed',
            })

    has_corrections = any(c.get('added_edges') or c.get('removed_edges') for c in all_corrections)

    return {
        'original_text': md_text,
        'corrected_text': corrected_text if has_corrections else None,
        'has_corrections': has_corrections,
        'corrections': all_corrections,
    }


# ── 메인 ──

def process_sheet(xlsx_path, sheet_dir, sheet_name):
    """한 시트의 Vision 결과를 OOXML로 검증/보정한다.

    입력: sheet_dir/_vision_output/merged.md + 원본 XLSX
    출력: sheet_dir/_parse_ooxml_output/ (중간 결과)
    """
    vision_output_dir = os.path.join(sheet_dir, '_vision_output')
    parse_output_dir = os.path.join(sheet_dir, '_parse_ooxml_output')

    # 1. Vision merged.md 확인
    merged_path = os.path.join(vision_output_dir, 'merged.md')
    if not os.path.exists(merged_path):
        print(f"  [parse_ooxml] {sheet_name}: merged.md not found - skip")
        return None

    import time as _time
    t_total = _time.time()

    # 2. 출력 디렉토리 생성
    os.makedirs(parse_output_dir, exist_ok=True)

    # 3. 시트 -> drawing 매핑
    sheet_drawing_map = get_sheet_drawing_map(xlsx_path)
    drawing_path = sheet_drawing_map.get(sheet_name)

    has_mermaid_corrections = False
    mermaid_result = None
    shapes_count = 0
    connectors_count = 0
    groups_count = 0

    t_ooxml = _time.time()
    if drawing_path:
        # 4. OOXML 도형/커넥터 추출
        shapes, connectors = extract_shapes_and_connectors(xlsx_path, drawing_path)
        shapes_count = len(shapes)
        connectors_count = len(connectors)
        print(f"  [parse_ooxml] {sheet_name}: {shapes_count} shapes, {connectors_count} connectors from {drawing_path}")

        # 4.5. 도형/커넥터 raw 데이터 저장
        if shapes:
            shapes_path = os.path.join(parse_output_dir, 'shapes.json')
            with open(shapes_path, 'w', encoding='utf-8') as f:
                json.dump(shapes, f, ensure_ascii=False, indent=2)
        if connectors:
            connectors_path = os.path.join(parse_output_dir, 'connectors.json')
            with open(connectors_path, 'w', encoding='utf-8') as f:
                json.dump(connectors, f, ensure_ascii=False, indent=2)

        if connectors:
            # 5. 플로우차트 그룹핑
            groups = group_flowcharts(shapes, connectors)
            groups_count = len(groups)
            print(f"  [parse_ooxml] {sheet_name}: {groups_count} flowchart groups detected")
            for gi, g in enumerate(groups):
                node_texts = [s['text'] for s in g['shapes'].values() if s['text']][:5]
                print(f"    group[{gi}]: {len(g['shapes'])} shapes, {len(g['connectors'])} connectors - {node_texts}")

            # 6. Mermaid 보정 (in-memory, 원본 수정 안 함)
            mermaid_result = correct_md_file(merged_path, groups, shapes)

            # 7. 결과 보고
            for ci, corr in enumerate(mermaid_result.get('corrections', [])):
                res = corr.get('result', {})
                added = corr.get('added_edges', [])
                msg = corr.get('message', '')
                print(f"    mermaid[{ci}]: match={res.get('match_count', 0)}/{res.get('ooxml_edge_count', 0)} ooxml edges, "
                      f"missing={len(res.get('missing_edges', []))}, extra={len(res.get('extra_edges', []))}")
                if added:
                    for a in added:
                        print(f"      + ADDED: {a}")
                if msg:
                    print(f"      {msg}")

            # 8. 보정된 merged.md 저장 (변경이 있을 때만)
            has_mermaid_corrections = mermaid_result.get('has_corrections', False)
            if has_mermaid_corrections:
                corrected_md_path = os.path.join(parse_output_dir, 'merged.md')
                with open(corrected_md_path, 'w', encoding='utf-8') as f:
                    f.write(mermaid_result['corrected_text'])
                print(f"  [parse_ooxml] {sheet_name}: corrected merged.md saved → _parse_ooxml_output/")
            else:
                print(f"  [parse_ooxml] {sheet_name}: no mermaid corrections needed")
        else:
            print(f"  [parse_ooxml] {sheet_name}: no connectors - skip mermaid verification")
    else:
        print(f"  [parse_ooxml] {sheet_name}: no drawing XML found - skip mermaid verification")
    t_ooxml_done = _time.time()

    # 9. 등급 색상 추출
    t_colors = _time.time()
    grade_colors = extract_grade_colors(xlsx_path, sheet_name)
    if grade_colors:
        gc_path = os.path.join(parse_output_dir, 'grade_colors.json')
        with open(gc_path, 'w', encoding='utf-8') as f:
            json.dump(grade_colors, f, ensure_ascii=False, indent=2)
        print(f"  [parse_ooxml] {sheet_name}: {len(grade_colors)} grade colors extracted")

    t_colors_done = _time.time()

    # 10. OOXML 텍스트 코퍼스 추출 (OCR 교정용)
    t_corpus = _time.time()
    text_corpus = extract_ooxml_text_corpus(xlsx_path, sheet_name)
    if text_corpus:
        tc_path = os.path.join(parse_output_dir, 'text_corpus.json')
        with open(tc_path, 'w', encoding='utf-8') as f:
            json.dump(text_corpus, f, ensure_ascii=False, indent=2)
        print(f"  [parse_ooxml] {sheet_name}: {len(text_corpus)} text fragments extracted")

    t_corpus_done = _time.time()

    # 11. 셀 데이터 그리드 추출 (기조 데이터 테이블 합성용)
    t_grid = _time.time()
    cell_data_grid = extract_cell_data_grid(xlsx_path, sheet_name)
    if cell_data_grid and cell_data_grid.get('cells'):
        grid_path = os.path.join(parse_output_dir, 'cell_data_grid.json')
        with open(grid_path, 'w', encoding='utf-8') as f:
            json.dump(cell_data_grid, f, ensure_ascii=False, indent=2, default=str)
        print(f"  [parse_ooxml] {sheet_name}: {len(cell_data_grid['cells'])} cells extracted to cell_data_grid.json")
    t_grid_done = _time.time()

    t_elapsed = t_grid_done - t_total

    # 타이밍 로그
    timing = {
        'total_s': round(t_elapsed, 2),
        'ooxml_mermaid_s': round(t_ooxml_done - t_ooxml, 2),
        'grade_colors_s': round(t_colors_done - t_colors, 2),
        'text_corpus_s': round(t_corpus_done - t_corpus, 2),
        'cell_data_grid_s': round(t_grid_done - t_grid, 2),
    }
    print(f"  [parse_ooxml] {sheet_name}: {t_elapsed:.1f}s total "
          f"(ooxml={timing['ooxml_mermaid_s']:.1f}s, colors={timing['grade_colors_s']:.1f}s, "
          f"corpus={timing['text_corpus_s']:.1f}s, grid={timing['cell_data_grid_s']:.1f}s)")

    # 12. 메타데이터 저장 (항상)
    meta = {
        'sheet_name': sheet_name,
        'drawing_path': drawing_path,
        'shapes_count': shapes_count,
        'connectors_count': connectors_count,
        'flowchart_groups': groups_count,
        'has_mermaid_corrections': has_mermaid_corrections,
        'has_grade_colors': bool(grade_colors),
        'has_text_corpus': bool(text_corpus),
        'has_cell_data_grid': bool(cell_data_grid and cell_data_grid.get('cells')),
        'corrections': mermaid_result.get('corrections', []) if mermaid_result else [],
        'timing': timing,
    }
    meta_path = os.path.join(parse_output_dir, 'parse_meta.json')
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2, default=str)

    return {
        'has_mermaid_corrections': has_mermaid_corrections,
        'has_grade_colors': bool(grade_colors),
        'has_text_corpus': bool(text_corpus),
        'has_cell_data_grid': bool(cell_data_grid and cell_data_grid.get('cells')),
        'corrections': mermaid_result.get('corrections', []) if mermaid_result else [],
    }


# ── OOXML 셀 색상 추출 ──

# 표준 Office 테마 색상 (기본 테마 기준, theme1.xml 없으면 이 값 사용)
_DEFAULT_THEME_COLORS = {
    0: 'FFFFFF',  # lt1 (배경1)
    1: '000000',  # dk1 (텍스트1)
    2: 'E7E6E6',  # lt2 (배경2)
    3: '44546A',  # dk2 (텍스트2)
    4: '4472C4',  # accent1
    5: 'ED7D31',  # accent2
    6: 'A5A5A5',  # accent3
    7: 'FFC000',  # accent4
    8: '5B9BD5',  # accent5
    9: '70AD47',  # accent6
}


def _resolve_theme_color(theme_idx, tint=0.0, theme_colors=None):
    """테마 색상 인덱스 + tint를 RGB hex 문자열로 변환한다."""
    colors = theme_colors or _DEFAULT_THEME_COLORS
    base_hex = colors.get(theme_idx, 'FFFFFF')
    r = int(base_hex[0:2], 16)
    g = int(base_hex[2:4], 16)
    b = int(base_hex[4:6], 16)
    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    r, g, b = min(255, max(0, r)), min(255, max(0, g)), min(255, max(0, b))
    return f'{r:02X}{g:02X}{b:02X}'


def extract_cell_colors(xlsx_path, sheet_name):
    """시트에서 셀 배경색(fill)이 있는 셀들의 좌표-RGB 매핑을 반환한다.

    Returns:
        dict: {(row, col): '#RRGGBB', ...}  (1-based row/col)
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    result = {}
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill is None or fill.patternType is None or fill.patternType == 'none':
                continue
            fg = fill.fgColor
            if fg is None:
                continue
            rgb = None
            try:
                if fg.rgb and isinstance(fg.rgb, str) and fg.rgb != '00000000':
                    rgb = fg.rgb[-6:]  # strip 'FF' alpha prefix
            except (TypeError, AttributeError):
                pass
            if rgb is None and fg.theme is not None and isinstance(fg.theme, int):
                tint = fg.tint if fg.tint else 0.0
                rgb = _resolve_theme_color(fg.theme, tint)
            if rgb:
                result[(cell.row, cell.column)] = f'#{rgb}'
    wb.close()
    return result


def extract_ooxml_text_corpus(xlsx_path, sheet_name):
    """시트의 모든 텍스트를 OOXML에서 추출한다 (셀 + 도형).

    Vision AI OCR 오타 교정의 ground truth로 사용.

    Returns:
        list[str]: 줄 단위로 분리된 텍스트 조각 목록
    """
    fragments = set()

    # 1. 셀 텍스트 (openpyxl)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val = cell.value.strip()
                        if val:
                            # 멀티라인 셀은 줄 단위로 분리
                            for line in val.split('\n'):
                                line = line.strip()
                                if len(line) >= 3:  # 너무 짧은 건 제외
                                    fragments.add(line)
        wb.close()
    except Exception:
        pass

    # 2. 도형 텍스트 (drawing XML)
    try:
        sheet_drawing = get_sheet_drawing_map(xlsx_path)
        drawing_path = sheet_drawing.get(sheet_name)
        if drawing_path:
            with zipfile.ZipFile(xlsx_path, 'r') as z:
                tree = ET.fromstring(z.read(drawing_path))
            a_ns = NS['a']
            xdr_ns = NS['xdr']
            for sp in tree.iter(f'{{{xdr_ns}}}sp'):
                texts = []
                for t in sp.iter(f'{{{a_ns}}}t'):
                    if t.text:
                        texts.append(t.text)
                joined = ''.join(texts).strip()
                if joined:
                    # 도형 텍스트도 줄 단위 분리
                    for line in joined.split('\n'):
                        line = line.strip()
                        if len(line) >= 3:
                            fragments.add(line)
                    # 전체 텍스트도 추가 (긴 구절 매칭용)
                    if len(joined) >= 5:
                        fragments.add(joined)
    except Exception:
        pass

    return list(fragments)


def extract_cell_data_grid(xlsx_path, sheet_name):
    """시트의 모든 셀 값을 위치+타입 정보와 함께 추출한다.

    기조 데이터 테이블 합성 시 openpyxl의 정확한 값을 Vision 구조 힌트와
    결합하기 위해 사용. Vision OCR은 숫자 인식 오류가 있으므로 이 데이터가
    ground truth 역할을 한다.

    Returns:
        dict: {
            "cells": [{"row": 5, "col": 3, "value": 27000, "display": "27,000",
                        "type": "number", "format": "#,##0"}, ...],
            "dimensions": {"min_row": 1, "max_row": 40, "min_col": 1, "max_col": 18},
            "merged_cells": [{"min_row": 1, "min_col": 1, "max_row": 1, "max_col": 3}, ...]
        }
    """
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]

        cells = []
        min_row = min_col = float('inf')
        max_row = max_col = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                r, c = cell.row, cell.column
                min_row = min(min_row, r)
                max_row = max(max_row, r)
                min_col = min(min_col, c)
                max_col = max(max_col, c)

                raw_value = cell.value
                number_format = cell.number_format or 'General'

                # 타입 판별
                if isinstance(raw_value, bool):
                    cell_type = 'boolean'
                    display = str(raw_value)
                elif isinstance(raw_value, (int, float)):
                    cell_type = 'number'
                    # 퍼센트 서식 감지
                    if '%' in number_format or '0.0%' in number_format:
                        cell_type = 'percent'
                        display = f"{raw_value * 100:.1f}%" if isinstance(raw_value, float) and raw_value < 1 else str(raw_value)
                    else:
                        # 정수면 정수로 표시
                        if isinstance(raw_value, float) and raw_value == int(raw_value):
                            display = f"{int(raw_value):,}"
                            raw_value = int(raw_value)
                        elif isinstance(raw_value, int):
                            display = f"{raw_value:,}"
                        else:
                            display = str(raw_value)
                elif isinstance(raw_value, str):
                    cell_type = 'string'
                    display = raw_value.strip()
                    if not display:
                        continue
                else:
                    # datetime 등 기타 타입
                    cell_type = 'string'
                    display = str(raw_value)

                cells.append({
                    'row': r,
                    'col': c,
                    'value': raw_value,
                    'display': display,
                    'type': cell_type,
                    'format': number_format,
                })

        # 병합 셀 정보 수집 + 값 복제
        merged = []
        cell_map = {(c['row'], c['col']): c for c in cells}  # (row, col) → cell

        for mc in ws.merged_cells.ranges:
            merged.append({
                'min_row': mc.min_row, 'min_col': mc.min_col,
                'max_row': mc.max_row, 'max_col': mc.max_col,
            })

            # 좌상단 셀의 값을 병합 범위 내 빈 셀에 복제
            src = cell_map.get((mc.min_row, mc.min_col))
            if src and src['display']:
                for r in range(mc.min_row, mc.max_row + 1):
                    for c in range(mc.min_col, mc.max_col + 1):
                        if (r, c) != (mc.min_row, mc.min_col) and (r, c) not in cell_map:
                            new_cell = {
                                'row': r,
                                'col': c,
                                'value': src['value'],
                                'display': src['display'],
                                'type': src['type'],
                                'format': src['format'],
                            }
                            cells.append(new_cell)
                            cell_map[(r, c)] = new_cell

        wb.close()

        if not cells:
            return None

        return {
            'cells': cells,
            'dimensions': {
                'min_row': min_row if min_row != float('inf') else 1,
                'max_row': max_row,
                'min_col': min_col if min_col != float('inf') else 1,
                'max_col': max_col,
            },
            'merged_cells': merged,
        }

    except Exception as e:
        print(f"  [parse_ooxml] extract_cell_data_grid error: {e}")
        return None


def extract_grade_colors(xlsx_path, sheet_name):
    """등급 테이블에서 등급명 → 색상 hex 매핑을 추출한다.

    등급 테이블의 패턴: 한글명 열 + 고유 색상 열 (배경색으로 표시)

    Returns:
        dict: {'신화': '#C00000', '전설': '#FFFF00', ...}
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    grade_names = {'에픽', '신화', '전설', '영웅', '희귀', '고급', '일반',
                   'Epic', 'Myth', 'Legendary', 'Unique', 'Rare', 'Uncommon', 'Common'}
    result = {}

    def _get_cell_rgb(cell):
        """셀의 배경색 RGB를 추출. 없으면 None."""
        fill = cell.fill
        if fill is None or fill.patternType is None or fill.patternType == 'none':
            return None
        fg = fill.fgColor
        if fg is None:
            return None
        rgb = None
        try:
            if fg.rgb and isinstance(fg.rgb, str) and fg.rgb != '00000000':
                rgb = fg.rgb[-6:]
        except (TypeError, AttributeError):
            pass
        if rgb is None and fg.theme is not None and isinstance(fg.theme, int):
            tint = fg.tint if fg.tint else 0.0
            rgb = _resolve_theme_color(fg.theme, tint)
        if rgb:
            r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            if r > 200 and g > 200 and b > 200:
                return None  # 너무 밝으면 skip
            if r < 30 and g < 30 and b < 30:
                return None  # 너무 어두우면 skip
            return rgb
        return None

    for row in ws.iter_rows(max_row=100, max_col=15):
        for cell in row:
            if cell.value and str(cell.value).strip() in grade_names:
                grade = str(cell.value).strip()
                if grade in result:
                    continue  # 첫 번째 발견된 매핑을 유지
                # 먼저 셀 자체의 배경색 확인 (등급명이 색상 셀인 경우)
                own_rgb = _get_cell_rgb(cell)
                if own_rgb:
                    result[grade] = f'#{own_rgb}'
                    continue
                # 인접 셀 탐색 (색상 전용 열, 보통 1~3칸 뒤)
                for offset in range(1, 6):
                    col = cell.column + offset
                    if col > ws.max_column:
                        break
                    color_cell = ws.cell(row=cell.row, column=col)
                    # 인접 셀이 다른 등급명이면 중단 (다른 테이블 레이아웃)
                    if color_cell.value and str(color_cell.value).strip() in grade_names:
                        break
                    rgb = _get_cell_rgb(color_cell)
                    if rgb:
                        result[grade] = f'#{rgb}'
                        break
    wb.close()
    return result


_RGB_TO_COLOR_NAME = {
    'C00000': '적색',   'FF0000': '빨간색',
    'FFFF00': '노란색', 'FFC000': '주황색',
    '7030A0': '보라색',
    '00B0F0': '파란색', '0070C0': '파란색',
    '92D050': '초록색', '00B050': '초록색',
    'BFBFBF': '회색',   'A5A5A5': '회색',   '808080': '회색',
    'FFFFFF': '흰색',
}


def rgb_to_color_name(hex_rgb):
    """#RRGGBB → 한글 색상명 (근사 매칭)"""
    hex6 = hex_rgb.lstrip('#').upper()
    if hex6 in _RGB_TO_COLOR_NAME:
        return _RGB_TO_COLOR_NAME[hex6]
    # 근사 매칭: 가장 가까운 색상
    r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
    best_name = hex_rgb
    best_dist = float('inf')
    for ref_hex, name in _RGB_TO_COLOR_NAME.items():
        rr, gg, bb = int(ref_hex[0:2], 16), int(ref_hex[2:4], 16), int(ref_hex[4:6], 16)
        dist = (r - rr) ** 2 + (g - gg) ** 2 + (b - bb) ** 2
        if dist < best_dist:
            best_dist = dist
            best_name = name
    return best_name


def main():
    if len(sys.argv) < 3:
        print("Usage: python parse_ooxml.py <xlsx_path> <output_base_dir> [--sheet <name>]")
        print("Example: python parse_ooxml.py ../../7_System/PK_변신.xlsx output/PK_변신/ --sheet 변신")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_base = sys.argv[2]
    target_sheet = None
    if '--sheet' in sys.argv:
        idx = sys.argv.index('--sheet')
        target_sheet = sys.argv[idx + 1]

    if not os.path.exists(xlsx_path):
        print(f"ERROR: {xlsx_path} not found")
        sys.exit(1)

    if target_sheet:
        sname = target_sheet
        for ch in '/\\:*?"<>|':
            sname = sname.replace(ch, '_')
        sheet_dir = os.path.join(output_base, sname)
        if not os.path.isdir(sheet_dir):
            print(f"ERROR: {sheet_dir} not found")
            sys.exit(1)
        process_sheet(xlsx_path, sheet_dir, target_sheet)
    else:
        # 모든 시트 처리
        for entry in sorted(os.listdir(output_base)):
            sheet_dir = os.path.join(output_base, entry)
            if not os.path.isdir(sheet_dir) or entry.startswith('_'):
                continue
            vision_dir = os.path.join(sheet_dir, '_vision_output')
            if os.path.isdir(vision_dir):
                # tile_manifest에서 실제 시트 이름 가져오기
                tm_path = os.path.join(sheet_dir, '_vision_input', 'tile_manifest.json')
                real_name = entry
                if os.path.exists(tm_path):
                    with open(tm_path, 'r', encoding='utf-8') as f:
                        tm = json.load(f)
                    real_name = tm.get('sheet_name', entry)
                process_sheet(xlsx_path, sheet_dir, real_name)


if __name__ == '__main__':
    main()
